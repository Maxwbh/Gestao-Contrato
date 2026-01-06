"""
Serviço de Relatórios Financeiros

Desenvolvedor: Maxwell da Silva Oliveira
Email: maxwbh@gmail.com
Empresa: M&S do Brasil LTDA

Este módulo implementa os relatórios de prestações a pagar e pagas,
com suporte a filtros, totalizadores e exportação.
"""
import logging
from datetime import date, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field
from enum import Enum

from django.db.models import Sum, Count, Q, F
from django.utils import timezone

logger = logging.getLogger(__name__)


class TipoRelatorio(Enum):
    """Tipos de relatório disponíveis"""
    PRESTACOES_A_PAGAR = 'prestacoes_a_pagar'
    PRESTACOES_PAGAS = 'prestacoes_pagas'
    POSICAO_CONTRATOS = 'posicao_contratos'
    PREVISAO_REAJUSTES = 'previsao_reajustes'


class StatusParcela(Enum):
    """Status para filtro de parcelas"""
    TODAS = 'todas'
    VENCIDAS = 'vencidas'
    A_VENCER = 'a_vencer'
    NO_PRAZO = 'no_prazo'


class TipoPrestacaoFiltro(Enum):
    """Tipos de prestação para filtro"""
    TODAS = 'todas'
    NORMAL = 'normal'
    INTERMEDIARIA = 'intermediaria'


@dataclass
class FiltroRelatorio:
    """Filtros para geração de relatórios"""
    contrato_id: Optional[int] = None
    imobiliaria_id: Optional[int] = None
    comprador_id: Optional[int] = None
    data_inicio: Optional[date] = None
    data_fim: Optional[date] = None
    status: StatusParcela = StatusParcela.TODAS
    tipo_prestacao: TipoPrestacaoFiltro = TipoPrestacaoFiltro.TODAS
    forma_pagamento: Optional[str] = None
    ciclo_reajuste: Optional[int] = None


@dataclass
class TotalizadorPrestacoes:
    """Totalizadores para relatórios de prestações"""
    total_parcelas: int = 0
    valor_total: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_principal: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_juros: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_multa: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_desconto: Decimal = field(default_factory=lambda: Decimal('0.00'))
    parcelas_vencidas: int = 0
    parcelas_a_vencer: int = 0
    valor_vencido: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_a_vencer: Decimal = field(default_factory=lambda: Decimal('0.00'))


@dataclass
class ItemRelatorioPrestacao:
    """Item individual do relatório de prestações"""
    contrato_numero: str
    comprador_nome: str
    comprador_documento: str
    numero_parcela: int
    tipo_parcela: str
    data_vencimento: date
    valor_original: Decimal
    valor_atual: Decimal
    valor_juros: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_multa: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_desconto: Decimal = field(default_factory=lambda: Decimal('0.00'))
    valor_total: Decimal = field(default_factory=lambda: Decimal('0.00'))
    dias_atraso: int = 0
    status_boleto: str = ''
    data_pagamento: Optional[date] = None
    valor_pago: Optional[Decimal] = None
    forma_pagamento: str = ''
    ciclo_reajuste: int = 1


class RelatorioService:
    """
    Serviço para geração de relatórios financeiros.

    Implementa relatórios de:
    - Prestações a pagar
    - Prestações pagas
    - Posição de contratos
    - Previsão de reajustes
    """

    def gerar_relatorio_prestacoes_a_pagar(
        self,
        filtro: FiltroRelatorio
    ) -> Dict[str, Any]:
        """
        Gera relatório de prestações a pagar.

        Args:
            filtro: Filtros para o relatório

        Returns:
            dict: Relatório com itens e totalizadores
        """
        from financeiro.models import Parcela, TipoParcela

        logger.info(f"Gerando relatório de prestações a pagar com filtros: {filtro}")

        # Query base: parcelas não pagas
        queryset = Parcela.objects.filter(pago=False).select_related(
            'contrato',
            'contrato__comprador',
            'contrato__imobiliaria'
        )

        # Aplicar filtros
        queryset = self._aplicar_filtros_parcela(queryset, filtro, pago=False)

        # Ordenar por data de vencimento
        queryset = queryset.order_by('data_vencimento', 'contrato__numero_contrato')

        # Construir itens do relatório
        hoje = date.today()
        itens = []
        totalizador = TotalizadorPrestacoes()

        for parcela in queryset:
            # Calcular juros e multa atualizados
            juros, multa = parcela.calcular_juros_multa(hoje)
            dias_atraso = (hoje - parcela.data_vencimento).days if hoje > parcela.data_vencimento else 0
            valor_total = parcela.valor_atual + juros + multa - parcela.valor_desconto

            item = ItemRelatorioPrestacao(
                contrato_numero=parcela.contrato.numero_contrato,
                comprador_nome=parcela.contrato.comprador.nome,
                comprador_documento=parcela.contrato.comprador.documento,
                numero_parcela=parcela.numero_parcela,
                tipo_parcela=parcela.tipo_parcela,
                data_vencimento=parcela.data_vencimento,
                valor_original=parcela.valor_original,
                valor_atual=parcela.valor_atual,
                valor_juros=juros,
                valor_multa=multa,
                valor_desconto=parcela.valor_desconto,
                valor_total=valor_total,
                dias_atraso=dias_atraso,
                status_boleto=parcela.status_boleto,
                ciclo_reajuste=parcela.ciclo_reajuste,
            )
            itens.append(item)

            # Atualizar totalizadores
            totalizador.total_parcelas += 1
            totalizador.valor_total += valor_total
            totalizador.valor_principal += parcela.valor_atual
            totalizador.valor_juros += juros
            totalizador.valor_multa += multa
            totalizador.valor_desconto += parcela.valor_desconto

            if dias_atraso > 0:
                totalizador.parcelas_vencidas += 1
                totalizador.valor_vencido += valor_total
            else:
                totalizador.parcelas_a_vencer += 1
                totalizador.valor_a_vencer += valor_total

        logger.info(f"Relatório gerado com {len(itens)} itens")

        return {
            'tipo': TipoRelatorio.PRESTACOES_A_PAGAR.value,
            'data_geracao': timezone.now(),
            'filtros_aplicados': self._filtros_para_dict(filtro),
            'itens': itens,
            'totalizador': totalizador,
            'quantidade_itens': len(itens),
        }

    def gerar_relatorio_prestacoes_pagas(
        self,
        filtro: FiltroRelatorio
    ) -> Dict[str, Any]:
        """
        Gera relatório de prestações pagas.

        Args:
            filtro: Filtros para o relatório

        Returns:
            dict: Relatório com itens e totalizadores
        """
        from financeiro.models import Parcela

        logger.info(f"Gerando relatório de prestações pagas com filtros: {filtro}")

        # Query base: parcelas pagas
        queryset = Parcela.objects.filter(pago=True).select_related(
            'contrato',
            'contrato__comprador',
            'contrato__imobiliaria'
        )

        # Aplicar filtros
        queryset = self._aplicar_filtros_parcela(queryset, filtro, pago=True)

        # Ordenar por data de pagamento (mais recente primeiro)
        queryset = queryset.order_by('-data_pagamento', 'contrato__numero_contrato')

        # Construir itens do relatório
        itens = []
        totalizador = TotalizadorPrestacoes()

        for parcela in queryset:
            item = ItemRelatorioPrestacao(
                contrato_numero=parcela.contrato.numero_contrato,
                comprador_nome=parcela.contrato.comprador.nome,
                comprador_documento=parcela.contrato.comprador.documento,
                numero_parcela=parcela.numero_parcela,
                tipo_parcela=parcela.tipo_parcela,
                data_vencimento=parcela.data_vencimento,
                valor_original=parcela.valor_original,
                valor_atual=parcela.valor_atual,
                valor_juros=parcela.valor_juros,
                valor_multa=parcela.valor_multa,
                valor_desconto=parcela.valor_desconto,
                valor_total=parcela.valor_pago or Decimal('0.00'),
                dias_atraso=0,
                status_boleto=parcela.status_boleto,
                data_pagamento=parcela.data_pagamento,
                valor_pago=parcela.valor_pago,
                ciclo_reajuste=parcela.ciclo_reajuste,
            )
            itens.append(item)

            # Atualizar totalizadores
            totalizador.total_parcelas += 1
            totalizador.valor_total += parcela.valor_pago or Decimal('0.00')
            totalizador.valor_principal += parcela.valor_atual
            totalizador.valor_juros += parcela.valor_juros
            totalizador.valor_multa += parcela.valor_multa
            totalizador.valor_desconto += parcela.valor_desconto

        logger.info(f"Relatório gerado com {len(itens)} itens")

        return {
            'tipo': TipoRelatorio.PRESTACOES_PAGAS.value,
            'data_geracao': timezone.now(),
            'filtros_aplicados': self._filtros_para_dict(filtro),
            'itens': itens,
            'totalizador': totalizador,
            'quantidade_itens': len(itens),
        }

    def gerar_relatorio_posicao_contratos(
        self,
        filtro: FiltroRelatorio
    ) -> Dict[str, Any]:
        """
        Gera relatório de posição de contratos.

        Args:
            filtro: Filtros para o relatório

        Returns:
            dict: Relatório com posição de cada contrato
        """
        from contratos.models import Contrato, StatusContrato

        logger.info(f"Gerando relatório de posição de contratos")

        queryset = Contrato.objects.filter(
            status=StatusContrato.ATIVO
        ).select_related('comprador', 'imobiliaria', 'imovel')

        # Aplicar filtros
        if filtro.contrato_id:
            queryset = queryset.filter(id=filtro.contrato_id)
        if filtro.imobiliaria_id:
            queryset = queryset.filter(imobiliaria_id=filtro.imobiliaria_id)
        if filtro.comprador_id:
            queryset = queryset.filter(comprador_id=filtro.comprador_id)

        itens = []
        for contrato in queryset:
            resumo = contrato.get_resumo_financeiro()
            proxima_parcela = contrato.get_parcelas_a_pagar().first()

            itens.append({
                'contrato_numero': contrato.numero_contrato,
                'comprador_nome': contrato.comprador.nome,
                'imovel': str(contrato.imovel),
                'data_contrato': contrato.data_contrato,
                'valor_total': contrato.valor_total,
                'valor_entrada': contrato.valor_entrada,
                'valor_financiado': contrato.valor_financiado,
                'total_parcelas': resumo['total_parcelas'],
                'parcelas_pagas': resumo['parcelas_pagas'],
                'parcelas_a_pagar': resumo['parcelas_a_pagar'],
                'parcelas_vencidas': resumo['parcelas_vencidas'],
                'total_pago': resumo['total_pago'],
                'saldo_devedor': resumo['saldo_devedor'],
                'progresso_percentual': resumo['progresso_percentual'],
                'proxima_parcela_vencimento': proxima_parcela.data_vencimento if proxima_parcela else None,
                'proxima_parcela_valor': proxima_parcela.valor_atual if proxima_parcela else None,
                'ciclo_atual': resumo['ciclo_atual'],
                'bloqueio_reajuste': resumo['bloqueio_reajuste'],
                'tipo_correcao': contrato.tipo_correcao,
                'data_proximo_reajuste': contrato.data_proximo_reajuste,
            })

        # Calcular totalizadores
        total_valor_contratos = sum(i['valor_total'] for i in itens)
        total_pago = sum(i['total_pago'] for i in itens)
        total_saldo = sum(i['saldo_devedor'] for i in itens)

        return {
            'tipo': TipoRelatorio.POSICAO_CONTRATOS.value,
            'data_geracao': timezone.now(),
            'filtros_aplicados': self._filtros_para_dict(filtro),
            'itens': itens,
            'totalizadores': {
                'total_contratos': len(itens),
                'valor_total_contratos': total_valor_contratos,
                'total_pago': total_pago,
                'total_saldo_devedor': total_saldo,
            },
            'quantidade_itens': len(itens),
        }

    def gerar_relatorio_previsao_reajustes(
        self,
        dias_antecedencia: int = 30
    ) -> Dict[str, Any]:
        """
        Gera relatório de contratos com reajuste próximo.

        Args:
            dias_antecedencia: Dias de antecedência para alertar

        Returns:
            dict: Relatório com contratos que precisam de reajuste
        """
        from contratos.models import Contrato, StatusContrato, TipoCorrecao
        from financeiro.models import Reajuste

        logger.info(f"Gerando relatório de previsão de reajustes")

        hoje = date.today()
        data_limite = hoje + timedelta(days=dias_antecedencia)

        # Buscar contratos ativos que não usam valor fixo
        contratos = Contrato.objects.filter(
            status=StatusContrato.ATIVO
        ).exclude(
            tipo_correcao=TipoCorrecao.FIXO
        ).select_related('comprador', 'imobiliaria')

        itens = []
        for contrato in contratos:
            data_proximo = contrato.data_proximo_reajuste
            if not data_proximo:
                continue

            # Verificar se está dentro do período de alerta
            if data_proximo <= data_limite:
                # Verificar se reajuste já foi aplicado
                proximo_ciclo = contrato.ciclo_reajuste_atual + 1
                reajuste_existente = Reajuste.objects.filter(
                    contrato=contrato,
                    ciclo=proximo_ciclo
                ).first()

                status_reajuste = 'PENDENTE'
                if reajuste_existente:
                    status_reajuste = 'APLICADO' if reajuste_existente.aplicado else 'CRIADO'

                # Calcular parcelas afetadas
                parcela_inicial = (proximo_ciclo - 1) * contrato.prazo_reajuste_meses + 1
                parcela_final = min(proximo_ciclo * contrato.prazo_reajuste_meses, contrato.numero_parcelas)
                parcelas_afetadas = contrato.parcelas.filter(
                    numero_parcela__gte=parcela_inicial,
                    numero_parcela__lte=parcela_final,
                    pago=False
                ).count()

                itens.append({
                    'contrato_numero': contrato.numero_contrato,
                    'comprador_nome': contrato.comprador.nome,
                    'tipo_correcao': contrato.tipo_correcao,
                    'data_proximo_reajuste': data_proximo,
                    'dias_para_reajuste': (data_proximo - hoje).days,
                    'ciclo_proximo': proximo_ciclo,
                    'status_reajuste': status_reajuste,
                    'parcelas_afetadas': parcelas_afetadas,
                    'bloqueio_ativo': contrato.bloqueio_boleto_reajuste,
                    'ultimo_reajuste': contrato.data_ultimo_reajuste,
                })

        # Ordenar por data de reajuste
        itens.sort(key=lambda x: x['data_proximo_reajuste'])

        return {
            'tipo': TipoRelatorio.PREVISAO_REAJUSTES.value,
            'data_geracao': timezone.now(),
            'dias_antecedencia': dias_antecedencia,
            'itens': itens,
            'totalizadores': {
                'total_contratos': len(itens),
                'contratos_pendentes': sum(1 for i in itens if i['status_reajuste'] == 'PENDENTE'),
                'contratos_bloqueados': sum(1 for i in itens if i['bloqueio_ativo']),
            },
            'quantidade_itens': len(itens),
        }

    def _aplicar_filtros_parcela(
        self,
        queryset,
        filtro: FiltroRelatorio,
        pago: bool
    ):
        """Aplica filtros comuns ao queryset de parcelas"""
        from financeiro.models import TipoParcela

        if filtro.contrato_id:
            queryset = queryset.filter(contrato_id=filtro.contrato_id)

        if filtro.imobiliaria_id:
            queryset = queryset.filter(contrato__imobiliaria_id=filtro.imobiliaria_id)

        if filtro.comprador_id:
            queryset = queryset.filter(contrato__comprador_id=filtro.comprador_id)

        # Filtro de data
        if pago:
            # Para pagas, filtrar por data de pagamento
            if filtro.data_inicio:
                queryset = queryset.filter(data_pagamento__gte=filtro.data_inicio)
            if filtro.data_fim:
                queryset = queryset.filter(data_pagamento__lte=filtro.data_fim)
        else:
            # Para a pagar, filtrar por data de vencimento
            if filtro.data_inicio:
                queryset = queryset.filter(data_vencimento__gte=filtro.data_inicio)
            if filtro.data_fim:
                queryset = queryset.filter(data_vencimento__lte=filtro.data_fim)

        # Filtro de status (apenas para a pagar)
        if not pago and filtro.status != StatusParcela.TODAS:
            hoje = date.today()
            if filtro.status == StatusParcela.VENCIDAS:
                queryset = queryset.filter(data_vencimento__lt=hoje)
            elif filtro.status == StatusParcela.A_VENCER:
                queryset = queryset.filter(data_vencimento__gte=hoje)

        # Filtro de tipo de prestação
        if filtro.tipo_prestacao != TipoPrestacaoFiltro.TODAS:
            if filtro.tipo_prestacao == TipoPrestacaoFiltro.NORMAL:
                queryset = queryset.filter(tipo_parcela=TipoParcela.NORMAL)
            elif filtro.tipo_prestacao == TipoPrestacaoFiltro.INTERMEDIARIA:
                queryset = queryset.filter(tipo_parcela=TipoParcela.INTERMEDIARIA)

        # Filtro de ciclo de reajuste
        if filtro.ciclo_reajuste:
            queryset = queryset.filter(ciclo_reajuste=filtro.ciclo_reajuste)

        return queryset

    def _filtros_para_dict(self, filtro: FiltroRelatorio) -> Dict:
        """Converte filtros para dicionário para registro"""
        return {
            'contrato_id': filtro.contrato_id,
            'imobiliaria_id': filtro.imobiliaria_id,
            'comprador_id': filtro.comprador_id,
            'data_inicio': str(filtro.data_inicio) if filtro.data_inicio else None,
            'data_fim': str(filtro.data_fim) if filtro.data_fim else None,
            'status': filtro.status.value,
            'tipo_prestacao': filtro.tipo_prestacao.value,
            'ciclo_reajuste': filtro.ciclo_reajuste,
        }

    def exportar_para_csv(self, relatorio: Dict[str, Any]) -> str:
        """
        Exporta relatório para formato CSV.

        Args:
            relatorio: Relatório gerado

        Returns:
            str: Conteúdo CSV
        """
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)

        tipo = relatorio.get('tipo')

        if tipo in [TipoRelatorio.PRESTACOES_A_PAGAR.value, TipoRelatorio.PRESTACOES_PAGAS.value]:
            # Cabeçalho
            cabecalho = [
                'Contrato', 'Comprador', 'Documento', 'Parcela', 'Tipo',
                'Vencimento', 'Valor Original', 'Valor Atual', 'Juros',
                'Multa', 'Desconto', 'Valor Total', 'Dias Atraso', 'Status Boleto'
            ]
            if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                cabecalho.extend(['Data Pagamento', 'Valor Pago'])

            writer.writerow(cabecalho)

            # Itens
            for item in relatorio.get('itens', []):
                linha = [
                    item.contrato_numero,
                    item.comprador_nome,
                    item.comprador_documento,
                    item.numero_parcela,
                    item.tipo_parcela,
                    item.data_vencimento.strftime('%d/%m/%Y'),
                    f'{item.valor_original:.2f}',
                    f'{item.valor_atual:.2f}',
                    f'{item.valor_juros:.2f}',
                    f'{item.valor_multa:.2f}',
                    f'{item.valor_desconto:.2f}',
                    f'{item.valor_total:.2f}',
                    item.dias_atraso,
                    item.status_boleto,
                ]
                if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                    linha.extend([
                        item.data_pagamento.strftime('%d/%m/%Y') if item.data_pagamento else '',
                        f'{item.valor_pago:.2f}' if item.valor_pago else '',
                    ])
                writer.writerow(linha)

        return output.getvalue()

    def exportar_para_json(self, relatorio: Dict[str, Any]) -> str:
        """
        Exporta relatório para formato JSON.

        Args:
            relatorio: Relatório gerado

        Returns:
            str: Conteúdo JSON
        """
        import json
        from datetime import datetime

        def converter(obj):
            if isinstance(obj, (date, datetime)):
                return obj.isoformat()
            if isinstance(obj, Decimal):
                return float(obj)
            if hasattr(obj, '__dict__'):
                return obj.__dict__
            return str(obj)

        return json.dumps(relatorio, default=converter, indent=2, ensure_ascii=False)

    def exportar_para_excel(self, relatorio: Dict[str, Any]) -> bytes:
        """
        Exporta relatório para formato Excel (XLSX).

        Args:
            relatorio: Relatório gerado

        Returns:
            bytes: Conteúdo do arquivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from io import BytesIO
        except ImportError:
            logger.error("openpyxl não instalado. Execute: pip install openpyxl")
            raise ImportError("Biblioteca openpyxl é necessária para exportar Excel")

        wb = Workbook()
        ws = wb.active

        tipo = relatorio.get('tipo')
        ws.title = tipo.replace('_', ' ').title()[:31]

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '#,##0.00'
        date_format = 'DD/MM/YYYY'

        row_num = 1

        # Título
        ws.merge_cells(f'A{row_num}:N{row_num}')
        ws[f'A{row_num}'] = f"Relatório: {tipo.replace('_', ' ').title()}"
        ws[f'A{row_num}'].font = Font(bold=True, size=14)
        row_num += 1

        # Data de geração
        data_geracao = relatorio.get('data_geracao')
        if data_geracao:
            ws[f'A{row_num}'] = f"Gerado em: {data_geracao.strftime('%d/%m/%Y %H:%M')}"
            row_num += 2

        # Cabeçalhos e dados conforme tipo
        if tipo in [TipoRelatorio.PRESTACOES_A_PAGAR.value, TipoRelatorio.PRESTACOES_PAGAS.value]:
            headers = [
                'Contrato', 'Comprador', 'Documento', 'Parcela', 'Tipo',
                'Vencimento', 'Valor Original', 'Valor Atual', 'Juros',
                'Multa', 'Desconto', 'Valor Total', 'Dias Atraso', 'Status Boleto'
            ]
            if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                headers.extend(['Data Pagamento', 'Valor Pago'])

            # Escrever cabeçalho
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            row_num += 1

            # Escrever dados
            for item in relatorio.get('itens', []):
                data_row = [
                    item.contrato_numero,
                    item.comprador_nome,
                    item.comprador_documento,
                    item.numero_parcela,
                    item.tipo_parcela,
                    item.data_vencimento,
                    float(item.valor_original),
                    float(item.valor_atual),
                    float(item.valor_juros),
                    float(item.valor_multa),
                    float(item.valor_desconto),
                    float(item.valor_total),
                    item.dias_atraso,
                    item.status_boleto,
                ]
                if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                    data_row.extend([
                        item.data_pagamento,
                        float(item.valor_pago) if item.valor_pago else 0,
                    ])

                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border

                    # Formato de moeda para colunas de valor
                    if col in [7, 8, 9, 10, 11, 12, 16]:
                        cell.number_format = currency_format
                    # Formato de data
                    if col in [6, 15]:
                        cell.number_format = date_format

                row_num += 1

            # Totalizadores
            row_num += 1
            totalizador = relatorio.get('totalizador')
            if totalizador:
                ws[f'A{row_num}'] = 'TOTALIZADORES'
                ws[f'A{row_num}'].font = Font(bold=True)
                row_num += 1

                ws[f'A{row_num}'] = 'Total de Parcelas:'
                ws[f'B{row_num}'] = totalizador.total_parcelas
                row_num += 1

                ws[f'A{row_num}'] = 'Valor Total:'
                ws[f'B{row_num}'] = float(totalizador.valor_total)
                ws[f'B{row_num}'].number_format = currency_format
                row_num += 1

                if tipo == TipoRelatorio.PRESTACOES_A_PAGAR.value:
                    ws[f'A{row_num}'] = 'Parcelas Vencidas:'
                    ws[f'B{row_num}'] = totalizador.parcelas_vencidas
                    row_num += 1

                    ws[f'A{row_num}'] = 'Valor Vencido:'
                    ws[f'B{row_num}'] = float(totalizador.valor_vencido)
                    ws[f'B{row_num}'].number_format = currency_format

        elif tipo == TipoRelatorio.POSICAO_CONTRATOS.value:
            headers = [
                'Contrato', 'Comprador', 'Imóvel', 'Data Contrato',
                'Valor Total', 'Entrada', 'Financiado', 'Total Parcelas',
                'Pagas', 'A Pagar', 'Vencidas', 'Total Pago',
                'Saldo Devedor', 'Progresso %', 'Próximo Vencimento'
            ]

            # Escrever cabeçalho
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            row_num += 1

            # Escrever dados
            for item in relatorio.get('itens', []):
                data_row = [
                    item['contrato_numero'],
                    item['comprador_nome'],
                    item['imovel'],
                    item['data_contrato'],
                    float(item['valor_total']),
                    float(item['valor_entrada']),
                    float(item['valor_financiado']),
                    item['total_parcelas'],
                    item['parcelas_pagas'],
                    item['parcelas_a_pagar'],
                    item['parcelas_vencidas'],
                    float(item['total_pago']),
                    float(item['saldo_devedor']),
                    round(item['progresso_percentual'], 2),
                    item['proxima_parcela_vencimento'],
                ]

                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border

                    # Formato de moeda
                    if col in [5, 6, 7, 12, 13]:
                        cell.number_format = currency_format
                    # Formato de data
                    if col in [4, 15]:
                        cell.number_format = date_format
                    # Formato de percentual
                    if col == 14:
                        cell.number_format = '0.00%'

                row_num += 1

        elif tipo == TipoRelatorio.PREVISAO_REAJUSTES.value:
            headers = [
                'Contrato', 'Comprador', 'Índice', 'Próximo Reajuste',
                'Dias Restantes', 'Ciclo', 'Status', 'Parcelas Afetadas',
                'Bloqueado', 'Último Reajuste'
            ]

            # Escrever cabeçalho
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            row_num += 1

            # Escrever dados
            for item in relatorio.get('itens', []):
                data_row = [
                    item['contrato_numero'],
                    item['comprador_nome'],
                    item['tipo_correcao'],
                    item['data_proximo_reajuste'],
                    item['dias_para_reajuste'],
                    item['ciclo_proximo'],
                    item['status_reajuste'],
                    item['parcelas_afetadas'],
                    'Sim' if item['bloqueio_ativo'] else 'Não',
                    item['ultimo_reajuste'],
                ]

                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border

                    if col in [4, 10]:
                        cell.number_format = date_format

                row_num += 1

        # Ajustar largura das colunas
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def exportar_para_pdf(self, relatorio: Dict[str, Any]) -> bytes:
        """
        Exporta relatório para formato PDF.

        Args:
            relatorio: Relatório gerado

        Returns:
            bytes: Conteúdo do arquivo PDF
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from io import BytesIO
        except ImportError:
            logger.error("reportlab não instalado. Execute: pip install reportlab")
            raise ImportError("Biblioteca reportlab é necessária para exportar PDF")

        output = BytesIO()
        tipo = relatorio.get('tipo')

        # Usar paisagem para tabelas largas
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )

        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=12
        )
        title = Paragraph(f"Relatório: {tipo.replace('_', ' ').title()}", title_style)
        elements.append(title)

        # Data de geração
        data_geracao = relatorio.get('data_geracao')
        if data_geracao:
            date_text = Paragraph(
                f"Gerado em: {data_geracao.strftime('%d/%m/%Y %H:%M')}",
                styles['Normal']
            )
            elements.append(date_text)
            elements.append(Spacer(1, 10*mm))

        # Construir tabela conforme tipo
        if tipo in [TipoRelatorio.PRESTACOES_A_PAGAR.value, TipoRelatorio.PRESTACOES_PAGAS.value]:
            headers = [
                'Contrato', 'Comprador', 'Parcela', 'Tipo',
                'Vencimento', 'Valor Atual', 'Juros', 'Multa',
                'Valor Total', 'Status'
            ]
            if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                headers.extend(['Dt Pgto', 'Valor Pago'])

            data = [headers]

            for item in relatorio.get('itens', []):
                row = [
                    item.contrato_numero[:15],
                    item.comprador_nome[:20],
                    str(item.numero_parcela),
                    item.tipo_parcela[:10],
                    item.data_vencimento.strftime('%d/%m/%Y'),
                    f'R$ {item.valor_atual:,.2f}',
                    f'R$ {item.valor_juros:,.2f}',
                    f'R$ {item.valor_multa:,.2f}',
                    f'R$ {item.valor_total:,.2f}',
                    item.status_boleto[:10],
                ]
                if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                    row.extend([
                        item.data_pagamento.strftime('%d/%m/%Y') if item.data_pagamento else '-',
                        f'R$ {item.valor_pago:,.2f}' if item.valor_pago else '-',
                    ])
                data.append(row)

            # Linha de totais
            totalizador = relatorio.get('totalizador')
            if totalizador:
                total_row = [
                    'TOTAL', '', str(totalizador.total_parcelas), '',
                    '', f'R$ {totalizador.valor_principal:,.2f}',
                    f'R$ {totalizador.valor_juros:,.2f}',
                    f'R$ {totalizador.valor_multa:,.2f}',
                    f'R$ {totalizador.valor_total:,.2f}', ''
                ]
                if tipo == TipoRelatorio.PRESTACOES_PAGAS.value:
                    total_row.extend(['', ''])
                data.append(total_row)

        elif tipo == TipoRelatorio.POSICAO_CONTRATOS.value:
            headers = [
                'Contrato', 'Comprador', 'Valor Total', 'Pago',
                'Saldo', 'Parcelas', 'Pagas', 'Progresso'
            ]
            data = [headers]

            for item in relatorio.get('itens', []):
                row = [
                    item['contrato_numero'][:15],
                    item['comprador_nome'][:25],
                    f"R$ {float(item['valor_total']):,.2f}",
                    f"R$ {float(item['total_pago']):,.2f}",
                    f"R$ {float(item['saldo_devedor']):,.2f}",
                    str(item['total_parcelas']),
                    str(item['parcelas_pagas']),
                    f"{item['progresso_percentual']:.1f}%",
                ]
                data.append(row)

        elif tipo == TipoRelatorio.PREVISAO_REAJUSTES.value:
            headers = [
                'Contrato', 'Comprador', 'Índice', 'Próx. Reajuste',
                'Dias', 'Ciclo', 'Status', 'Bloqueado'
            ]
            data = [headers]

            for item in relatorio.get('itens', []):
                row = [
                    item['contrato_numero'][:15],
                    item['comprador_nome'][:25],
                    item['tipo_correcao'],
                    item['data_proximo_reajuste'].strftime('%d/%m/%Y') if item['data_proximo_reajuste'] else '-',
                    str(item['dias_para_reajuste']),
                    str(item['ciclo_proximo']),
                    item['status_reajuste'],
                    'Sim' if item['bloqueio_ativo'] else 'Não',
                ]
                data.append(row)

        else:
            data = [['Relatório não suportado para exportação PDF']]

        # Criar tabela
        col_widths = None
        if len(data) > 0 and len(data[0]) > 0:
            available_width = landscape(A4)[0] - 20*mm
            col_widths = [available_width / len(data[0])] * len(data[0])

        table = Table(data, colWidths=col_widths)

        # Estilo da tabela
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F0FE')]),
        ])

        # Destacar última linha (totais) se houver
        if len(data) > 1:
            style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D6DCE4'))
            style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')

        table.setStyle(style)
        elements.append(table)

        # Construir PDF
        doc.build(elements)
        output.seek(0)

        return output.getvalue()

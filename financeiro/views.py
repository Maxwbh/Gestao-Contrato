"""
Views do app Financeiro

Desenvolvedor: Maxwell da Silva Oliveira
Email: maxwbh@gmail.com
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, FileResponse
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Count, Q, Min
from django.views.generic import TemplateView
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator
from django.views.decorators.clickjacking import xframe_options_sameorigin
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from decimal import Decimal
import logging

from django.core.cache import cache
from .models import Parcela, Reajuste, StatusBoleto, HistoricoPagamento
from core.models import Imobiliaria, ContaBancaria, BancoBrasil
from contratos.models import Contrato, StatusContrato

logger = logging.getLogger(__name__)


def _voltar_url(request, default):
    """Retorna o HTTP_REFERER se for da mesma origem, senão o default."""
    from urllib.parse import urlparse
    ref = request.META.get('HTTP_REFERER', '')
    if ref:
        p = urlparse(ref)
        if not p.netloc or p.netloc == request.get_host():
            return ref
    return default


class DashboardFinanceiroView(LoginRequiredMixin, TemplateView):
    """Dashboard Financeiro por Imobiliária"""
    template_name = 'financeiro/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoje = timezone.now().date()

        # Filtro por imobiliária
        imobiliaria_id = self.request.GET.get('imobiliaria')
        imobiliarias = Imobiliaria.objects.filter(ativo=True)
        imobiliaria_selecionada = None

        if imobiliaria_id:
            try:
                imobiliaria_selecionada = Imobiliaria.objects.get(pk=imobiliaria_id, ativo=True)
            except Imobiliaria.DoesNotExist:
                pass

        # Base queryset de contratos
        contratos_qs = Contrato.objects.all()
        if imobiliaria_selecionada:
            contratos_qs = contratos_qs.filter(imobiliaria=imobiliaria_selecionada)

        # Base queryset de parcelas
        parcelas_qs = Parcela.objects.all()
        if imobiliaria_selecionada:
            parcelas_qs = parcelas_qs.filter(contrato__imobiliaria=imobiliaria_selecionada)

        # K-01: Lotes
        from core.models import Imovel
        imoveis_qs = Imovel.objects.all()
        if imobiliaria_selecionada:
            imoveis_qs = imoveis_qs.filter(imobiliaria=imobiliaria_selecionada)
        context['total_lotes'] = imoveis_qs.count()
        context['lotes_disponiveis'] = imoveis_qs.filter(disponivel=True).count()
        context['lotes_vendidos'] = imoveis_qs.filter(disponivel=False).count()

        # K-06: Reajustes pendentes
        contratos_nao_fixo = contratos_qs.filter(
            status=StatusContrato.ATIVO
        ).exclude(tipo_correcao='FIXO').only(
            'tipo_correcao', 'prazo_reajuste_meses', 'data_contrato', 'ciclo_reajuste_atual'
        )
        reajustes_pendentes = sum(
            1 for c in contratos_nao_fixo
            if Reajuste.calcular_ciclo_pendente(c) is not None
        )
        context['reajustes_pendentes'] = reajustes_pendentes

        # Estatísticas de Contratos
        context['total_contratos'] = contratos_qs.count()
        context['contratos_ativos'] = contratos_qs.filter(status=StatusContrato.ATIVO).count()
        context['contratos_quitados'] = contratos_qs.filter(status=StatusContrato.QUITADO).count()
        context['contratos_cancelados'] = contratos_qs.filter(status=StatusContrato.CANCELADO).count()

        # Valor total dos contratos
        context['valor_total_contratos'] = contratos_qs.aggregate(
            total=Sum('valor_total')
        )['total'] or Decimal('0.00')

        # Estatísticas de Parcelas
        context['total_parcelas'] = parcelas_qs.count()
        context['parcelas_pagas'] = parcelas_qs.filter(pago=True).count()
        context['parcelas_pendentes'] = parcelas_qs.filter(pago=False).count()
        context['parcelas_vencidas'] = parcelas_qs.filter(
            pago=False, data_vencimento__lt=hoje
        ).count()

        # Valores
        context['valor_recebido'] = parcelas_qs.filter(pago=True).aggregate(
            total=Sum('valor_pago')
        )['total'] or Decimal('0.00')

        context['valor_a_receber'] = parcelas_qs.filter(pago=False).aggregate(
            total=Sum('valor_atual')
        )['total'] or Decimal('0.00')

        context['valor_em_atraso'] = parcelas_qs.filter(
            pago=False, data_vencimento__lt=hoje
        ).aggregate(total=Sum('valor_atual'))['total'] or Decimal('0.00')

        # Períodos para filtros
        primeiro_dia_mes = hoje.replace(day=1)
        ultimo_dia_mes_passado = primeiro_dia_mes - timedelta(days=1)
        primeiro_dia_mes_passado = ultimo_dia_mes_passado.replace(day=1)
        seis_meses_atras = hoje - relativedelta(months=6)
        primeiro_dia_ano = hoje.replace(month=1, day=1)

        # Parcelas do mês atual
        context['parcelas_mes_atual'] = self._get_estatisticas_periodo(
            parcelas_qs, primeiro_dia_mes, hoje
        )

        # Parcelas do mês passado
        context['parcelas_mes_passado'] = self._get_estatisticas_periodo(
            parcelas_qs, primeiro_dia_mes_passado, ultimo_dia_mes_passado
        )

        # Parcelas últimos 6 meses
        context['parcelas_6_meses'] = self._get_estatisticas_periodo(
            parcelas_qs, seis_meses_atras, hoje
        )

        # Parcelas do ano
        context['parcelas_ano'] = self._get_estatisticas_periodo(
            parcelas_qs, primeiro_dia_ano, hoje
        )

        # Top 10 contratos com mais atraso
        context['contratos_mais_atraso'] = self._get_contratos_mais_atraso(
            contratos_qs, limite=10
        )

        # D-04: Parcelas vencendo esta semana
        fim_semana = hoje + timedelta(days=7)
        context['parcelas_semana'] = (
            parcelas_qs.filter(
                pago=False,
                data_vencimento__gte=hoje,
                data_vencimento__lte=fim_semana,
                tipo_parcela='NORMAL',
            )
            .select_related('contrato__comprador', 'contrato__imovel')
            .order_by('data_vencimento')[:20]
        )

        # G-05: Top 5 contratos com maior saldo devedor estimado (soma valor_atual parcelas NORMAL não pagas)
        context['top5_saldo_devedor'] = (
            contratos_qs.filter(status=StatusContrato.ATIVO)
            .annotate(
                saldo_est=Sum(
                    'parcelas__valor_atual',
                    filter=Q(parcelas__pago=False, parcelas__tipo_parcela='NORMAL')
                )
            )
            .filter(saldo_est__isnull=False)
            .select_related('comprador', 'imovel')
            .order_by('-saldo_est')[:5]
        )

        # Imobiliárias e selecionada
        context['imobiliarias'] = imobiliarias
        context['imobiliaria_selecionada'] = imobiliaria_selecionada

        return context

    def _get_estatisticas_periodo(self, parcelas_qs, data_inicio, data_fim):
        """Retorna estatísticas de parcelas em um período"""
        parcelas_periodo = parcelas_qs.filter(
            data_vencimento__gte=data_inicio,
            data_vencimento__lte=data_fim
        )

        return {
            'total': parcelas_periodo.count(),
            'pagas': parcelas_periodo.filter(pago=True).count(),
            'pendentes': parcelas_periodo.filter(pago=False).count(),
            'vencidas': parcelas_periodo.filter(
                pago=False, data_vencimento__lt=timezone.now().date()
            ).count(),
            'valor_esperado': parcelas_periodo.aggregate(
                total=Sum('valor_atual')
            )['total'] or Decimal('0.00'),
            'valor_recebido': parcelas_periodo.filter(pago=True).aggregate(
                total=Sum('valor_pago')
            )['total'] or Decimal('0.00'),
        }

    def _get_contratos_mais_atraso(self, contratos_qs, limite=10):
        """Retorna contratos com mais parcelas em atraso.

        Usa annotate para calcular tudo em 1 query em vez de N+1.
        """
        hoje = timezone.now().date()
        filtro_atraso = Q(parcelas__pago=False, parcelas__data_vencimento__lt=hoje)

        qs = (
            contratos_qs
            .filter(status=StatusContrato.ATIVO)
            .filter(filtro_atraso)
            .annotate(
                parcelas_atraso_count=Count('parcelas', filter=filtro_atraso),
                valor_atraso_total=Sum('parcelas__valor_atual', filter=filtro_atraso),
                primeira_vencida=Min('parcelas__data_vencimento', filter=filtro_atraso),
            )
            .filter(parcelas_atraso_count__gt=0)
            .select_related('comprador', 'imovel')
            .order_by('-valor_atraso_total')[:limite]
        )

        result = []
        for contrato in qs:
            result.append({
                'contrato': contrato,
                'parcelas_atraso': contrato.parcelas_atraso_count,
                'valor_atraso': contrato.valor_atraso_total or Decimal('0.00'),
                'dias_atraso': (hoje - contrato.primeira_vencida).days,
            })
        return result


@login_required
def api_dashboard_dados(request):
    """API para retornar dados do dashboard em JSON (para gráficos)"""
    hoje = timezone.now().date()

    # Filtro por imobiliária
    imobiliaria_id = request.GET.get('imobiliaria') or ''

    # Cache por data + imobiliária (5 minutos) — evita múltiplas queries pesadas
    cache_key = f'dashboard:dados:{hoje.isoformat()}:{imobiliaria_id}'
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse(cached)

    parcelas_qs = Parcela.objects.all()
    contratos_qs = Contrato.objects.all()

    if imobiliaria_id:
        parcelas_qs = parcelas_qs.filter(contrato__imobiliaria_id=imobiliaria_id)
        contratos_qs = contratos_qs.filter(imobiliaria_id=imobiliaria_id)

    # Dados para gráfico de pizza - Status das parcelas
    status_parcelas = {
        'labels': ['Pagas', 'Pendentes', 'Vencidas'],
        'data': [
            parcelas_qs.filter(pago=True).count(),
            parcelas_qs.filter(pago=False, data_vencimento__gte=hoje).count(),
            parcelas_qs.filter(pago=False, data_vencimento__lt=hoje).count(),
        ],
        'colors': ['#28a745', '#ffc107', '#dc3545']
    }

    # Dados para gráfico de pizza - Status dos contratos
    status_contratos = {
        'labels': ['Ativos', 'Quitados', 'Cancelados', 'Suspensos'],
        'data': [
            contratos_qs.filter(status=StatusContrato.ATIVO).count(),
            contratos_qs.filter(status=StatusContrato.QUITADO).count(),
            contratos_qs.filter(status=StatusContrato.CANCELADO).count(),
            contratos_qs.filter(status=StatusContrato.SUSPENSO).count(),
        ],
        'colors': ['#007bff', '#28a745', '#dc3545', '#6c757d']
    }

    # Dados para gráfico de barras - Recebimentos por mês (últimos 12 meses)
    recebimentos_mensais = {'labels': [], 'recebido': [], 'esperado': []}
    meses = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

    for i in range(11, -1, -1):
        data = hoje - relativedelta(months=i)
        primeiro_dia = data.replace(day=1)
        ultimo_dia = (primeiro_dia + relativedelta(months=1)) - timedelta(days=1)

        parcelas_mes = parcelas_qs.filter(
            data_vencimento__gte=primeiro_dia,
            data_vencimento__lte=ultimo_dia
        )

        recebido = parcelas_mes.filter(pago=True).aggregate(
            total=Sum('valor_pago')
        )['total'] or 0
        esperado = parcelas_mes.aggregate(
            total=Sum('valor_atual')
        )['total'] or 0

        recebimentos_mensais['labels'].append(f"{meses[data.month-1]}/{data.year % 100}")
        recebimentos_mensais['recebido'].append(float(recebido))
        recebimentos_mensais['esperado'].append(float(esperado))

    # Dados para gráfico de linha - Inadimplência por mês
    inadimplencia_mensal = {'labels': [], 'valores': [], 'quantidades': []}

    for i in range(11, -1, -1):
        data = hoje - relativedelta(months=i)
        primeiro_dia = data.replace(day=1)
        ultimo_dia = (primeiro_dia + relativedelta(months=1)) - timedelta(days=1)

        parcelas_vencidas = parcelas_qs.filter(
            pago=False,
            data_vencimento__gte=primeiro_dia,
            data_vencimento__lte=ultimo_dia,
            data_vencimento__lt=hoje
        )

        valor_inadimplente = parcelas_vencidas.aggregate(
            total=Sum('valor_atual')
        )['total'] or 0

        inadimplencia_mensal['labels'].append(f"{meses[data.month-1]}/{data.year % 100}")
        inadimplencia_mensal['valores'].append(float(valor_inadimplente))
        inadimplencia_mensal['quantidades'].append(parcelas_vencidas.count())

    # Tabela vencimentos consolidados - próximos 3 meses (3.6)
    vencimentos_proximos = []
    for i in range(3):
        data = hoje + relativedelta(months=i)
        primeiro_dia = data.replace(day=1)
        ultimo_dia = (primeiro_dia + relativedelta(months=1)) - timedelta(days=1)

        parcelas_mes = parcelas_qs.filter(
            pago=False,
            data_vencimento__gte=primeiro_dia,
            data_vencimento__lte=ultimo_dia
        )
        agg = parcelas_mes.aggregate(
            total_valor=Sum('valor_atual'),
            quantidade=Count('id')
        )
        vencimentos_proximos.append({
            'mes': f"{meses[data.month - 1]}/{data.year}",
            'quantidade': agg['quantidade'] or 0,
            'valor_total': float(agg['total_valor'] or 0),
        })

    # G-02: Inadimplência por faixa de atraso
    inadimplencia_faixas = {
        'labels': ['1–30 dias', '31–60 dias', '61–90 dias', '90+ dias'],
        'data': [
            parcelas_qs.filter(
                pago=False,
                data_vencimento__gte=hoje - timedelta(days=30),
                data_vencimento__lt=hoje
            ).count(),
            parcelas_qs.filter(
                pago=False,
                data_vencimento__gte=hoje - timedelta(days=60),
                data_vencimento__lt=hoje - timedelta(days=30)
            ).count(),
            parcelas_qs.filter(
                pago=False,
                data_vencimento__gte=hoje - timedelta(days=90),
                data_vencimento__lt=hoje - timedelta(days=60)
            ).count(),
            parcelas_qs.filter(
                pago=False,
                data_vencimento__lt=hoje - timedelta(days=90)
            ).count(),
        ],
        'colors': ['#ffc107', '#fd7e14', '#dc3545', '#6f1212'],
    }

    # G-03: Fluxo de caixa previsto vs. realizado (6 meses passados + mês atual + 6 futuros)
    fluxo_caixa = {'labels': [], 'realizado': [], 'previsto': [], 'is_future': []}
    for i in range(-5, 7):
        ref = hoje + relativedelta(months=i)
        primeiro_dia = ref.replace(day=1)
        ultimo_dia = (primeiro_dia + relativedelta(months=1)) - timedelta(days=1)

        qs_mes = parcelas_qs.filter(data_vencimento__gte=primeiro_dia, data_vencimento__lte=ultimo_dia)
        agg = qs_mes.aggregate(
            previsto=Sum('valor_atual'),
            realizado=Sum('valor_pago', filter=Q(pago=True)),
        )

        fluxo_caixa['labels'].append(f"{meses[ref.month - 1]}/{ref.year % 100}")
        fluxo_caixa['previsto'].append(float(agg['previsto'] or 0))
        fluxo_caixa['realizado'].append(float(agg['realizado'] or 0) if i <= 0 else None)
        fluxo_caixa['is_future'].append(i > 0)

    data = {
        'status_parcelas': status_parcelas,
        'status_contratos': status_contratos,
        'recebimentos_mensais': recebimentos_mensais,
        'inadimplencia_mensal': inadimplencia_mensal,
        'inadimplencia_faixas': inadimplencia_faixas,
        'vencimentos_proximos': vencimentos_proximos,
        'fluxo_caixa': fluxo_caixa,
    }
    cache.set(cache_key, data, timeout=300)  # 5 minutos
    return JsonResponse(data)


@login_required
def listar_parcelas(request):
    """
    Lista todas as parcelas com filtros avançados.

    Filtros disponíveis:
    - status: pagas, pendentes, vencidas, a_vencer
    - imobiliaria: ID da imobiliária
    - comprador: ID do comprador
    - contrato: número do contrato
    - data_inicio: data de vencimento inicial
    - data_fim: data de vencimento final
    - q: busca textual
    """
    from core.models import Comprador

    parcelas = Parcela.objects.select_related(
        'contrato',
        'contrato__comprador',
        'contrato__imovel',
        'contrato__imovel__imobiliaria'
    ).order_by('-data_vencimento')

    # Dados para os filtros
    imobiliarias = Imobiliaria.objects.filter(ativo=True).order_by('nome')
    compradores = Comprador.objects.filter(ativo=True).order_by('nome')
    contas_bancarias = ContaBancaria.objects.filter(ativo=True).select_related('imobiliaria').order_by('imobiliaria__nome', 'banco')

    # Filtro por Status de Pagamento
    status = request.GET.get('status', '')
    if status == 'pagas':
        parcelas = parcelas.filter(pago=True)
    elif status == 'pendentes':
        parcelas = parcelas.filter(pago=False)
    elif status == 'vencidas':
        parcelas = parcelas.filter(pago=False, data_vencimento__lt=timezone.now().date())
    elif status == 'a_vencer':
        parcelas = parcelas.filter(pago=False, data_vencimento__gte=timezone.now().date())

    # Filtro por Status do Boleto
    status_boleto_filtro = request.GET.get('status_boleto', '')
    if status_boleto_filtro:
        parcelas = parcelas.filter(status_boleto=status_boleto_filtro)

    # Filtro por Imobiliária
    imobiliaria_id = request.GET.get('imobiliaria', '')
    if imobiliaria_id:
        parcelas = parcelas.filter(contrato__imovel__imobiliaria_id=imobiliaria_id)

    # Filtro por Comprador
    comprador_id = request.GET.get('comprador', '')
    if comprador_id:
        parcelas = parcelas.filter(contrato__comprador_id=comprador_id)

    # Filtro por Número do Contrato ou ID
    contrato_param = request.GET.get('contrato', '').strip()
    if contrato_param:
        # Se for número, filtra por ID do contrato
        if contrato_param.isdigit():
            parcelas = parcelas.filter(contrato_id=contrato_param)
        else:
            # Senão, filtra por número do contrato
            parcelas = parcelas.filter(contrato__numero_contrato__icontains=contrato_param)

    # Filtro por Período de Vencimento
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    if data_inicio:
        parcelas = parcelas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        parcelas = parcelas.filter(data_vencimento__lte=data_fim)

    # Busca textual (nome do comprador ou identificação do imóvel)
    busca = request.GET.get('q', '').strip()
    if busca:
        parcelas = parcelas.filter(
            Q(contrato__comprador__nome__icontains=busca) |
            Q(contrato__imovel__identificacao__icontains=busca) |
            Q(contrato__numero_contrato__icontains=busca)
        )

    # Estatísticas (sobre o queryset filtrado completo)
    hoje = timezone.now().date()
    total_parcelas = parcelas.count()
    valor_total = parcelas.aggregate(total=Sum('valor_atual'))['total'] or Decimal('0.00')
    parcelas_vencidas_count = parcelas.filter(pago=False, data_vencimento__lt=hoje).count()

    # Paginação
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = min(int(per_page), 100)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(parcelas, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except Exception:
        page_obj = paginator.get_page(1)

    # Calcular juros/multa dinâmico para parcelas vencidas não pagas na página atual
    for p in page_obj:
        if not p.pago and p.data_vencimento < hoje:
            juros, multa = p.calcular_juros_multa(hoje)
            p.juros_dinamico = juros
            p.multa_dinamico = multa
            p.total_com_encargos = p.valor_atual + juros + multa
        else:
            p.juros_dinamico = None
            p.multa_dinamico = None
            p.total_com_encargos = None

    context = {
        'parcelas': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
        'imobiliarias': imobiliarias,
        'compradores': compradores,
        'contas_bancarias': contas_bancarias,
        # Valores atuais dos filtros
        'filtro_status': status,
        'filtro_status_boleto': status_boleto_filtro,
        'filtro_imobiliaria': imobiliaria_id,
        'filtro_comprador': comprador_id,
        'filtro_contrato': contrato_param,
        'filtro_data_inicio': data_inicio,
        'filtro_data_fim': data_fim,
        'filtro_busca': busca,
        # Estatísticas
        'total_parcelas': total_parcelas,
        'valor_total': valor_total,
        'parcelas_vencidas_count': parcelas_vencidas_count,
    }
    return render(request, 'financeiro/listar_parcelas.html', context)


@login_required
def detalhe_parcela(request, pk):
    """Exibe detalhes de uma parcela específica"""
    parcela = get_object_or_404(
        Parcela.objects.select_related(
            'contrato',
            'contrato__comprador',
            'contrato__imovel'
        ),
        pk=pk
    )

    # Atualizar juros e multa se estiver vencida
    if parcela.esta_vencida and not parcela.pago:
        parcela.atualizar_juros_multa()

    from django.urls import reverse
    context = {
        'parcela': parcela,
        'voltar_url': _voltar_url(request, reverse('financeiro:listar_parcelas')),
    }
    return render(request, 'financeiro/detalhe_parcela.html', context)


@login_required
@require_POST
def notificar_inadimplente(request, pk):
    """
    3.25 — Envia notificação de inadimplência manualmente para o comprador de uma parcela.
    Retorna JSON {sucesso, mensagem/erro}.
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    if parcela.pago:
        return JsonResponse({'sucesso': False, 'erro': 'Parcela já paga — notificação não aplicável.'}, status=400)

    from datetime import date as _date
    hoje = _date.today()
    if parcela.data_vencimento > hoje:
        return JsonResponse({'sucesso': False, 'erro': 'Parcela ainda não vencida.'}, status=400)

    try:
        from core.tasks import (
            _notificacao_ja_enviada_hoje,
            _registrar_notificacao,
            _enviar_pelo_canal,
        )
        from notificacoes.models import TipoNotificacao

        comprador = parcela.contrato.comprador
        dias_atraso = (hoje - parcela.data_vencimento).days
        imob_nome = getattr(parcela.contrato.imobiliaria, 'nome', 'Gestão de Contratos')

        PREFIXO = '[INADIMPLENCIA-MANUAL]'

        canais_enviados = []

        # Email
        email = getattr(comprador, 'email', None)
        if email:
            if not _notificacao_ja_enviada_hoje(parcela, PREFIXO):
                assunto = (
                    f"{PREFIXO} Parcela {parcela.numero_parcela} em atraso há {dias_atraso} dia(s)"
                )
                mensagem = (
                    f"Olá {comprador.nome},\n\n"
                    f"Verificamos que a parcela {parcela.numero_parcela}/{parcela.contrato.numero_parcelas} "
                    f"do contrato {parcela.contrato.numero_contrato} encontra-se em atraso.\n\n"
                    f"Vencimento: {parcela.data_vencimento.strftime('%d/%m/%Y')} ({dias_atraso} dia(s) em atraso)\n"
                    f"Valor: R$ {parcela.valor_atual:,.2f}\n\n"
                    f"Regularize sua situação para evitar acréscimo de juros e multa.\n\n"
                    f"Atenciosamente,\n{imob_nome}"
                )
                notif = _registrar_notificacao(parcela, TipoNotificacao.EMAIL, email, assunto, mensagem)
                enviado = _enviar_pelo_canal(TipoNotificacao.EMAIL, email, assunto, mensagem)
                if enviado and notif:
                    from notificacoes.models import StatusNotificacao
                    notif.status = StatusNotificacao.ENVIADO
                    notif.save(update_fields=['status'])
                    canais_enviados.append('email')

        # WhatsApp
        celular = getattr(comprador, 'celular', None)
        if celular:
            msg_wa = (
                f"Olá {comprador.nome}! Sua parcela {parcela.numero_parcela} do contrato "
                f"{parcela.contrato.numero_contrato} está {dias_atraso} dia(s) em atraso "
                f"(venc. {parcela.data_vencimento.strftime('%d/%m/%Y')}). "
                f"Valor: R$ {parcela.valor_atual:,.2f}. Entre em contato para regularizar."
            )
            notif_wa = _registrar_notificacao(parcela, TipoNotificacao.WHATSAPP, celular, 'Inadimplência', msg_wa)
            enviado_wa = _enviar_pelo_canal(TipoNotificacao.WHATSAPP, celular, 'Inadimplência', msg_wa)
            if enviado_wa and notif_wa:
                from notificacoes.models import StatusNotificacao
                notif_wa.status = StatusNotificacao.ENVIADO
                notif_wa.save(update_fields=['status'])
                canais_enviados.append('WhatsApp')

        if canais_enviados:
            return JsonResponse({
                'sucesso': True,
                'mensagem': f'Notificação enviada via {", ".join(canais_enviados)}.'
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Nenhum canal de notificação disponível para este comprador (sem e-mail ou celular configurado).'
            }, status=400)

    except Exception as e:
        logger.exception(f'Erro ao notificar inadimplente (parcela {pk}): {e}')
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
def registrar_pagamento(request, pk):
    """Registra o pagamento de uma parcela"""
    from datetime import datetime

    parcela = get_object_or_404(Parcela, pk=pk)

    FORMAS_PAGAMENTO = [
        ('DINHEIRO', 'Dinheiro'),
        ('PIX', 'PIX'),
        ('TRANSFERENCIA', 'Transferência Bancária'),
        ('BOLETO', 'Boleto'),
        ('CARTAO_CREDITO', 'Cartão de Crédito'),
        ('CARTAO_DEBITO', 'Cartão de Débito'),
        ('CHEQUE', 'Cheque'),
    ]

    if request.method == 'POST':
        valor_pago_str = request.POST.get('valor_pago', '0')
        data_pagamento_str = request.POST.get('data_pagamento', '')
        observacoes = request.POST.get('observacoes', '')
        forma_pagamento = request.POST.get('forma_pagamento', 'DINHEIRO')
        comprovante = request.FILES.get('comprovante')

        try:
            valor_pago = Decimal(valor_pago_str.replace(',', '.'))

            if data_pagamento_str:
                data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
            else:
                data_pagamento = timezone.now().date()

            # Calcular juros/multa para o HistoricoPagamento
            valor_juros = Decimal('0.00')
            valor_multa = Decimal('0.00')
            if data_pagamento > parcela.data_vencimento:
                valor_juros, valor_multa = parcela.calcular_juros_multa(data_pagamento)

            parcela.registrar_pagamento(
                valor_pago=valor_pago,
                data_pagamento=data_pagamento,
                observacoes=observacoes
            )

            # Criar registro no histórico com comprovante e forma de pagamento
            historico = HistoricoPagamento.objects.create(
                parcela=parcela,
                data_pagamento=data_pagamento,
                valor_pago=valor_pago,
                valor_parcela=parcela.valor_atual,
                valor_juros=valor_juros,
                valor_multa=valor_multa,
                forma_pagamento=forma_pagamento,
                observacoes=observacoes,
                origem_pagamento='MANUAL',
            )
            if comprovante:
                historico.comprovante = comprovante
                historico.save(update_fields=['comprovante'])

            messages.success(request, 'Pagamento registrado com sucesso!')
            return redirect('financeiro:detalhe_parcela', pk=pk)
        except Exception as e:
            logger.exception("Erro ao registrar pagamento parcela pk=%s: %s", pk, e)
            messages.error(request, f'Erro ao registrar pagamento: {str(e)}')

    context = {
        'parcela': parcela,
        'formas_pagamento': FORMAS_PAGAMENTO,
    }
    return render(request, 'financeiro/registrar_pagamento.html', context)


@login_required
def listar_reajustes(request):
    """Lista todos os reajustes"""
    reajustes = Reajuste.objects.select_related('contrato', 'usuario').order_by('-data_reajuste')

    per_page = request.GET.get('per_page', '25')
    try:
        per_page = min(int(per_page), 100)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(reajustes, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'reajustes': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
    }
    return render(request, 'financeiro/listar_reajustes.html', context)


# =============================================================================
# GESTÃO DE PARCELAS POR IMOBILIÁRIA
# =============================================================================

@login_required
def parcelas_mes_atual(request):
    """
    Lista parcelas que vencem no mês atual, agrupadas por imobiliária.
    """
    hoje = timezone.now().date()
    primeiro_dia_mes = hoje.replace(day=1)

    # Último dia do mês
    if hoje.month == 12:
        ultimo_dia_mes = hoje.replace(day=31)
    else:
        ultimo_dia_mes = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)

    # Filtro por imobiliária
    imobiliaria_id = request.GET.get('imobiliaria', '')
    imobiliarias = Imobiliaria.objects.filter(ativo=True).order_by('nome')

    # Base queryset
    parcelas = Parcela.objects.select_related(
        'contrato',
        'contrato__comprador',
        'contrato__imovel',
        'contrato__imovel__imobiliaria'
    ).filter(
        data_vencimento__gte=primeiro_dia_mes,
        data_vencimento__lte=ultimo_dia_mes
    ).order_by('data_vencimento', 'contrato__imovel__imobiliaria__nome')

    if imobiliaria_id:
        parcelas = parcelas.filter(contrato__imovel__imobiliaria_id=imobiliaria_id)

    # Filtro por status
    status = request.GET.get('status', '')
    if status == 'pagas':
        parcelas = parcelas.filter(pago=True)
    elif status == 'pendentes':
        parcelas = parcelas.filter(pago=False)
    elif status == 'vencidas':
        parcelas = parcelas.filter(pago=False, data_vencimento__lt=hoje)

    # Estatísticas do mês
    stats = parcelas.aggregate(
        total=Count('id'),
        valor_total=Sum('valor_atual'),
        pagas=Count('id', filter=Q(pago=True)),
        valor_pago=Sum('valor_pago', filter=Q(pago=True)),
        pendentes=Count('id', filter=Q(pago=False)),
        valor_pendente=Sum('valor_atual', filter=Q(pago=False)),
        vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
        valor_vencido=Sum('valor_atual', filter=Q(pago=False, data_vencimento__lt=hoje)),
    )

    # Agrupar por imobiliária para resumo
    parcelas_por_imobiliaria = {}
    for parcela in parcelas:
        imob_nome = parcela.contrato.imovel.imobiliaria.nome if parcela.contrato.imovel.imobiliaria else 'Sem Imobiliária'
        if imob_nome not in parcelas_por_imobiliaria:
            parcelas_por_imobiliaria[imob_nome] = {
                'total': 0,
                'pagas': 0,
                'pendentes': 0,
                'valor_total': Decimal('0.00'),
                'valor_pago': Decimal('0.00'),
                'valor_pendente': Decimal('0.00'),
            }
        parcelas_por_imobiliaria[imob_nome]['total'] += 1
        parcelas_por_imobiliaria[imob_nome]['valor_total'] += parcela.valor_atual
        if parcela.pago:
            parcelas_por_imobiliaria[imob_nome]['pagas'] += 1
            parcelas_por_imobiliaria[imob_nome]['valor_pago'] += parcela.valor_pago or Decimal('0.00')
        else:
            parcelas_por_imobiliaria[imob_nome]['pendentes'] += 1
            parcelas_por_imobiliaria[imob_nome]['valor_pendente'] += parcela.valor_atual

    per_page = request.GET.get('per_page', '25')
    try:
        per_page = min(int(per_page), 100)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(parcelas, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'parcelas': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
        'imobiliarias': imobiliarias,
        'filtro_imobiliaria': imobiliaria_id,
        'filtro_status': status,
        'mes_atual': hoje.strftime('%B/%Y'),
        'primeiro_dia': primeiro_dia_mes,
        'ultimo_dia': ultimo_dia_mes,
        'stats': stats,
        'parcelas_por_imobiliaria': parcelas_por_imobiliaria,
    }
    return render(request, 'financeiro/parcelas_mes.html', context)


@login_required
def dashboard_imobiliaria(request, imobiliaria_id):
    """
    Dashboard financeiro detalhado para uma imobiliária específica.

    Inclui:
    - Estatísticas gerais de parcelas e contratos
    - Contratos com reajuste pendente
    - Parcelas bloqueadas por reajuste
    - Prestações intermediárias
    - Top 10 compradores com mais atraso
    """
    imobiliaria = get_object_or_404(Imobiliaria, pk=imobiliaria_id, ativo=True)
    hoje = timezone.now().date()

    # Período do mês atual
    primeiro_dia_mes = hoje.replace(day=1)
    if hoje.month == 12:
        ultimo_dia_mes = hoje.replace(day=31)
    else:
        ultimo_dia_mes = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)

    # Todas as parcelas da imobiliária
    parcelas_imob = Parcela.objects.select_related(
        'contrato',
        'contrato__comprador',
        'contrato__imovel'
    ).filter(
        contrato__imovel__imobiliaria=imobiliaria
    )

    # Estatísticas gerais
    stats_geral = parcelas_imob.aggregate(
        total=Count('id'),
        pagas=Count('id', filter=Q(pago=True)),
        pendentes=Count('id', filter=Q(pago=False)),
        vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
        valor_total=Sum('valor_atual'),
        valor_recebido=Sum('valor_pago', filter=Q(pago=True)),
        valor_pendente=Sum('valor_atual', filter=Q(pago=False)),
        valor_vencido=Sum('valor_atual', filter=Q(pago=False, data_vencimento__lt=hoje)),
    )

    # Parcelas do mês atual
    parcelas_mes = parcelas_imob.filter(
        data_vencimento__gte=primeiro_dia_mes,
        data_vencimento__lte=ultimo_dia_mes
    )

    stats_mes = parcelas_mes.aggregate(
        total=Count('id'),
        pagas=Count('id', filter=Q(pago=True)),
        pendentes=Count('id', filter=Q(pago=False)),
        vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
        valor_total=Sum('valor_atual'),
        valor_recebido=Sum('valor_pago', filter=Q(pago=True)),
        valor_pendente=Sum('valor_atual', filter=Q(pago=False)),
    )

    # Parcelas vencidas (não pagas e data < hoje)
    parcelas_vencidas = parcelas_imob.filter(
        pago=False,
        data_vencimento__lt=hoje
    ).order_by('data_vencimento')[:20]

    # Próximas parcelas a vencer (próximos 30 dias)
    proximas_parcelas = parcelas_imob.filter(
        pago=False,
        data_vencimento__gte=hoje,
        data_vencimento__lte=hoje + timedelta(days=30)
    ).order_by('data_vencimento')[:20]

    # Contratos da imobiliária
    contratos = Contrato.objects.filter(
        imovel__imobiliaria=imobiliaria
    ).select_related('comprador', 'imovel')

    stats_contratos = contratos.aggregate(
        total=Count('id'),
        ativos=Count('id', filter=Q(status=StatusContrato.ATIVO)),
        quitados=Count('id', filter=Q(status=StatusContrato.QUITADO)),
        valor_total=Sum('valor_total'),
    )

    # =========================================================================
    # CONTRATOS COM REAJUSTE PENDENTE
    # =========================================================================
    contratos_reajuste_pendente = []
    contratos_com_boleto_bloqueado = []

    for contrato in contratos.filter(status=StatusContrato.ATIVO):
        # Verificar se tem bloqueio de reajuste
        # verificar_bloqueio_reajuste() retorna bool (True = bloqueado)
        if hasattr(contrato, 'verificar_bloqueio_reajuste') and contrato.verificar_bloqueio_reajuste():
            contratos_com_boleto_bloqueado.append({
                'contrato': contrato,
                'ciclo_atual': 1,
                'ciclo_pendente': None,
                'motivo': '',
            })

        # Verificar reajuste pendente
        if hasattr(contrato, 'data_ultimo_reajuste') and hasattr(contrato, 'prazo_reajuste_meses'):
            data_base = contrato.data_ultimo_reajuste or contrato.data_contrato
            proxima_data_reajuste = data_base + relativedelta(months=contrato.prazo_reajuste_meses)

            if proxima_data_reajuste <= hoje:
                meses_atraso = (hoje.year - proxima_data_reajuste.year) * 12 + (hoje.month - proxima_data_reajuste.month)
                contratos_reajuste_pendente.append({
                    'contrato': contrato,
                    'data_ultimo_reajuste': contrato.data_ultimo_reajuste,
                    'proxima_data_reajuste': proxima_data_reajuste,
                    'meses_atraso': meses_atraso,
                    'tipo_correcao': contrato.get_tipo_correcao_display() if hasattr(contrato, 'get_tipo_correcao_display') else 'N/A',
                })

    # Ordenar por meses em atraso
    contratos_reajuste_pendente.sort(key=lambda x: x['meses_atraso'], reverse=True)

    # =========================================================================
    # PRESTAÇÕES INTERMEDIÁRIAS
    # =========================================================================
    from contratos.models import PrestacaoIntermediaria

    intermediarias_pendentes = PrestacaoIntermediaria.objects.filter(
        contrato__imovel__imobiliaria=imobiliaria,
        paga=False
    ).select_related('contrato', 'contrato__comprador').order_by('mes_vencimento')[:10]

    stats_intermediarias = PrestacaoIntermediaria.objects.filter(
        contrato__imovel__imobiliaria=imobiliaria
    ).aggregate(
        total=Count('id'),
        pagas=Count('id', filter=Q(paga=True)),
        pendentes=Count('id', filter=Q(paga=False)),
        valor_total=Sum('valor'),
        valor_pago=Sum('valor_pago', filter=Q(paga=True)),
        valor_pendente=Sum('valor', filter=Q(paga=False)),
    )

    # Top 10 compradores com mais atraso
    compradores_atraso = []
    for contrato in contratos.filter(status=StatusContrato.ATIVO):
        parcelas_atraso = contrato.parcelas.filter(pago=False, data_vencimento__lt=hoje)
        if parcelas_atraso.exists():
            valor_atraso = parcelas_atraso.aggregate(total=Sum('valor_atual'))['total'] or Decimal('0.00')
            compradores_atraso.append({
                'comprador': contrato.comprador,
                'contrato': contrato,
                'qtd_parcelas': parcelas_atraso.count(),
                'valor_atraso': valor_atraso,
                'dias_atraso': (hoje - parcelas_atraso.first().data_vencimento).days,
            })
    compradores_atraso.sort(key=lambda x: x['valor_atraso'], reverse=True)

    context = {
        'imobiliaria': imobiliaria,
        'stats_geral': stats_geral,
        'stats_mes': stats_mes,
        'stats_contratos': stats_contratos,
        'parcelas_vencidas': parcelas_vencidas,
        'proximas_parcelas': proximas_parcelas,
        'compradores_atraso': compradores_atraso[:10],
        'mes_atual': hoje.strftime('%B/%Y'),
        # Novos campos
        'contratos_reajuste_pendente': contratos_reajuste_pendente[:10],
        'contratos_com_boleto_bloqueado': contratos_com_boleto_bloqueado[:10],
        'total_contratos_bloqueados': len(contratos_com_boleto_bloqueado),
        'total_contratos_reajuste_pendente': len(contratos_reajuste_pendente),
        'intermediarias_pendentes': intermediarias_pendentes,
        'stats_intermediarias': stats_intermediarias,
    }
    return render(request, 'financeiro/dashboard_imobiliaria.html', context)


# =============================================================================
# VIEWS DE BOLETO BANCÁRIO
# =============================================================================

@login_required
@require_POST
def gerar_boleto_parcela(request, pk):
    """
    Gera boleto para uma parcela específica.

    IMPORTANTE: Verifica se a parcela pode ter boleto gerado considerando
    o ciclo de reajuste. Boletos após o 12º mês só podem ser gerados
    após aplicação do reajuste correspondente.
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    if parcela.pago:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Parcela já está paga'
        }, status=400)

    # =========================================================================
    # VERIFICAÇÃO DE BLOQUEIO POR REAJUSTE
    # =========================================================================
    contrato = parcela.contrato
    force = request.POST.get('force', 'false').lower() == 'true'

    # Só verifica bloqueio se não for forçado (force=true ignora o bloqueio)
    if not force and hasattr(contrato, 'pode_gerar_boleto'):
        pode_gerar, motivo = contrato.pode_gerar_boleto(parcela.numero_parcela)
        if not pode_gerar:
            return JsonResponse({
                'sucesso': False,
                'erro': f'Boleto bloqueado: {motivo}',
                'bloqueado_reajuste': True,
                'ciclo_atual': getattr(contrato, 'ciclo_reajuste_atual', 1),
                'numero_parcela': parcela.numero_parcela
            }, status=400)

    # Obter conta bancária (pode vir do request ou usar a padrão)
    conta_id = request.POST.get('conta_bancaria_id')
    conta_bancaria = None

    if conta_id:
        conta_bancaria = get_object_or_404(ContaBancaria, pk=conta_id, ativo=True)
    else:
        # Usar conta principal da imobiliária
        imobiliaria = parcela.contrato.imovel.imobiliaria
        conta_bancaria = imobiliaria.contas_bancarias.filter(
            principal=True, ativo=True
        ).first()

    if not conta_bancaria:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Nenhuma conta bancária configurada'
        }, status=400)

    try:
        resultado = parcela.gerar_boleto(conta_bancaria, force=force)

        if resultado and resultado.get('sucesso'):
            # Atualizar último mês com boleto gerado no contrato
            if hasattr(contrato, 'ultimo_mes_boleto_gerado'):
                if parcela.numero_parcela > contrato.ultimo_mes_boleto_gerado:
                    contrato.ultimo_mes_boleto_gerado = parcela.numero_parcela
                    contrato.save(update_fields=['ultimo_mes_boleto_gerado'])

            return JsonResponse({
                'sucesso': True,
                'parcela_id': parcela.id,
                'nosso_numero': resultado.get('nosso_numero', ''),
                'linha_digitavel': resultado.get('linha_digitavel', ''),
                'codigo_barras': resultado.get('codigo_barras', ''),
                'tem_pdf': parcela.boleto_pdf.name if parcela.boleto_pdf else False,
                'mensagem': 'Boleto gerado com sucesso'
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'erro': resultado.get('erro', 'Erro ao gerar boleto') if resultado else 'Erro desconhecido'
            }, status=400)

    except Exception as e:
        logger.exception(f"Erro ao gerar boleto: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
@require_POST
def gerar_boletos_contrato(request, contrato_id):
    """
    Gera boletos para todas as parcelas pendentes de um contrato.

    IMPORTANTE: Verifica o bloqueio por reajuste antes de gerar cada boleto.
    Parcelas de ciclos não reajustados serão marcadas como bloqueadas.
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    # Obter conta bancária
    conta_id = request.POST.get('conta_bancaria_id')
    conta_bancaria = None
    force = request.POST.get('force', 'false').lower() == 'true'

    if conta_id:
        conta_bancaria = get_object_or_404(ContaBancaria, pk=conta_id, ativo=True)

    try:
        # =====================================================================
        # VERIFICAÇÃO DE BLOQUEIO POR REAJUSTE
        # =====================================================================
        parcelas_pendentes = contrato.parcelas.filter(pago=False).order_by('numero_parcela')

        if not parcelas_pendentes.exists():
            return JsonResponse({
                'sucesso': False,
                'erro': 'Não há parcelas pendentes para gerar boletos'
            }, status=400)

        resultados = []
        gerados = 0
        bloqueados = 0
        erros = 0

        for parcela in parcelas_pendentes:
            # Verificar bloqueio por reajuste
            if not force and hasattr(contrato, 'pode_gerar_boleto'):
                pode_gerar, motivo = contrato.pode_gerar_boleto(parcela.numero_parcela)
                if not pode_gerar:
                    resultados.append({
                        'parcela_id': parcela.id,
                        'numero_parcela': parcela.numero_parcela,
                        'sucesso': False,
                        'bloqueado_reajuste': True,
                        'erro': motivo
                    })
                    bloqueados += 1
                    continue

            # Verificar se já tem boleto
            if parcela.tem_boleto and not force:
                resultados.append({
                    'parcela_id': parcela.id,
                    'numero_parcela': parcela.numero_parcela,
                    'sucesso': True,
                    'mensagem': 'Boleto já existente'
                })
                continue

            # Gerar boleto
            try:
                resultado = parcela.gerar_boleto(conta_bancaria, force=force)
                if resultado and resultado.get('sucesso'):
                    gerados += 1
                    resultados.append({
                        'parcela_id': parcela.id,
                        'numero_parcela': parcela.numero_parcela,
                        'sucesso': True,
                        'nosso_numero': resultado.get('nosso_numero')
                    })

                    # Atualizar último mês com boleto gerado
                    if hasattr(contrato, 'ultimo_mes_boleto_gerado'):
                        if parcela.numero_parcela > contrato.ultimo_mes_boleto_gerado:
                            contrato.ultimo_mes_boleto_gerado = parcela.numero_parcela
                else:
                    erros += 1
                    resultados.append({
                        'parcela_id': parcela.id,
                        'numero_parcela': parcela.numero_parcela,
                        'sucesso': False,
                        'erro': resultado.get('erro') if resultado else 'Erro desconhecido'
                    })
            except Exception as e:
                logger.exception('Erro ao gerar boleto parcela pk=%s: %s', parcela.id, e)
                erros += 1
                resultados.append({
                    'parcela_id': parcela.id,
                    'numero_parcela': parcela.numero_parcela,
                    'sucesso': False,
                    'erro': str(e)
                })

        # Salvar atualização do último mês com boleto gerado
        if hasattr(contrato, 'ultimo_mes_boleto_gerado'):
            contrato.save(update_fields=['ultimo_mes_boleto_gerado'])

        total = len(resultados)

        return JsonResponse({
            'sucesso': True,
            'total': total,
            'gerados': gerados,
            'bloqueados': bloqueados,
            'erros': erros,
            'resultados': resultados,
            'mensagem': f'{gerados} boletos gerados, {bloqueados} bloqueados por reajuste, {erros} erros'
        })

    except Exception as e:
        logger.exception(f"Erro ao gerar boletos do contrato: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
@require_GET
def download_boleto(request, pk):
    """
    Download do PDF do boleto.
    Regenera automaticamente se o arquivo não existir (storage efêmero no Render).
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    if not parcela.boleto_pdf:
        messages.error(request, 'Boleto não disponível para download.')
        return redirect('financeiro:detalhe_parcela', pk=pk)

    # Se o arquivo não existir no disco (storage efêmero do Render), serve do banco de dados
    if not parcela.boleto_pdf.storage.exists(parcela.boleto_pdf.name):
        logger.warning(
            "Arquivo de boleto ausente no storage (%s), buscando do banco de dados...",
            parcela.boleto_pdf.name
        )
        if parcela.boleto_pdf_db:
            filename = f'boleto_{parcela.contrato.numero_contrato}_{parcela.numero_parcela}.pdf'
            response = HttpResponse(bytes(parcela.boleto_pdf_db), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # Sem cópia em DB: tenta regenerar via BRCobrança (usa Parcela.gerar_boleto para
        # resolver conta_bancaria automaticamente e não exigir parâmetro extra)
        logger.warning("Sem cópia em DB para pk=%s, tentando regenerar...", pk)
        try:
            resultado = parcela.gerar_boleto(force=True, enviar_email=False)
            if resultado and resultado.get('sucesso'):
                parcela.refresh_from_db()
                logger.info("Boleto regenerado com sucesso para parcela pk=%s", pk)
            else:
                logger.error("Falha ao regenerar boleto pk=%s: %s", pk, resultado.get('erro'))
                messages.error(
                    request,
                    'O arquivo do boleto foi perdido. Clique em "Gerar Boleto" para criar um novo.'
                )
                return redirect('financeiro:detalhe_parcela', pk=pk)
        except Exception as e:
            logger.exception("Erro ao regenerar boleto pk=%s: %s", pk, e)
            messages.error(request, 'Não foi possível recuperar o boleto. Gere novamente.')
            return redirect('financeiro:detalhe_parcela', pk=pk)

    try:
        response = FileResponse(
            parcela.boleto_pdf.open('rb'),
            as_attachment=True,
            filename=f'boleto_{parcela.contrato.numero_contrato}_{parcela.numero_parcela}.pdf'
        )
        response['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        logger.exception(f"Erro ao fazer download do boleto: {e}")
        messages.error(request, 'Erro ao baixar o boleto.')
        return redirect('financeiro:detalhe_parcela', pk=pk)


@login_required
def download_zip_boletos(request, contrato_id):
    """
    Download em ZIP de todos os boletos com PDF de um contrato.
    POST opcionalmente com lista de parcela_ids para filtrar.
    """
    import io
    import zipfile

    from contratos.models import Contrato
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    if request.method == 'POST':
        ids_raw = request.POST.getlist('parcela_ids')
        if ids_raw:
            parcelas = Parcela.objects.filter(
                contrato=contrato,
                id__in=[int(i) for i in ids_raw if i.isdigit()],
            ).exclude(boleto_pdf='')
        else:
            parcelas = Parcela.objects.filter(contrato=contrato).exclude(boleto_pdf='')
    else:
        parcelas = Parcela.objects.filter(contrato=contrato).exclude(boleto_pdf='')

    parcelas = [p for p in parcelas if p.boleto_pdf]

    if not parcelas:
        messages.error(request, 'Nenhum boleto disponível para download neste contrato.')
        return redirect('contratos:detalhe', pk=contrato_id)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in parcelas:
            try:
                fname = f'boleto_{contrato.numero_contrato}_parcela_{p.numero_parcela:03d}.pdf'
                with p.boleto_pdf.open('rb') as f:
                    zf.writestr(fname, f.read())
            except Exception as e:
                logger.warning("Erro ao incluir boleto parcela %s no ZIP: %s", p.pk, e)

    buf.seek(0)
    zip_name = f'boletos_{contrato.numero_contrato}.zip'
    response = HttpResponse(buf.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    return response


@login_required
def segunda_via_boleto(request, pk):
    """
    Segunda via de boleto com juros/multa atualizados para hoje.
    GET → página de preview com valores atualizados.
    GET ?download=1 → gera PDF fresco via BRCobrança e retorna download.
    """
    from financeiro.services.boleto_service import BoletoService

    parcela = get_object_or_404(Parcela, pk=pk)
    contrato = parcela.contrato
    hoje = timezone.localdate()

    if parcela.pago:
        messages.info(request, f'Parcela {parcela.numero_parcela} já foi paga em {parcela.data_pagamento}.')
        return redirect('financeiro:detalhe_parcela', pk=pk)

    # Calcular juros/multa atualizados
    valor_juros = Decimal('0.00')
    valor_multa = Decimal('0.00')
    if hoje > parcela.data_vencimento:
        valor_juros, valor_multa = parcela.calcular_juros_multa(hoje)
    valor_total = parcela.valor_atual + valor_juros + valor_multa

    if request.GET.get('download') == '1':
        # Obter conta bancária
        conta_bancaria = None
        if hasattr(contrato, 'get_conta_bancaria'):
            conta_bancaria = contrato.get_conta_bancaria()
        if not conta_bancaria:
            conta_bancaria = contrato.imobiliaria.contas_bancarias.filter(principal=True, ativo=True).first()
        if not conta_bancaria:
            conta_bancaria = contrato.imobiliaria.contas_bancarias.filter(ativo=True).first()

        if not conta_bancaria:
            messages.error(request, 'Nenhuma conta bancária configurada para esta imobiliária.')
            return redirect('financeiro:detalhe_parcela', pk=pk)

        servico = BoletoService()
        resultado = servico.gerar_segunda_via(parcela, conta_bancaria, data_referencia=hoje)

        if resultado.get('sucesso') and resultado.get('pdf_content'):
            fname = f'segunda_via_{contrato.numero_contrato}_parcela_{parcela.numero_parcela}.pdf'
            response = HttpResponse(resultado['pdf_content'], content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{fname}"'
            return response
        else:
            messages.error(request, f"Erro ao gerar segunda via: {resultado.get('erro', 'Erro desconhecido')}")
            return redirect('financeiro:detalhe_parcela', pk=pk)

    context = {
        'parcela': parcela,
        'contrato': contrato,
        'hoje': hoje,
        'valor_juros': valor_juros,
        'valor_multa': valor_multa,
        'valor_total': valor_total,
        'dias_atraso': (hoje - parcela.data_vencimento).days if hoje > parcela.data_vencimento else 0,
    }
    return render(request, 'financeiro/segunda_via_boleto.html', context)


@login_required
@require_GET
@xframe_options_sameorigin
def visualizar_boleto(request, pk):
    """
    Exibe página com dados do boleto de uma parcela.
    Permite carregamento em iframe do mesmo domínio.
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    if not parcela.tem_boleto:
        messages.warning(request, 'Boleto ainda não foi gerado para esta parcela.')
        return redirect('financeiro:detalhe_parcela', pk=pk)

    contrato = parcela.contrato
    comprador = contrato.comprador
    imobiliaria = contrato.imobiliaria

    # Calcular valores para hoje
    valores_hoje = parcela.calcular_valores_hoje()

    # Verificar se e popup (sem navbar) - verificando se tem parametro popup ou se nao tem referer
    popup = request.GET.get('popup', 'false').lower() == 'true' or not request.META.get('HTTP_REFERER')

    from django.urls import reverse
    context = {
        'parcela': parcela,
        'contrato': contrato,
        'comprador': comprador,
        'imobiliaria': imobiliaria,
        'valores_hoje': valores_hoje,
        'popup': popup,
        'voltar_url': _voltar_url(
            request, reverse('financeiro:detalhe_parcela', args=[parcela.pk])
        ),
    }

    return render(request, 'financeiro/visualizar_boleto.html', context)


@login_required
@require_POST
def cancelar_boleto(request, pk):
    """
    Cancela um boleto.
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    if parcela.status_boleto in [StatusBoleto.NAO_GERADO, StatusBoleto.CANCELADO]:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Boleto não pode ser cancelado'
        }, status=400)

    if parcela.pago:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Não é possível cancelar boleto de parcela paga'
        }, status=400)

    motivo = request.POST.get('motivo', '')

    try:
        parcela.cancelar_boleto(motivo)
        return JsonResponse({
            'sucesso': True,
            'mensagem': 'Boleto cancelado com sucesso'
        })
    except Exception as e:
        logger.exception(f"Erro ao cancelar boleto: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
def api_status_boleto(request, pk):
    """
    Retorna o status do boleto de uma parcela (para atualização via AJAX).
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    return JsonResponse({
        'parcela_id': parcela.id,
        'status': parcela.status_boleto,
        'status_display': parcela.get_status_boleto_display(),
        'tem_boleto': parcela.tem_boleto,
        'nosso_numero': parcela.nosso_numero,
        'linha_digitavel': parcela.linha_digitavel,
        'tem_pdf': bool(parcela.boleto_pdf),
        'data_geracao': parcela.data_geracao_boleto.isoformat() if parcela.data_geracao_boleto else None,
    })


# =============================================================================
# VIEWS CNAB - ARQUIVOS DE REMESSA E RETORNO
# =============================================================================

@login_required
def listar_arquivos_remessa(request):
    """Lista todos os arquivos de remessa"""
    from .models import ArquivoRemessa
    from .services.cnab_service import CNABService

    arquivos = ArquivoRemessa.objects.select_related(
        'conta_bancaria', 'conta_bancaria__imobiliaria'
    ).order_by('-data_geracao')

    # Filtros
    conta_id = request.GET.get('conta')
    status = request.GET.get('status')
    imobiliaria_id = request.GET.get('imobiliaria')

    if conta_id:
        arquivos = arquivos.filter(conta_bancaria_id=conta_id)
    if status:
        arquivos = arquivos.filter(status=status)
    if imobiliaria_id:
        arquivos = arquivos.filter(conta_bancaria__imobiliaria_id=imobiliaria_id)

    contas = ContaBancaria.objects.filter(ativo=True).select_related('imobiliaria')
    imobiliarias = Imobiliaria.objects.filter(ativo=True).order_by('nome')

    per_page = request.GET.get('per_page', '25')
    try:
        per_page = min(int(per_page), 100)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(arquivos, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # Contador de boletos disponíveis para nova remessa
    boletos_pendentes_count = len(CNABService().obter_boletos_sem_remessa())

    context = {
        'arquivos': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
        'contas_bancarias': contas,
        'imobiliarias': imobiliarias,
        'filtro_conta': conta_id,
        'filtro_status': status,
        'filtro_imobiliaria': imobiliaria_id,
        'boletos_pendentes_count': boletos_pendentes_count,
    }
    return render(request, 'financeiro/cnab/listar_remessas.html', context)


@login_required
def detalhe_arquivo_remessa(request, pk):
    """Exibe detalhes de um arquivo de remessa"""
    from .models import ArquivoRemessa

    arquivo = get_object_or_404(
        ArquivoRemessa.objects.select_related(
            'conta_bancaria', 'conta_bancaria__imobiliaria'
        ).prefetch_related(
            'itens', 'itens__parcela', 'itens__parcela__contrato',
            'itens__parcela__contrato__comprador'
        ),
        pk=pk
    )

    context = {
        'arquivo': arquivo,
        'itens': arquivo.itens.all(),
    }
    return render(request, 'financeiro/cnab/detalhe_remessa.html', context)


@login_required
def gerar_arquivo_remessa(request):
    """
    View para gerar novo arquivo de remessa.

    GET: Exibe boletos disponíveis com filtros por escopo
         Escopos: tudo | imobiliaria | contrato | conta
    POST: Gera 1 arquivo de remessa por conta bancária a partir
          das parcelas selecionadas (auto-split por conta).
    """
    from .services.cnab_service import CNABService

    service = CNABService()
    imobiliarias = Imobiliaria.objects.filter(ativo=True).order_by('nome')
    contas = ContaBancaria.objects.filter(ativo=True).select_related('imobiliaria')
    contratos = Contrato.objects.filter(
        status='ATIVO'
    ).select_related('comprador', 'imobiliaria').order_by('numero_contrato')

    if request.method == 'POST':
        layout = request.POST.get('layout', 'CNAB_240')
        parcela_ids = [int(x) for x in request.POST.getlist('parcelas') if x.isdigit()]

        if not parcela_ids:
            messages.error(request, 'Selecione pelo menos uma parcela.')
            return redirect('financeiro:gerar_remessa')

        # Verificar conflito de layout CNAB entre as contas selecionadas
        from .models import Parcela as _Parcela
        contas_selecionadas = (
            _Parcela.objects.filter(pk__in=parcela_ids)
            .select_related('conta_bancaria')
            .values_list(
                'conta_bancaria__layout_cnab',
                'conta_bancaria__banco',
            )
            .distinct()
        )
        layouts_por_banco = {}  # {layout: set(banco_codigo)}
        for lay, banco in contas_selecionadas:
            layouts_por_banco.setdefault(lay, set()).add(banco or '000')
        if len(layouts_por_banco) > 1:
            banco_display = dict(BancoBrasil.choices)
            partes = []
            for lay, bancos in layouts_por_banco.items():
                nomes = ', '.join(banco_display.get(b, b) for b in sorted(bancos))
                partes.append(f"{nomes} ({lay.replace('_', ' ')})")
            messages.error(
                request,
                f"{' e '.join(partes)} não usam o mesmo padrão de CNAB. "
                "Gere os CNAB's separadamente."
            )
            return redirect('financeiro:gerar_remessa')

        # Verificar se há parcelas pagas na seleção — informar ao usuário
        pagas_selecionadas = _Parcela.objects.filter(pk__in=parcela_ids, pago=True).count()
        if pagas_selecionadas:
            messages.warning(
                request,
                f'{pagas_selecionadas} parcela(s) PAGA(S) foram ignoradas — '
                'parcelas pagas não entram em arquivo de remessa.'
            )

        try:
            resultado = service.gerar_remessas_por_escopo(parcela_ids, layout)

            remessas = resultado['remessas_geradas']
            erros = resultado['erros']

            for r in remessas:
                arq = r['arquivo_remessa']
                aviso = f" ({r['aviso']})" if r.get('aviso') else ''
                messages.success(
                    request,
                    f"Remessa #{arq.numero_remessa} — {arq.conta_bancaria}: "
                    f"{r['quantidade_boletos']} boleto(s), "
                    f"R$ {r['valor_total']:,.2f}{aviso}"
                )

            for erro in erros:
                messages.warning(request, erro)

            if not remessas:
                messages.error(request, 'Nenhuma remessa foi gerada.')
                return redirect('financeiro:gerar_remessa')

            # Se gerou apenas 1, vai direto para o detalhe
            if len(remessas) == 1:
                return redirect('financeiro:detalhe_remessa', pk=remessas[0]['arquivo_remessa'].pk)

            return redirect('financeiro:listar_remessas')

        except Exception as e:
            logger.exception(f"Erro ao gerar remessa(s): {e}")
            messages.error(request, f"Erro ao gerar remessa: {str(e)}")
            return redirect('financeiro:gerar_remessa')

    # GET — filtros de escopo
    escopo = request.GET.get('escopo', 'tudo')
    imobiliaria_id = request.GET.get('imobiliaria') or None
    contrato_id = request.GET.get('contrato') or None
    conta_id = request.GET.get('conta') or None

    conta_obj = None
    if conta_id:
        conta_obj = ContaBancaria.objects.filter(pk=conta_id, ativo=True).first()

    # Parcelas disponíveis (sem remessa)
    boletos_disponiveis = service.obter_boletos_sem_remessa(
        conta_bancaria=conta_obj,
        imobiliaria_id=imobiliaria_id,
        contrato_id=contrato_id,
    )

    # Parcelas em remessa pendente (já em GERADO, não enviada) — para aviso
    boletos_pendentes = service.obter_boletos_em_remessa_pendente(
        conta_bancaria=conta_obj,
        imobiliaria_id=imobiliaria_id,
        contrato_id=contrato_id,
    )

    # Agrupar disponíveis por conta_bancaria para exibição
    from collections import defaultdict
    grupos_conta = defaultdict(list)
    for p in boletos_disponiveis:
        grupos_conta[p.conta_bancaria].append(p)

    context = {
        'contas_bancarias': contas,
        'imobiliarias': imobiliarias,
        'contratos': contratos,
        'escopo': escopo,
        'filtro_conta': conta_id,
        'filtro_imobiliaria': imobiliaria_id,
        'filtro_contrato': contrato_id,
        'grupos_conta': dict(grupos_conta),  # {ContaBancaria: [Parcela, ...]}
        'boletos_disponiveis': boletos_disponiveis,
        'boletos_pendentes': boletos_pendentes,
        'total_disponivel': len(boletos_disponiveis),
        'today': timezone.localdate(),
    }
    return render(request, 'financeiro/cnab/gerar_remessa.html', context)


@login_required
@require_POST
def regenerar_arquivo_remessa(request, pk):
    """Regenera um arquivo de remessa existente"""
    from .models import ArquivoRemessa
    from .services.cnab_service import CNABService

    arquivo = get_object_or_404(ArquivoRemessa, pk=pk)

    if not arquivo.pode_reenviar:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Este arquivo nao pode ser regenerado'
        }, status=400)

    try:
        service = CNABService()
        resultado = service.regenerar_remessa(arquivo)

        if resultado.get('sucesso'):
            return JsonResponse({
                'sucesso': True,
                'mensagem': 'Remessa regenerada com sucesso',
                'redirect': f"/financeiro/cnab/remessa/{resultado.get('arquivo_remessa').pk}/"
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'erro': resultado.get('erro')
            }, status=400)

    except Exception as e:
        logger.exception(f"Erro ao regenerar remessa: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
@require_POST
def marcar_remessa_enviada(request, pk):
    """Marca um arquivo de remessa como enviado ao banco"""
    from .models import ArquivoRemessa

    arquivo = get_object_or_404(ArquivoRemessa, pk=pk)

    if arquivo.status != 'GERADO':
        return JsonResponse({
            'sucesso': False,
            'erro': 'Apenas arquivos com status GERADO podem ser marcados como enviados'
        }, status=400)

    arquivo.marcar_enviado()
    return JsonResponse({
        'sucesso': True,
        'mensagem': 'Remessa marcada como enviada'
    })


@login_required
def download_arquivo_remessa(request, pk):
    """Download do arquivo de remessa"""
    from .models import ArquivoRemessa

    arquivo = get_object_or_404(ArquivoRemessa, pk=pk)

    if not arquivo.arquivo:
        messages.error(request, 'Arquivo nao disponivel para download.')
        return redirect('financeiro:detalhe_remessa', pk=pk)

    try:
        response = FileResponse(
            arquivo.arquivo.open('rb'),
            as_attachment=True,
            filename=arquivo.nome_arquivo
        )
        response['Content-Type'] = 'text/plain'
        return response
    except Exception as e:
        logger.exception(f"Erro ao fazer download da remessa: {e}")
        messages.error(request, 'Erro ao baixar o arquivo.')
        return redirect('financeiro:detalhe_remessa', pk=pk)


# =============================================================================
# DASHBOARD DE CONCILIAÇÃO BANCÁRIA
# =============================================================================

@login_required
def dashboard_conciliacao(request):
    """
    Hub de conciliação bancária — reúne os 3 métodos de baixa:
      1. Retorno CNAB (arquivo do banco)
      2. Extrato OFX
      3. Baixa manual

    Exibe KPIs, fila de boletos pendentes de conciliação e histórico recente.
    """
    from django.db.models import Count, Sum
    from .models import (
        Parcela, HistoricoPagamento, ArquivoRetorno, StatusBoleto,
        ItemRetorno,
    )
    from core.models import Imobiliaria

    imobiliarias = Imobiliaria.objects.filter(ativo=True).order_by('nome')
    imob_id = request.GET.get('imobiliaria')
    periodo = request.GET.get('periodo', '30')  # dias
    try:
        dias = int(periodo)
    except ValueError:
        dias = 30

    desde = timezone.now().date() - timezone.timedelta(days=dias)

    # ── Boletos aguardando conciliação ────────────────────────────────────────
    qs_pendentes = Parcela.objects.filter(
        status_boleto__in=[StatusBoleto.GERADO, StatusBoleto.REGISTRADO, StatusBoleto.VENCIDO],
        pago=False,
    ).select_related('contrato', 'contrato__comprador', 'contrato__imobiliaria', 'conta_bancaria_boleto')

    if imob_id:
        qs_pendentes = qs_pendentes.filter(contrato__imobiliaria_id=imob_id)

    total_pendentes = qs_pendentes.count()
    valor_pendente = qs_pendentes.aggregate(s=Sum('valor_boleto'))['s'] or 0

    # ── Histórico de conciliações (últimos N dias) ────────────────────────────
    qs_hist = HistoricoPagamento.objects.filter(
        data_pagamento__gte=desde,
    ).select_related('parcela', 'parcela__contrato', 'parcela__contrato__comprador')

    if imob_id:
        qs_hist = qs_hist.filter(parcela__contrato__imobiliaria_id=imob_id)

    hist_por_origem = (
        qs_hist.values('origem_pagamento')
        .annotate(total=Count('id'), valor=Sum('valor_pago'))
        .order_by('origem_pagamento')
    )
    # normaliza em dict para template
    origem_stats = {
        'CNAB':      {'total': 0, 'valor': 0},
        'OFX':       {'total': 0, 'valor': 0},
        'MANUAL':    {'total': 0, 'valor': 0},
        'ANTECIPACAO': {'total': 0, 'valor': 0},
    }
    for row in hist_por_origem:
        key = row['origem_pagamento']
        if key in origem_stats:
            origem_stats[key] = {'total': row['total'], 'valor': float(row['valor'] or 0)}

    historico_recente = qs_hist.order_by('-data_pagamento', '-id')[:50]

    # ── Arquivos de retorno CNAB recentes ─────────────────────────────────────
    arquivos_retorno = ArquivoRetorno.objects.filter(
        data_upload__date__gte=desde
    ).select_related('conta_bancaria', 'conta_bancaria__imobiliaria', 'processado_por').order_by('-data_upload')[:10]

    # ── ItemRetornos com erro (não processados) ───────────────────────────────
    itens_erro = ItemRetorno.objects.filter(
        processado=False,
        parcela__isnull=False,
    ).exclude(erro_processamento='').select_related(
        'arquivo_retorno', 'parcela', 'parcela__contrato'
    ).order_by('-id')[:20]

    context = {
        'total_pendentes': total_pendentes,
        'valor_pendente': valor_pendente,
        'origem_stats': origem_stats,
        'historico_recente': historico_recente,
        'arquivos_retorno': arquivos_retorno,
        'itens_erro': itens_erro,
        'imobiliarias': imobiliarias,
        'imob_id': imob_id or '',
        'periodo': str(dias),
        'parcelas_pendentes': qs_pendentes.order_by('data_vencimento')[:100],
        'today': timezone.localdate(),
    }
    return render(request, 'financeiro/conciliacao/dashboard.html', context)


# =============================================================================
# VIEWS DE ARQUIVO DE RETORNO
# =============================================================================

@login_required
def listar_arquivos_retorno(request):
    """Lista todos os arquivos de retorno"""
    from .models import ArquivoRetorno

    arquivos = ArquivoRetorno.objects.select_related(
        'conta_bancaria', 'conta_bancaria__imobiliaria', 'processado_por'
    ).order_by('-data_upload')

    # Filtros
    conta_id = request.GET.get('conta')
    status = request.GET.get('status')

    if conta_id:
        arquivos = arquivos.filter(conta_bancaria_id=conta_id)
    if status:
        arquivos = arquivos.filter(status=status)

    contas = ContaBancaria.objects.filter(ativo=True).select_related('imobiliaria')

    per_page = request.GET.get('per_page', '25')
    try:
        per_page = min(int(per_page), 100)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(arquivos, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'arquivos': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
        'contas_bancarias': contas,
        'filtro_conta': conta_id,
        'filtro_status': status,
    }
    return render(request, 'financeiro/cnab/listar_retornos.html', context)


@login_required
def detalhe_arquivo_retorno(request, pk):
    """Exibe detalhes de um arquivo de retorno"""
    from .models import ArquivoRetorno

    arquivo = get_object_or_404(
        ArquivoRetorno.objects.select_related(
            'conta_bancaria', 'conta_bancaria__imobiliaria', 'processado_por'
        ).prefetch_related(
            'itens', 'itens__parcela', 'itens__parcela__contrato',
            'itens__parcela__contrato__comprador'
        ),
        pk=pk
    )

    # Agrupar itens por tipo de ocorrencia
    itens_por_tipo = {}
    for item in arquivo.itens.all():
        tipo = item.tipo_ocorrencia
        if tipo not in itens_por_tipo:
            itens_por_tipo[tipo] = []
        itens_por_tipo[tipo].append(item)

    context = {
        'arquivo': arquivo,
        'itens': arquivo.itens.all(),
        'itens_por_tipo': itens_por_tipo,
    }
    return render(request, 'financeiro/cnab/detalhe_retorno.html', context)


@login_required
def upload_arquivo_retorno(request):
    """Upload de arquivo de retorno para processamento"""
    from .models import ArquivoRetorno

    contas = ContaBancaria.objects.filter(ativo=True).select_related('imobiliaria')

    if request.method == 'POST':
        conta_id = request.POST.get('conta_bancaria')
        arquivo_upload = request.FILES.get('arquivo')

        if not conta_id:
            messages.error(request, 'Selecione uma conta bancaria.')
            return redirect('financeiro:upload_retorno')

        if not arquivo_upload:
            messages.error(request, 'Selecione um arquivo para upload.')
            return redirect('financeiro:upload_retorno')

        conta = get_object_or_404(ContaBancaria, pk=conta_id, ativo=True)

        try:
            arquivo_retorno = ArquivoRetorno.objects.create(
                conta_bancaria=conta,
                nome_arquivo=arquivo_upload.name,
            )
            arquivo_retorno.arquivo.save(arquivo_upload.name, arquivo_upload)

            messages.success(
                request,
                f"Arquivo '{arquivo_upload.name}' carregado com sucesso! "
                f"Clique em 'Processar' para realizar a baixa dos boletos."
            )
            return redirect('financeiro:detalhe_retorno', pk=arquivo_retorno.pk)

        except Exception as e:
            logger.exception(f"Erro ao fazer upload do retorno: {e}")
            messages.error(request, f"Erro ao fazer upload: {str(e)}")

        return redirect('financeiro:upload_retorno')

    context = {
        'contas_bancarias': contas,
    }
    return render(request, 'financeiro/cnab/upload_retorno.html', context)


@login_required
@require_POST
def processar_arquivo_retorno(request, pk):
    """Processa um arquivo de retorno (realiza as baixas)"""
    from .models import ArquivoRetorno
    from .services.cnab_service import CNABService

    arquivo = get_object_or_404(ArquivoRetorno, pk=pk)

    if not arquivo.pode_reprocessar:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Este arquivo ja foi processado e nao pode ser reprocessado'
        }, status=400)

    try:
        service = CNABService()
        resultado = service.processar_retorno(arquivo, request.user)

        if resultado.get('sucesso'):
            return JsonResponse({
                'sucesso': True,
                'mensagem': 'Arquivo processado com sucesso',
                'total_registros': resultado.get('total_registros'),
                'registros_processados': resultado.get('registros_processados'),
                'registros_erro': resultado.get('registros_erro'),
                'valor_total_pago': str(resultado.get('valor_total_pago')),
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'erro': resultado.get('erro')
            }, status=400)

    except Exception as e:
        logger.exception(f"Erro ao processar retorno: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
def download_arquivo_retorno(request, pk):
    """Download do arquivo de retorno"""
    from .models import ArquivoRetorno

    arquivo = get_object_or_404(ArquivoRetorno, pk=pk)

    if not arquivo.arquivo:
        messages.error(request, 'Arquivo nao disponivel para download.')
        return redirect('financeiro:detalhe_retorno', pk=pk)

    try:
        response = FileResponse(
            arquivo.arquivo.open('rb'),
            as_attachment=True,
            filename=arquivo.nome_arquivo
        )
        response['Content-Type'] = 'text/plain'
        return response
    except Exception as e:
        logger.exception(f"Erro ao fazer download do retorno: {e}")
        messages.error(request, 'Erro ao baixar o arquivo.')
        return redirect('financeiro:detalhe_retorno', pk=pk)


# =============================================================================
# VIEWS DE PAGAMENTO E CARNE
# =============================================================================

@login_required
@require_POST
def pagar_parcela_ajax(request, pk):
    """
    Registra o pagamento de uma parcela via AJAX.
    Permite editar juros, multa e desconto manualmente.
    """
    from .models import HistoricoPagamento

    parcela = get_object_or_404(Parcela, pk=pk)

    if parcela.pago:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Esta parcela ja esta paga'
        }, status=400)

    try:
        from datetime import date as _date
        data_pagamento_str = request.POST.get('data_pagamento', '').strip()
        try:
            data_pagamento = _date.fromisoformat(data_pagamento_str) if data_pagamento_str else timezone.localdate()
        except ValueError:
            data_pagamento = timezone.localdate()
        valor_pago = Decimal(request.POST.get('valor_pago', 0))
        valor_juros = Decimal(request.POST.get('valor_juros', 0))
        valor_multa = Decimal(request.POST.get('valor_multa', 0))
        valor_desconto = Decimal(request.POST.get('valor_desconto', 0))
        forma_pagamento = request.POST.get('forma_pagamento', 'DINHEIRO')
        observacoes = request.POST.get('observacoes', '')

        # Atualizar parcela
        parcela.valor_juros = valor_juros
        parcela.valor_multa = valor_multa
        parcela.valor_desconto = valor_desconto
        parcela.pago = True
        parcela.data_pagamento = data_pagamento
        parcela.valor_pago = valor_pago

        if observacoes:
            parcela.observacoes = observacoes

        # Atualizar status do boleto se houver
        if parcela.tem_boleto:
            parcela.status_boleto = StatusBoleto.PAGO
            parcela.data_pagamento_boleto = timezone.now()
            parcela.valor_pago_boleto = valor_pago

        parcela.save()

        # Registrar historico
        HistoricoPagamento.objects.create(
            parcela=parcela,
            data_pagamento=data_pagamento,
            valor_pago=valor_pago,
            valor_parcela=parcela.valor_atual,
            valor_juros=valor_juros,
            valor_multa=valor_multa,
            valor_desconto=valor_desconto,
            forma_pagamento=forma_pagamento,
            observacoes=observacoes,
            origem_pagamento='MANUAL',
        )

        return JsonResponse({
            'sucesso': True,
            'mensagem': 'Pagamento registrado com sucesso',
            'parcela_id': parcela.id
        })

    except Exception as e:
        logger.exception(f"Erro ao registrar pagamento: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
@require_GET
def api_calcular_encargos(request, pk):
    """
    Retorna juros, multa e valor total calculados para uma parcela em uma data de pagamento.
    GET /financeiro/parcelas/<pk>/calcular-encargos/?data=YYYY-MM-DD
    """
    parcela = get_object_or_404(Parcela, pk=pk)
    data_str = request.GET.get('data', '')
    try:
        from datetime import date as _date
        if data_str:
            data_ref = _date.fromisoformat(data_str)
        else:
            data_ref = timezone.localdate()
    except ValueError:
        return JsonResponse({'erro': 'Data inválida. Use YYYY-MM-DD.'}, status=400)

    juros, multa = parcela.calcular_juros_multa(data_ref)
    valor_total = parcela.valor_atual + juros + multa
    return JsonResponse({
        'valor_original': float(parcela.valor_original),
        'valor_atual': float(parcela.valor_atual),
        'data_vencimento': parcela.data_vencimento.isoformat(),
        'juros': float(juros),
        'multa': float(multa),
        'valor_total': float(valor_total),
    })


@login_required
@require_POST
def gerar_carne(request, contrato_id):
    """
    Gera boletos para multiplas parcelas de um contrato (carne).

    IMPORTANTE: Utiliza o método pode_gerar_boleto() do contrato para
    verificar bloqueio por reajuste. Boletos de parcelas de ciclos
    sem reajuste aplicado serão bloqueados.
    """
    import json

    contrato = get_object_or_404(Contrato, pk=contrato_id)

    try:
        data = json.loads(request.body)
        parcela_ids = data.get('parcelas', [])
        force = data.get('force', False)

        if not parcela_ids:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Nenhuma parcela selecionada'
            }, status=400)

        # Buscar parcelas
        parcelas = Parcela.objects.filter(
            pk__in=parcela_ids,
            contrato=contrato,
            pago=False
        ).order_by('numero_parcela')

        if not parcelas.exists():
            return JsonResponse({
                'sucesso': False,
                'erro': 'Nenhuma parcela valida encontrada'
            }, status=400)

        # =====================================================================
        # LIMITE DE LOTE: apenas até o último boleto do ciclo atual
        # Regra: em lote só é permitido gerar boletos dentro do ciclo atual
        # (último ciclo onde todos os reajustes foram aplicados).
        # Ciclos futuros (data ainda não chegou) só são permitidos individualmente.
        # =====================================================================
        max_parcela_lote = None  # None = sem limite (FIXO ou todos ciclos quitados)
        if not force and contrato.tipo_correcao != 'FIXO':
            from dateutil.relativedelta import relativedelta as _rd
            prazo_lote = contrato.prazo_reajuste_meses or 12
            hoje_lote = timezone.now().date()
            total_ciclos_lote = (contrato.numero_parcelas - 1) // prazo_lote + 1
            for _ciclo in range(2, total_ciclos_lote + 2):
                data_rd = contrato.data_contrato + _rd(months=(_ciclo - 1) * prazo_lote)
                if hoje_lote < data_rd:
                    # Ciclo ainda futuro → lote só até o ciclo anterior
                    max_parcela_lote = (_ciclo - 1) * prazo_lote
                    break
                from financeiro.models import Reajuste as _Reaj
                if not _Reaj.objects.filter(contrato=contrato, ciclo=_ciclo, aplicado=True).exists():
                    # Ciclo vencido mas não reajustado → lote só até o ciclo anterior
                    max_parcela_lote = (_ciclo - 1) * prazo_lote
                    break

        resultados = []
        gerados = 0
        bloqueados = 0
        erros = 0

        for parcela in parcelas:
            # =====================================================================
            # BLOQUEIO DE LOTE: parcelas além do ciclo atual não são permitidas em lote
            # =====================================================================
            if not force and max_parcela_lote is not None and parcela.numero_parcela > max_parcela_lote:
                resultados.append({
                    'parcela_id': parcela.id,
                    'numero_parcela': parcela.numero_parcela,
                    'sucesso': False,
                    'bloqueado_reajuste': True,
                    'erro': (
                        f"Parcela {parcela.numero_parcela} pertence a um ciclo futuro ou com reajuste "
                        f"pendente. Boletos de ciclos futuros só podem ser gerados individualmente."
                    )
                })
                bloqueados += 1
                continue

            # Verificar se ja tem boleto
            if parcela.tem_boleto and not force:
                resultados.append({
                    'parcela_id': parcela.id,
                    'numero_parcela': parcela.numero_parcela,
                    'sucesso': True,
                    'mensagem': 'Boleto ja existente'
                })
                continue

            # Gerar boleto
            try:
                resultado = parcela.gerar_boleto(enviar_email=False)
                if resultado and resultado.get('sucesso'):
                    gerados += 1
                    resultados.append({
                        'parcela_id': parcela.id,
                        'numero_parcela': parcela.numero_parcela,
                        'sucesso': True,
                        'nosso_numero': resultado.get('nosso_numero')
                    })

                    # Atualizar último mês com boleto gerado
                    if hasattr(contrato, 'ultimo_mes_boleto_gerado'):
                        if parcela.numero_parcela > contrato.ultimo_mes_boleto_gerado:
                            contrato.ultimo_mes_boleto_gerado = parcela.numero_parcela
                else:
                    erros += 1
                    resultados.append({
                        'parcela_id': parcela.id,
                        'numero_parcela': parcela.numero_parcela,
                        'sucesso': False,
                        'erro': resultado.get('erro') if resultado else 'Erro desconhecido'
                    })
            except Exception as e:
                logger.exception('Erro ao gerar boleto parcela pk=%s: %s', parcela.id, e)
                erros += 1
                resultados.append({
                    'parcela_id': parcela.id,
                    'numero_parcela': parcela.numero_parcela,
                    'sucesso': False,
                    'erro': str(e)
                })

        # Salvar atualização do último mês com boleto gerado
        if hasattr(contrato, 'ultimo_mes_boleto_gerado'):
            contrato.save(update_fields=['ultimo_mes_boleto_gerado'])

        return JsonResponse({
            'sucesso': True,
            'gerados': gerados,
            'bloqueados': bloqueados,
            'erros': erros,
            'total': len(parcela_ids),
            'resultados': resultados,
            'mensagem': f'{gerados} boletos gerados, {bloqueados} bloqueados por reajuste'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Dados invalidos'
        }, status=400)
    except Exception as e:
        logger.exception(f"Erro ao gerar carne: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
def download_carne_pdf(request, contrato_id):
    """
    Gera e baixa o carnê em PDF para um contrato.

    POST body JSON:
        parcela_ids: list[int]  — IDs das parcelas a incluir no carnê
        apenas_com_boleto: bool — se True, inclui só parcelas com boleto gerado (padrão False)

    GET:  retorna JSON com parcelas disponíveis para carnê

    Retorna: application/pdf (Content-Disposition: attachment)
    """
    import json as _json
    from financeiro.services.carne_service import gerar_carne_pdf

    contrato = get_object_or_404(Contrato, pk=contrato_id)

    if request.method == 'GET':
        # Lista parcelas elegíveis
        parcelas = (
            Parcela.objects.filter(contrato=contrato, pago=False, tipo_parcela='NORMAL')
            .order_by('numero_parcela')
            .values('id', 'numero_parcela', 'data_vencimento', 'valor_atual',
                    'nosso_numero', 'linha_digitavel', 'status_boleto')
        )
        return JsonResponse({
            'parcelas': [
                {
                    **p,
                    'data_vencimento': str(p['data_vencimento']),
                    'valor_atual': str(p['valor_atual']),
                    'tem_boleto': bool(p['nosso_numero']),
                }
                for p in parcelas
            ]
        })

    # POST — gerar PDF
    try:
        body = _json.loads(request.body)
    except _json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido'}, status=400)

    parcela_ids = body.get('parcela_ids', [])
    apenas_com_boleto = body.get('apenas_com_boleto', False)

    if not parcela_ids:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela selecionada'}, status=400)

    if len(parcela_ids) > 60:
        return JsonResponse({'sucesso': False, 'erro': 'Máximo de 60 parcelas por carnê'}, status=400)

    qs = Parcela.objects.filter(
        pk__in=parcela_ids,
        contrato=contrato,
        tipo_parcela='NORMAL',
    ).select_related('contrato__comprador', 'contrato__imovel', 'contrato__imobiliaria').order_by('numero_parcela')

    if apenas_com_boleto:
        qs = qs.exclude(nosso_numero='')

    parcelas = list(qs)
    if not parcelas:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela válida encontrada'}, status=400)

    try:
        pdf_bytes = gerar_carne_pdf(parcelas, contrato)
    except Exception as e:
        logger.exception('Erro ao gerar carnê PDF contrato %s: %s', contrato_id, e)
        return JsonResponse({'sucesso': False, 'erro': f'Erro ao gerar PDF: {e}'}, status=500)

    filename = f"carne_{contrato.numero_contrato}_{len(parcelas)}parcelas.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def download_carne_pdf_multiplos(request):
    """
    Gera carnê PDF consolidado para múltiplos contratos.

    POST body JSON:
        contratos: list[{contrato_id: int, parcela_ids: list[int]}]

    Retorna: application/pdf com todos os carnês concatenados.
    """
    import json as _json
    from financeiro.services.carne_service import gerar_carne_multiplos_contratos

    try:
        body = _json.loads(request.body)
    except _json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido'}, status=400)

    contratos_data = body.get('contratos', [])
    if not contratos_data:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhum contrato informado'}, status=400)

    if len(contratos_data) > 50:
        return JsonResponse({'sucesso': False, 'erro': 'Máximo de 50 contratos por vez'}, status=400)

    contratos_parcelas = []
    for item in contratos_data:
        cid = item.get('contrato_id')
        pids = item.get('parcela_ids', [])
        if not cid or not pids:
            continue
        try:
            contrato = Contrato.objects.select_related(
                'comprador', 'imovel', 'imobiliaria'
            ).get(pk=cid)
        except Contrato.DoesNotExist:
            continue
        parcelas = Parcela.objects.filter(
            pk__in=pids, contrato=contrato, tipo_parcela='NORMAL'
        ).order_by('numero_parcela')
        contratos_parcelas.append({'contrato': contrato, 'parcelas': parcelas})

    if not contratos_parcelas:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhum dado válido encontrado'}, status=400)

    try:
        pdf_bytes = gerar_carne_multiplos_contratos(contratos_parcelas)
    except Exception as e:
        logger.exception('Erro ao gerar carnê multiplos: %s', e)
        return JsonResponse({'sucesso': False, 'erro': f'Erro ao gerar PDF: {e}'}, status=500)

    total = sum(len(list(c['parcelas'])) for c in contratos_parcelas)
    filename = f"carnes_{len(contratos_parcelas)}contratos_{total}parcelas.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_GET
def api_parcelas_elegibilidade(request, contrato_id):
    """
    Retorna informacoes sobre elegibilidade de geracao de boletos para parcelas.

    Regras de elegibilidade por ciclo de reajuste:
    - Ciclo 1 (parcelas 1-12 para prazo_reajuste_meses=12): Sempre liberado
    - Ciclo 2+ (parcelas 13+): Requer reajuste aplicado para o ciclo
    - tipo_correcao='FIXO': Todas as parcelas liberadas (sem reajuste necessario)

    Retorna lista de parcelas com status de elegibilidade e informacoes do ciclo.
    """
    try:
        return _api_parcelas_elegibilidade_logic(request, contrato_id)
    except Exception as exc:
        import logging
        logging.getLogger(__name__).exception('api_parcelas_elegibilidade erro contrato_id=%s', contrato_id)
        return JsonResponse({'sucesso': False, 'erro': str(exc)}, status=500)


def _api_parcelas_elegibilidade_logic(request, contrato_id):
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    # Obter todas as parcelas nao pagas
    parcelas = contrato.parcelas.filter(pago=False).order_by('numero_parcela')

    # Verificar se o tipo de correcao e FIXO (sem reajuste necessario)
    tipo_correcao_fixo = contrato.tipo_correcao == 'FIXO'

    # Informacoes do ciclo atual
    prazo_reajuste = contrato.prazo_reajuste_meses or 12
    ciclo_atual = contrato.ciclo_reajuste_atual or 1

    # Reajustes aplicados
    from financeiro.models import Reajuste
    reajustes_aplicados = set(
        Reajuste.objects.filter(contrato=contrato, aplicado=True)
        .values_list('ciclo', flat=True)
    )

    # Calcular maximo de parcelas disponiveis para geracao
    parcelas_disponiveis = []
    primeiro_ciclo_bloqueado = None
    total_disponiveis = 0
    total_bloqueadas = 0

    for parcela in parcelas:
        ciclo_parcela = contrato.calcular_ciclo_parcela(parcela.numero_parcela)

        # Determinar elegibilidade
        if tipo_correcao_fixo:
            # FIXO: todas liberadas
            pode_gerar = True
            motivo = "Indice FIXO - sem reajuste necessario"
        elif ciclo_parcela == 1:
            # Primeiro ciclo: sempre liberado
            pode_gerar = True
            motivo = "Primeiro ciclo - liberado"
        elif ciclo_parcela in reajustes_aplicados:
            # Reajuste ja aplicado para este ciclo
            pode_gerar = True
            motivo = f"Reajuste do ciclo {ciclo_parcela} aplicado"
        else:
            # Reajuste pendente
            pode_gerar = False
            motivo = f"Aguardando reajuste do ciclo {ciclo_parcela}"
            if primeiro_ciclo_bloqueado is None:
                primeiro_ciclo_bloqueado = ciclo_parcela

        if pode_gerar:
            total_disponiveis += 1
        else:
            total_bloqueadas += 1

        parcelas_disponiveis.append({
            'id': parcela.id,
            'numero_parcela': parcela.numero_parcela,
            'data_vencimento': parcela.data_vencimento.strftime('%d/%m/%Y'),
            'valor_atual': float(parcela.valor_atual),
            'ciclo': ciclo_parcela,
            'pode_gerar': pode_gerar,
            'tem_boleto': parcela.tem_boleto,
            'motivo': motivo
        })

    # Calcular informacoes do proximo ciclo de reajuste
    proximo_ciclo_reajuste = None
    if primeiro_ciclo_bloqueado:
        parcela_inicio_ciclo = ((primeiro_ciclo_bloqueado - 1) * prazo_reajuste) + 1
        parcela_fim_ciclo = primeiro_ciclo_bloqueado * prazo_reajuste
        proximo_ciclo_reajuste = {
            'ciclo': primeiro_ciclo_bloqueado,
            'parcela_inicial': parcela_inicio_ciclo,
            'parcela_final': min(parcela_fim_ciclo, contrato.numero_parcelas),
            'mensagem': f"Aplique o reajuste para liberar as parcelas {parcela_inicio_ciclo} a {min(parcela_fim_ciclo, contrato.numero_parcelas)}"
        }

    return JsonResponse({
        'sucesso': True,
        'contrato_id': contrato.id,
        'tipo_correcao': contrato.tipo_correcao,
        'tipo_correcao_fixo': tipo_correcao_fixo,
        'prazo_reajuste_meses': prazo_reajuste,
        'ciclo_atual': ciclo_atual,
        'reajustes_aplicados': list(reajustes_aplicados),
        'total_parcelas_pendentes': parcelas.count(),
        'total_disponiveis': total_disponiveis,
        'total_bloqueadas': total_bloqueadas,
        'proximo_ciclo_reajuste': proximo_ciclo_reajuste,
        'parcelas': parcelas_disponiveis
    })


@login_required
@require_POST
def api_gerar_boletos_parcelas(request):
    """
    Gera boletos para uma lista de parcelas selecionadas individualmente.

    POST JSON: {"parcela_ids": [1, 2, 3], "force": false}
    """
    import json
    try:
        data = json.loads(request.body)
        parcela_ids = [int(x) for x in data.get('parcela_ids', []) if str(x).isdigit()]
        force = bool(data.get('force', False))
    except (ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=400)

    if not parcela_ids:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela selecionada'}, status=400)

    gerados, bloqueados, erros = 0, 0, 0
    detalhes = []

    for parcela_id in parcela_ids:
        try:
            parcela = Parcela.objects.select_related(
                'contrato', 'contrato__imovel', 'contrato__imovel__imobiliaria'
            ).get(pk=parcela_id)
        except Parcela.DoesNotExist:
            erros += 1
            detalhes.append({'parcela_id': parcela_id, 'sucesso': False, 'erro': 'Parcela não encontrada'})
            continue

        if parcela.pago:
            detalhes.append({'parcela_id': parcela_id, 'sucesso': False, 'erro': 'Parcela já paga'})
            erros += 1
            continue

        contrato = parcela.contrato
        if not force and hasattr(contrato, 'pode_gerar_boleto'):
            pode_gerar, motivo = contrato.pode_gerar_boleto(parcela.numero_parcela)
            if not pode_gerar:
                bloqueados += 1
                detalhes.append({'parcela_id': parcela_id, 'sucesso': False, 'bloqueado': True, 'erro': motivo})
                continue

        # Obter conta bancária
        conta_bancaria = None
        try:
            imobiliaria = contrato.imovel.imobiliaria
            conta_bancaria = imobiliaria.contas_bancarias.filter(principal=True, ativo=True).first()
        except Exception:
            pass

        if not conta_bancaria:
            erros += 1
            detalhes.append({'parcela_id': parcela_id, 'sucesso': False, 'erro': 'Nenhuma conta bancária principal configurada'})
            continue

        try:
            resultado = parcela.gerar_boleto(conta_bancaria, force=force, enviar_email=False)
            if resultado and resultado.get('sucesso'):
                gerados += 1
                detalhes.append({
                    'parcela_id': parcela_id,
                    'sucesso': True,
                    'nosso_numero': resultado.get('nosso_numero', ''),
                })
            else:
                erros += 1
                detalhes.append({
                    'parcela_id': parcela_id, 'sucesso': False,
                    'erro': resultado.get('erro', 'Erro desconhecido') if resultado else 'Sem resposta'})
        except Exception as e:
            erros += 1
            logger.exception("Erro ao gerar boleto parcela %s: %s", parcela_id, e)
            detalhes.append({'parcela_id': parcela_id, 'sucesso': False, 'erro': str(e)})

    return JsonResponse({
        'sucesso': True,
        'gerados': gerados,
        'bloqueados': bloqueados,
        'erros': erros,
        'detalhes': detalhes,
    })


@login_required
@require_POST
def api_gerar_boletos_lote(request):
    """
    Gera boletos em lote para multiplos contratos.

    Recebe:
    - contratos: lista de IDs de contratos
    - quantidade: quantidade de boletos por contrato (opcional, padrao=1)
    - force: ignorar bloqueio de reajuste (opcional, padrao=False)

    Respeita a logica de reajuste:
    - Gera apenas parcelas cujo ciclo ja teve reajuste aplicado
    - tipo_correcao='FIXO' permite gerar todas as parcelas
    """
    import json

    try:
        data = json.loads(request.body)
        contrato_ids = data.get('contratos', [])
        quantidade = int(data.get('quantidade', 1))
        force = data.get('force', False)

        if not contrato_ids:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Nenhum contrato selecionado'
            }, status=400)

        if quantidade < 1:
            quantidade = 1

        resultados = []
        total_gerados = 0
        total_bloqueados = 0
        total_erros = 0

        for contrato_id in contrato_ids:
            try:
                contrato = Contrato.objects.get(pk=contrato_id, status='ATIVO')
            except Contrato.DoesNotExist:
                resultados.append({
                    'contrato_id': contrato_id,
                    'sucesso': False,
                    'erro': 'Contrato nao encontrado ou inativo'
                })
                total_erros += 1
                continue

            # Obter parcelas elegiveis (nao pagas, sem boleto, elegíveis por reajuste)
            parcelas = contrato.parcelas.filter(
                pago=False,
                status_boleto='NAO_GERADO'
            ).order_by('numero_parcela')

            gerados_contrato = 0
            bloqueados_contrato = 0
            parcelas_resultado = []

            for parcela in parcelas[:quantidade + 10]:  # Pegar algumas extras caso algumas sejam bloqueadas
                if gerados_contrato >= quantidade:
                    break

                # Verificar elegibilidade por reajuste
                if not force:
                    pode_gerar, motivo = contrato.pode_gerar_boleto(parcela.numero_parcela)
                    if not pode_gerar:
                        bloqueados_contrato += 1
                        parcelas_resultado.append({
                            'parcela_id': parcela.id,
                            'numero_parcela': parcela.numero_parcela,
                            'sucesso': False,
                            'bloqueado': True,
                            'motivo': motivo
                        })
                        continue

                # Gerar boleto
                try:
                    resultado = parcela.gerar_boleto(enviar_email=False)
                    if resultado and resultado.get('sucesso'):
                        gerados_contrato += 1
                        total_gerados += 1

                        # Atualizar ultimo mes com boleto gerado
                        if parcela.numero_parcela > contrato.ultimo_mes_boleto_gerado:
                            contrato.ultimo_mes_boleto_gerado = parcela.numero_parcela

                        parcelas_resultado.append({
                            'parcela_id': parcela.id,
                            'numero_parcela': parcela.numero_parcela,
                            'sucesso': True,
                            'nosso_numero': resultado.get('nosso_numero')
                        })
                    else:
                        total_erros += 1
                        parcelas_resultado.append({
                            'parcela_id': parcela.id,
                            'numero_parcela': parcela.numero_parcela,
                            'sucesso': False,
                            'erro': resultado.get('erro') if resultado else 'Erro desconhecido'
                        })
                except Exception as e:
                    logger.exception('Erro ao gerar boleto parcela pk=%s em lote: %s', parcela.id, e)
                    total_erros += 1
                    parcelas_resultado.append({
                        'parcela_id': parcela.id,
                        'numero_parcela': parcela.numero_parcela,
                        'sucesso': False,
                        'erro': str(e)
                    })

            # Salvar atualizacao do contrato
            if gerados_contrato > 0:
                contrato.save(update_fields=['ultimo_mes_boleto_gerado'])

            total_bloqueados += bloqueados_contrato

            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': gerados_contrato > 0,
                'gerados': gerados_contrato,
                'bloqueados': bloqueados_contrato,
                'parcelas': parcelas_resultado
            })

        return JsonResponse({
            'sucesso': True,
            'total_contratos': len(contrato_ids),
            'total_gerados': total_gerados,
            'total_bloqueados': total_bloqueados,
            'total_erros': total_erros,
            'resultados': resultados,
            'mensagem': f'{total_gerados} boletos gerados em {len(contrato_ids)} contrato(s)'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Dados invalidos'
        }, status=400)
    except Exception as e:
        logger.exception(f"Erro ao gerar boletos em lote: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


# =============================================================================
# REAJUSTES DE CONTRATO
# =============================================================================

@login_required
def preview_reajuste_contrato(request, contrato_id):
    """
    Dry-run: calcula o reajuste para o ciclo pendente sem persistir nada.

    GET  → retorna ciclo pendente, índice, período de referência, % acumulado
           e tabela detalhada por parcela (valor atual → valor novo).

    Parâmetros GET opcionais:
      - ciclo              : forçar ciclo específico (padrão: ciclo pendente)
      - desconto_percentual: desconto em p.p. sobre o índice
      - desconto_valor     : desconto fixo em R$ por parcela
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    try:
        ciclo = request.GET.get('ciclo')
        if ciclo:
            ciclo = int(ciclo)
        else:
            ciclo = Reajuste.calcular_ciclo_pendente(contrato)

        if not ciclo:
            return JsonResponse({
                'sucesso': False,
                'ciclo_pendente': None,
                'mensagem': 'Nenhum reajuste pendente para este contrato.',
            })

        desconto_percentual = request.GET.get('desconto_percentual') or None
        desconto_valor = request.GET.get('desconto_valor') or None

        resultado = Reajuste.preview_reajuste(
            contrato, ciclo,
            desconto_percentual=desconto_percentual,
            desconto_valor=desconto_valor,
        )

        if 'erro' in resultado:
            return JsonResponse({
                'sucesso': False,
                'ciclo_pendente': ciclo,
                'erro': resultado['erro'],
                'indice_tipo': resultado.get('indice_tipo', ''),
                'periodo_referencia_inicio': resultado.get('periodo_referencia_inicio', ''),
                'periodo_referencia_fim': resultado.get('periodo_referencia_fim', ''),
                'parcela_inicial': resultado.get('parcela_inicial', ''),
                'parcela_final': resultado.get('parcela_final', ''),
            })

        # Serializar datas e Decimals para JSON
        parcelas_json = []
        for p in resultado['parcelas']:
            parcelas_json.append({
                'numero_parcela': p['numero_parcela'],
                'data_vencimento': p['data_vencimento'].strftime('%d/%m/%Y'),
                'valor_atual': float(p['valor_atual']),
                'valor_novo': float(p['valor_novo']),
                'diferenca': float(p['diferenca']),
                'tem_boleto': p['tem_boleto'],
            })

        return JsonResponse({
            'sucesso': True,
            'ciclo_pendente': ciclo,
            'indice_tipo': resultado['indice_tipo'],
            'periodo_referencia_inicio': resultado['periodo_referencia_inicio'].strftime('%b/%Y'),
            'periodo_referencia_fim': resultado['periodo_referencia_fim'].strftime('%b/%Y'),
            'percentual_bruto': float(resultado['percentual_bruto']),
            'spread': float(resultado['spread']),
            'percentual_bruto_com_spread': float(resultado['percentual_bruto_com_spread']),
            'desconto_percentual': float(resultado['desconto_percentual']),
            'desconto_valor': float(resultado['desconto_valor']),
            'percentual_liquido': float(resultado['percentual_liquido']),
            'percentual_final': float(resultado['percentual_final']),
            'piso': float(resultado['piso']) if resultado['piso'] is not None else None,
            'teto': float(resultado['teto']) if resultado['teto'] is not None else None,
            'piso_ativado': resultado['piso_ativado'],
            'teto_ativado': resultado['teto_ativado'],
            'parcela_inicial': resultado['parcela_inicial'],
            'parcela_final': resultado['parcela_final'],
            'parcelas': parcelas_json,
            'total_parcelas': resultado['total_parcelas'],
            'valor_anterior_total': float(resultado['valor_anterior_total']),
            'valor_novo_total': float(resultado['valor_novo_total']),
            'diferenca_total': float(resultado['diferenca_total']),
            'boletos_emitidos': resultado['boletos_emitidos'],
        })

    except Exception as e:
        logger.exception(f"Erro no preview de reajuste: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
def aplicar_reajuste_pagina(request, contrato_id):
    """
    Página dedicada para aplicar reajuste em um contrato.

    GET  → calcula o preview server-side e exibe o formulário.
           Query params opcionais: desconto_percentual, desconto_valor, modal=1
    POST → aplica o reajuste e redireciona para o detalhe do contrato.
           Se modal=1 ou X-Requested-With=XMLHttpRequest → retorna JSON.
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    is_modal = (
        request.GET.get('modal') == '1'
        or request.POST.get('modal') == '1'
        or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    )

    def get_client_ip(req):
        x_forwarded = req.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded:
            return x_forwarded.split(',')[0].strip()
        return req.META.get('REMOTE_ADDR')

    if request.method == 'POST':
        ciclo_param = request.POST.get('ciclo')
        desconto_percentual = request.POST.get('desconto_percentual') or None
        desconto_valor = request.POST.get('desconto_valor') or None
        observacoes = request.POST.get('observacoes', '')

        try:
            if not ciclo_param:
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': 'Ciclo de reajuste não informado.'}, status=400)
                messages.error(request, 'Ciclo de reajuste não informado.')
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            ciclo = int(ciclo_param)

            # V-09: Validar sequência obrigatória de ciclos (2.9 — não pular ciclos)
            ciclo_esperado = Reajuste.calcular_ciclo_pendente(contrato)
            if ciclo_esperado is None:
                erro_seq = 'Não há ciclos de reajuste pendentes para este contrato.'
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': erro_seq}, status=400)
                messages.error(request, erro_seq)
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)
            if ciclo > ciclo_esperado:
                erro_seq = (
                    f'Sequência incorreta: o ciclo {ciclo_esperado} ainda não foi aplicado. '
                    f'Execute os reajustes em ordem — aplique o ciclo {ciclo_esperado} antes do {ciclo}.'
                )
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': erro_seq}, status=400)
                messages.error(request, erro_seq)
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)
            if ciclo < ciclo_esperado:
                erro_seq = f'Ciclo {ciclo} já foi aplicado anteriormente.'
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': erro_seq}, status=400)
                messages.error(request, erro_seq)
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            preview = Reajuste.preview_reajuste(
                contrato, ciclo,
                desconto_percentual=desconto_percentual,
                desconto_valor=desconto_valor,
            )

            if 'erro' in preview:
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': preview['erro']}, status=400)
                messages.error(request, preview['erro'])
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            if not contrato.parcelas.filter(
                numero_parcela__gte=preview['parcela_inicial'],
                numero_parcela__lte=preview['parcela_final'],
                pago=False,
            ).exists():
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela pendente no intervalo selecionado.'}, status=400)
                messages.error(request, 'Nenhuma parcela pendente no intervalo selecionado.')
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            reajuste = Reajuste.objects.create(
                contrato=contrato,
                data_reajuste=timezone.now().date(),
                indice_tipo=preview['indice_tipo'],
                percentual=preview['percentual_final'],
                percentual_bruto=preview['percentual_bruto'],
                spread_aplicado=preview['spread'] if preview.get('spread') else None,
                desconto_percentual=Decimal(str(desconto_percentual)) if desconto_percentual else None,
                desconto_valor=Decimal(str(desconto_valor)) if desconto_valor else None,
                piso_aplicado=preview.get('piso'),
                teto_aplicado=preview.get('teto'),
                parcela_inicial=preview['parcela_inicial'],
                parcela_final=preview['parcela_final'],
                ciclo=ciclo,
                periodo_referencia_inicio=preview['periodo_referencia_inicio'],
                periodo_referencia_fim=preview['periodo_referencia_fim'],
                aplicado_manual=True,
                usuario=request.user,
                ip_address=get_client_ip(request),
                observacoes=observacoes,
            )

            resultado = reajuste.aplicar_reajuste()
            msg = (
                f'Reajuste do ciclo {ciclo} ({preview["percentual_final"]:,.4f}%) '
                f'aplicado em {resultado["parcelas_reajustadas"]} parcela(s).'
            )
            if is_modal:
                return JsonResponse({'sucesso': True, 'mensagem': msg})
            messages.success(request, msg)
            return redirect('contratos:detalhe', pk=contrato_id)

        except Exception as e:
            logger.exception(f'Erro ao aplicar reajuste: {e}')
            if is_modal:
                return JsonResponse({'sucesso': False, 'erro': str(e)}, status=400)
            messages.error(request, f'Erro ao aplicar reajuste: {e}')
            return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

    # ── GET ──────────────────────────────────────────────────────────────────
    ciclo = Reajuste.calcular_ciclo_pendente(contrato)

    if not ciclo:
        # Verificar se o contrato sequer tem parcelas no ciclo 2
        prazo = contrato.prazo_reajuste_meses or 12
        reajuste_nao_aplicavel = (
            contrato.tipo_correcao != 'FIXO'
            and (contrato.numero_parcelas or 0) <= prazo
        )
        context = {
            'contrato': contrato,
            'sem_pendente': True,
            'reajuste_nao_aplicavel': reajuste_nao_aplicavel,
            'proximo_reajuste': contrato.data_proximo_reajuste,
        }
        tpl = 'financeiro/aplicar_reajuste_partial.html' if is_modal else 'financeiro/aplicar_reajuste.html'
        return render(request, tpl, context)

    desconto_percentual = request.GET.get('desconto_percentual') or None
    desconto_valor = request.GET.get('desconto_valor') or None

    preview = Reajuste.preview_reajuste(
        contrato, ciclo,
        desconto_percentual=desconto_percentual,
        desconto_valor=desconto_valor,
    )

    # Calcular histórico de reajustes já aplicados
    reajustes_anteriores = Reajuste.objects.filter(
        contrato=contrato, aplicado=True
    ).order_by('ciclo')

    context = {
        'contrato': contrato,
        'ciclo': ciclo,
        'preview': preview,
        'desconto_percentual': desconto_percentual or '',
        'desconto_valor': desconto_valor or '',
        'reajustes_anteriores': reajustes_anteriores,
        'sem_pendente': False,
        'tem_erro': 'erro' in preview,
        'is_modal': is_modal,
    }
    tpl = 'financeiro/aplicar_reajuste_partial.html' if is_modal else 'financeiro/aplicar_reajuste.html'
    return render(request, tpl, context)


@login_required
@require_POST
def aplicar_reajuste_contrato(request, contrato_id):
    """
    Página dedicada para aplicar reajuste em um contrato.

    GET  → calcula o preview server-side e exibe o formulário.
           Query params opcionais: desconto_percentual, desconto_valor, modal=1
    POST → aplica o reajuste e redireciona para o detalhe do contrato.
           Se modal=1 ou X-Requested-With=XMLHttpRequest → retorna JSON.
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    is_modal = (
        request.GET.get('modal') == '1'
        or request.POST.get('modal') == '1'
        or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    )

    # V-08: Contratos FIXO não têm reajuste periódico
    if contrato.tipo_correcao == 'FIXO':
        return JsonResponse({
            'sucesso': False,
            'erro': 'Contratos com índice FIXO são pré-fixados e não possuem reajuste periódico.'
        }, status=400)

    # Capturar IP do usuário
    def get_client_ip(req):
        x_forwarded = req.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded:
            return x_forwarded.split(',')[0].strip()
        return req.META.get('REMOTE_ADDR')

    if request.method == 'POST':
        ciclo_param = request.POST.get('ciclo')
        desconto_percentual = request.POST.get('desconto_percentual') or None
        desconto_valor = request.POST.get('desconto_valor') or None
        observacoes = request.POST.get('observacoes', '')

        try:
            if not ciclo_param:
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': 'Ciclo de reajuste não informado.'}, status=400)
                messages.error(request, 'Ciclo de reajuste não informado.')
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            ciclo = int(ciclo_param)

            # V-09: Validar sequência obrigatória de ciclos (2.9 — não pular ciclos)
            ciclo_esperado = Reajuste.calcular_ciclo_pendente(contrato)
            if ciclo_esperado is None:
                erro_seq = 'Não há ciclos de reajuste pendentes para este contrato.'
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': erro_seq}, status=400)
                messages.error(request, erro_seq)
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)
            if ciclo > ciclo_esperado:
                erro_seq = (
                    f'Sequência incorreta: o ciclo {ciclo_esperado} ainda não foi aplicado. '
                    f'Execute os reajustes em ordem — aplique o ciclo {ciclo_esperado} antes do {ciclo}.'
                )
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': erro_seq}, status=400)
                messages.error(request, erro_seq)
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)
            if ciclo < ciclo_esperado:
                erro_seq = f'Ciclo {ciclo} já foi aplicado anteriormente.'
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': erro_seq}, status=400)
                messages.error(request, erro_seq)
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            preview = Reajuste.preview_reajuste(
                contrato, ciclo,
                desconto_percentual=desconto_percentual,
                desconto_valor=desconto_valor,
            )

            if 'erro' in preview:
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': preview['erro']}, status=400)
                messages.error(request, preview['erro'])
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            if not contrato.parcelas.filter(
                numero_parcela__gte=preview['parcela_inicial'],
                numero_parcela__lte=preview['parcela_final'],
                pago=False,
            ).exists():
                if is_modal:
                    return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela pendente no intervalo selecionado.'}, status=400)
                messages.error(request, 'Nenhuma parcela pendente no intervalo selecionado.')
                return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

            reajuste = Reajuste.objects.create(
                contrato=contrato,
                data_reajuste=timezone.now().date(),
                indice_tipo=preview['indice_tipo'],
                percentual=preview['percentual_final'],
                percentual_bruto=preview['percentual_bruto'],
                spread_aplicado=preview['spread'] if preview.get('spread') else None,
                desconto_percentual=Decimal(str(desconto_percentual)) if desconto_percentual else None,
                desconto_valor=Decimal(str(desconto_valor)) if desconto_valor else None,
                piso_aplicado=preview.get('piso'),
                teto_aplicado=preview.get('teto'),
                parcela_inicial=preview['parcela_inicial'],
                parcela_final=preview['parcela_final'],
                ciclo=ciclo,
                periodo_referencia_inicio=preview['periodo_referencia_inicio'],
                periodo_referencia_fim=preview['periodo_referencia_fim'],
                aplicado_manual=True,
                usuario=request.user,
                ip_address=get_client_ip(request),
                observacoes=observacoes,
            )

            resultado = reajuste.aplicar_reajuste()
            msg = (
                f'Reajuste do ciclo {ciclo} ({preview["percentual_final"]:,.4f}%) '
                f'aplicado em {resultado["parcelas_reajustadas"]} parcela(s).'
            )
            if is_modal:
                return JsonResponse({'sucesso': True, 'mensagem': msg})
            messages.success(request, msg)
            return redirect('contratos:detalhe', pk=contrato_id)

        except Exception as e:
            logger.exception(f'Erro ao aplicar reajuste: {e}')
            if is_modal:
                return JsonResponse({'sucesso': False, 'erro': str(e)}, status=400)
            messages.error(request, f'Erro ao aplicar reajuste: {e}')
            return redirect('financeiro:aplicar_reajuste', contrato_id=contrato_id)

    # ── GET ──────────────────────────────────────────────────────────────────
    ciclo = Reajuste.calcular_ciclo_pendente(contrato)

    if not ciclo:
        context = {
            'contrato': contrato,
            'sem_pendente': True,
            'proximo_reajuste': contrato.data_proximo_reajuste,
        }
        tpl = 'financeiro/aplicar_reajuste_partial.html' if is_modal else 'financeiro/aplicar_reajuste.html'
        return render(request, tpl, context)

    desconto_percentual = request.GET.get('desconto_percentual') or None
    desconto_valor = request.GET.get('desconto_valor') or None

    preview = Reajuste.preview_reajuste(
        contrato, ciclo,
        desconto_percentual=desconto_percentual,
        desconto_valor=desconto_valor,
    )

    # Calcular histórico de reajustes já aplicados
    reajustes_anteriores = Reajuste.objects.filter(
        contrato=contrato, aplicado=True
    ).order_by('ciclo')

    context = {
        'contrato': contrato,
        'ciclo': ciclo,
        'preview': preview,
        'desconto_percentual': desconto_percentual or '',
        'desconto_valor': desconto_valor or '',
        'reajustes_anteriores': reajustes_anteriores,
        'sem_pendente': False,
        'tem_erro': 'erro' in preview,
        'is_modal': is_modal,
    }
    tpl = 'financeiro/aplicar_reajuste_partial.html' if is_modal else 'financeiro/aplicar_reajuste.html'
    return render(request, tpl, context)


@login_required
@require_POST
def aplicar_reajuste_contrato(request, contrato_id):  # noqa: F811
    """
    Aplica o reajuste nas parcelas de um contrato.

    Modo automático (recomendado): passa apenas ciclo + desconto opcionais.
      O sistema determina índice, período de referência, % acumulado e parcelas.

    Modo manual (legado): passa indice_tipo + percentual + parcela_inicial/final.

    JSON body:
      - ciclo              : ciclo a aplicar (se omitido, usa o ciclo pendente)
      - desconto_percentual: desconto em p.p. sobre o índice (opcional)
      - desconto_valor     : desconto fixo em R$ por parcela (opcional)
      - observacoes        : texto livre
      --- parâmetros manuais (só usados se ciclo não for informado) ---
      - indice_tipo, percentual, parcela_inicial, parcela_final
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    # V-08: Contratos FIXO não têm reajuste periódico
    if contrato.tipo_correcao == 'FIXO':
        return JsonResponse({
            'sucesso': False,
            'erro': 'Contratos com índice FIXO são pré-fixados e não possuem reajuste periódico.'
        }, status=400)

    # Capturar IP do usuário
    def get_client_ip(req):
        x_forwarded = req.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded:
            return x_forwarded.split(',')[0].strip()
        return req.META.get('REMOTE_ADDR')

    try:
        import json
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST

        observacoes = data.get('observacoes', '')
        desconto_percentual = data.get('desconto_percentual') or None
        desconto_valor = data.get('desconto_valor') or None

        ciclo_param = data.get('ciclo')

        # ── Modo automático ──────────────────────────────────────────────────
        if ciclo_param:
            ciclo = int(ciclo_param)
            preview = Reajuste.preview_reajuste(
                contrato, ciclo,
                desconto_percentual=desconto_percentual,
                desconto_valor=desconto_valor,
            )
            if 'erro' in preview:
                return JsonResponse({'sucesso': False, 'erro': preview['erro']}, status=400)

            reajuste = Reajuste.objects.create(
                contrato=contrato,
                data_reajuste=timezone.now().date(),
                indice_tipo=preview['indice_tipo'],
                percentual=preview['percentual_final'],
                percentual_bruto=preview['percentual_bruto'],
                spread_aplicado=preview['spread'] if preview['spread'] else None,
                desconto_percentual=Decimal(str(desconto_percentual)) if desconto_percentual else None,
                desconto_valor=Decimal(str(desconto_valor)) if desconto_valor else None,
                piso_aplicado=preview['piso'],
                teto_aplicado=preview['teto'],
                parcela_inicial=preview['parcela_inicial'],
                parcela_final=preview['parcela_final'],
                ciclo=ciclo,
                periodo_referencia_inicio=preview['periodo_referencia_inicio'],
                periodo_referencia_fim=preview['periodo_referencia_fim'],
                aplicado_manual=True,
                usuario=request.user,
                ip_address=get_client_ip(request),
                observacoes=observacoes,
            )

        # ── Modo manual (legado) ─────────────────────────────────────────────
        else:
            indice_tipo = data.get('indice_tipo', 'MANUAL')
            percentual = Decimal(str(data.get('percentual', 0)))
            parcela_inicial = int(data.get('parcela_inicial', 1))
            parcela_final = int(data.get('parcela_final', contrato.numero_parcelas))

            if percentual == 0:
                return JsonResponse({'sucesso': False, 'erro': 'O percentual de reajuste nao pode ser zero'}, status=400)
            if parcela_inicial < 1 or parcela_final > contrato.numero_parcelas:
                return JsonResponse({'sucesso': False, 'erro': 'Numeros de parcela invalidos'}, status=400)
            if parcela_inicial > parcela_final:
                return JsonResponse({'sucesso': False, 'erro': 'Parcela inicial deve ser menor ou igual a parcela final'}, status=400)

            ciclo = contrato.ciclo_reajuste_atual + 1

            reajuste = Reajuste.objects.create(
                contrato=contrato,
                data_reajuste=timezone.now().date(),
                indice_tipo=indice_tipo,
                percentual=percentual,
                percentual_bruto=percentual,
                desconto_percentual=Decimal(str(desconto_percentual)) if desconto_percentual else None,
                desconto_valor=Decimal(str(desconto_valor)) if desconto_valor else None,
                parcela_inicial=parcela_inicial,
                parcela_final=parcela_final,
                ciclo=ciclo,
                aplicado_manual=True,
                usuario=request.user,
                ip_address=get_client_ip(request),
                observacoes=observacoes,
            )

        if not contrato.parcelas.filter(
            numero_parcela__gte=reajuste.parcela_inicial,
            numero_parcela__lte=reajuste.parcela_final,
            pago=False,
        ).exists():
            reajuste.delete()
            return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela pendente no intervalo selecionado'}, status=400)

        resultado = reajuste.aplicar_reajuste()

        return JsonResponse({
            'sucesso': True,
            'mensagem': (
                f'Reajuste de {reajuste.percentual}% aplicado com sucesso em '
                f'{resultado["parcelas_reajustadas"]} parcela(s)'
            ),
            'reajuste_id': reajuste.id,
            'parcelas_afetadas': resultado['parcelas_reajustadas'],
            'ciclo': reajuste.ciclo,
            'percentual_bruto': float(reajuste.percentual_bruto or reajuste.percentual),
            'percentual_liquido': float(reajuste.percentual),
        })

    except ValueError as e:
        return JsonResponse({'sucesso': False, 'erro': f'Valor invalido: {str(e)}'}, status=400)
    except Exception as e:
        logger.exception(f"Erro ao aplicar reajuste: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
def reajustes_pendentes(request):
    """
    Lista todos os contratos ativos com ciclo de reajuste pendente.

    Detecção mês-a-mês (não por dia exato): exibe contratos a partir de
    1 mês antes do aniversário, permitindo aplicar o reajuste assim que
    o índice de referência estiver disponível.

    Para cada contrato calcula: percentual do índice, prestação atual do
    ciclo e estimativa da prestação nova.
    """
    from contratos.models import Contrato as ContratoModel, IndiceReajuste, TabelaJurosContrato

    contratos_ativos = ContratoModel.objects.filter(
        status='ATIVO'
    ).select_related('comprador', 'imobiliaria', 'imovel').order_by(
        'imobiliaria__nome', 'data_contrato'
    )

    pendentes = []
    for contrato in contratos_ativos:
        ciclo = Reajuste.calcular_ciclo_pendente(contrato)
        if ciclo is None:
            continue

        inicio_ref, fim_ref = Reajuste.calcular_periodo_referencia(contrato, ciclo)
        prazo = contrato.prazo_reajuste_meses
        parcela_inicial = (ciclo - 1) * prazo + 1
        parcela_final = min(ciclo * prazo, contrato.numero_parcelas)

        # Busca o índice acumulado do período de referência
        indice_tipo = contrato.tipo_correcao
        percentual = IndiceReajuste.get_acumulado_periodo(
            indice_tipo,
            inicio_ref.year, inicio_ref.month,
            fim_ref.year, fim_ref.month,
        )
        # Tenta fallback se configurado e índice principal sem dados
        if percentual is None and contrato.tipo_correcao_fallback:
            percentual = IndiceReajuste.get_acumulado_periodo(
                contrato.tipo_correcao_fallback,
                inicio_ref.year, inicio_ref.month,
                fim_ref.year, fim_ref.month,
            )
            if percentual is not None:
                indice_tipo = contrato.tipo_correcao_fallback

        # Spread
        taxa_tabela = TabelaJurosContrato.get_juros_para_ciclo(contrato, ciclo)
        usa_price = taxa_tabela is not None
        spread = taxa_tabela if usa_price else (contrato.spread_reajuste or Decimal('0'))

        percentual_final = None
        if percentual is not None:
            perc_com_spread = percentual + spread
            piso = contrato.reajuste_piso or Decimal('-100')
            teto = contrato.reajuste_teto or Decimal('999')
            percentual_final = max(piso, min(teto, perc_com_spread))

        # Prestação atual: primeiro valor_atual do ciclo (parcela NORMAL não paga preferencial)
        from financeiro.models import Parcela as ParcelaModel  # evita circular
        parcela_ref = ParcelaModel.objects.filter(
            contrato=contrato,
            numero_parcela=parcela_inicial,
            tipo_parcela='NORMAL',
        ).values('valor_atual').first()
        prestacao_atual = parcela_ref['valor_atual'] if parcela_ref else None

        # Estimativa da prestação nova (somente modo SIMPLES — Price recalcula PMT)
        prestacao_nova = None
        if prestacao_atual and percentual_final is not None and not usa_price:
            prestacao_nova = prestacao_atual * (1 + percentual_final / 100)

        pendentes.append({
            'contrato': contrato,
            'ciclo': ciclo,
            'parcela_inicial': parcela_inicial,
            'parcela_final': parcela_final,
            'periodo_referencia_inicio': inicio_ref,
            'periodo_referencia_fim': fim_ref,
            'indice_tipo': indice_tipo,
            'percentual': percentual_final,        # None = sem dados ainda
            'prestacao_atual': prestacao_atual,
            'prestacao_nova': prestacao_nova,
            'usa_price': usa_price,
            'dados_disponiveis': percentual_final is not None,
        })

    # Paginação
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = min(int(per_page), 100)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(pendentes, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    from itertools import groupby
    pendentes_agrupados = []
    for imob_nome, grupo in groupby(list(page_obj), key=lambda x: x['contrato'].imobiliaria.nome):
        pendentes_agrupados.append({
            'imobiliaria': imob_nome,
            'contratos': list(grupo),
        })

    context = {
        'pendentes_agrupados': pendentes_agrupados,
        'total_pendentes': len(pendentes),
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
    }
    return render(request, 'financeiro/reajustes_pendentes.html', context)


@login_required
@require_POST
def aplicar_reajuste_lote(request):
    """
    Aplica reajuste automático em múltiplos contratos de uma vez.

    Corpo JSON:
      - contrato_ids: lista de IDs de contratos (int[])
      - desconto_percentual: desconto em p.p. aplicado a todos (opcional)
      - desconto_valor: desconto em R$ por parcela aplicado a todos (opcional)
      - observacoes: texto livre (opcional)

    Retorna JSON com resumo por contrato: sucesso, erro, parcelas_reajustadas.
    """
    import json

    try:
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
    except (json.JSONDecodeError, Exception):
        return JsonResponse({'sucesso': False, 'erro': 'Corpo da requisição inválido'}, status=400)

    contrato_ids = data.get('contrato_ids', [])
    if not contrato_ids:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhum contrato selecionado'}, status=400)

    desconto_percentual = data.get('desconto_percentual') or None
    desconto_valor = data.get('desconto_valor') or None
    observacoes = data.get('observacoes', 'Reajuste em lote')

    def get_client_ip(req):
        x_forwarded = req.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded:
            return x_forwarded.split(',')[0].strip()
        return req.META.get('REMOTE_ADDR')

    resultados = []
    total_ok = 0
    total_erro = 0

    for contrato_id in contrato_ids:
        try:
            contrato = Contrato.objects.get(pk=int(contrato_id))
        except (Contrato.DoesNotExist, ValueError):
            resultados.append({'contrato_id': contrato_id, 'sucesso': False, 'erro': 'Contrato não encontrado'})
            total_erro += 1
            continue

        ciclo = Reajuste.calcular_ciclo_pendente(contrato)
        if ciclo is None:
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': False,
                'erro': 'Nenhum ciclo pendente',
            })
            total_erro += 1
            continue

        try:
            preview = Reajuste.preview_reajuste(
                contrato, ciclo,
                desconto_percentual=desconto_percentual,
                desconto_valor=desconto_valor,
            )
        except Exception as e:
            logger.exception('Erro no preview reajuste contrato %s: %s', contrato_id, e)
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': False,
                'erro': f'Erro no preview: {e}',
            })
            total_erro += 1
            continue

        if 'erro' in preview:
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': False,
                'erro': preview['erro'],
            })
            total_erro += 1
            continue

        if not contrato.parcelas.filter(
            numero_parcela__gte=preview['parcela_inicial'],
            numero_parcela__lte=preview['parcela_final'],
            pago=False,
        ).exists():
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': False,
                'erro': 'Sem parcelas pendentes no intervalo',
            })
            total_erro += 1
            continue

        try:
            reajuste = Reajuste.objects.create(
                contrato=contrato,
                data_reajuste=timezone.now().date(),
                indice_tipo=preview['indice_tipo'],
                percentual=preview['percentual_final'],
                percentual_bruto=preview['percentual_bruto'],
                spread_aplicado=preview['spread'] if preview['spread'] else None,
                desconto_percentual=Decimal(str(desconto_percentual)) if desconto_percentual else None,
                desconto_valor=Decimal(str(desconto_valor)) if desconto_valor else None,
                piso_aplicado=preview['piso'],
                teto_aplicado=preview['teto'],
                parcela_inicial=preview['parcela_inicial'],
                parcela_final=preview['parcela_final'],
                ciclo=ciclo,
                periodo_referencia_inicio=preview['periodo_referencia_inicio'],
                periodo_referencia_fim=preview['periodo_referencia_fim'],
                aplicado_manual=True,
                usuario=request.user,
                ip_address=get_client_ip(request),
                observacoes=observacoes,
            )
            resultado = reajuste.aplicar_reajuste()
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': True,
                'ciclo': ciclo,
                'percentual': float(preview['percentual_final']),
                'parcelas_reajustadas': resultado.get('parcelas_reajustadas', 0),
            })
            total_ok += 1
        except Exception as e:
            logger.exception(f"Erro ao aplicar reajuste em lote no contrato {contrato_id}: {e}")
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': getattr(contrato, 'numero_contrato', str(contrato_id)),
                'sucesso': False,
                'erro': str(e),
            })
            total_erro += 1

    return JsonResponse({
        'sucesso': True,
        'total_processados': len(resultados),
        'total_ok': total_ok,
        'total_erro': total_erro,
        'resultados': resultados,
    })


@login_required
@require_POST
def aplicar_reajuste_informado_lote(request):
    """
    Aplica um percentual INFORMADO pelo usuário em múltiplos contratos.

    Diferente de aplicar_reajuste_lote (que usa o índice calculado),
    este endpoint recebe um percentual_informado fixo e o aplica a todos
    os contratos selecionados — útil para acordos, arredondamentos ou
    quando o índice ainda não está disponível.

    Corpo JSON:
      - contrato_ids: lista de IDs de contratos (int[])
      - percentual_informado: percentual a aplicar (Decimal)
      - desconto_percentual: desconto em p.p. (opcional)
      - desconto_valor: desconto em R$/parcela (opcional)
      - observacoes: texto livre (opcional)
    """
    import json

    try:
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
    except (json.JSONDecodeError, Exception):
        return JsonResponse({'sucesso': False, 'erro': 'Corpo da requisição inválido'}, status=400)

    contrato_ids = data.get('contrato_ids', [])
    if not contrato_ids:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhum contrato selecionado'}, status=400)

    try:
        percentual_informado = Decimal(str(data.get('percentual_informado', '')))
    except Exception:
        return JsonResponse({'sucesso': False, 'erro': 'percentual_informado inválido'}, status=400)

    desconto_percentual = data.get('desconto_percentual') or None
    desconto_valor = data.get('desconto_valor') or None
    observacoes = data.get('observacoes', 'Reajuste informado em lote')

    def get_client_ip(req):
        x_forwarded = req.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded:
            return x_forwarded.split(',')[0].strip()
        return req.META.get('REMOTE_ADDR')

    resultados = []
    total_ok = 0
    total_erro = 0

    for contrato_id in contrato_ids:
        try:
            contrato = Contrato.objects.get(pk=int(contrato_id))
        except (Contrato.DoesNotExist, ValueError):
            resultados.append({'contrato_id': contrato_id, 'sucesso': False, 'erro': 'Contrato não encontrado'})
            total_erro += 1
            continue

        ciclo = Reajuste.calcular_ciclo_pendente(contrato)
        if ciclo is None:
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': False,
                'erro': 'Nenhum ciclo pendente',
            })
            total_erro += 1
            continue

        inicio_ref, fim_ref = Reajuste.calcular_periodo_referencia(contrato, ciclo)
        prazo = contrato.prazo_reajuste_meses
        parcela_inicial = (ciclo - 1) * prazo + 1
        parcela_final = min(ciclo * prazo, contrato.numero_parcelas)

        # Aplica piso/teto ao percentual informado
        piso = contrato.reajuste_piso or Decimal('-100')
        teto = contrato.reajuste_teto or Decimal('999')
        perc_final = max(piso, min(teto, percentual_informado))

        # Desconto
        desc_pct = Decimal(str(desconto_percentual)) if desconto_percentual else Decimal('0')
        desc_val = Decimal(str(desconto_valor)) if desconto_valor else None
        perc_final = max(piso, perc_final - desc_pct)

        try:
            with transaction.atomic():
                reajuste = Reajuste.objects.create(
                    contrato=contrato,
                    data_reajuste=timezone.now().date(),
                    indice_tipo=contrato.tipo_correcao,
                    percentual=perc_final,
                    percentual_bruto=percentual_informado,
                    desconto_percentual=Decimal(str(desconto_percentual)) if desconto_percentual else None,
                    desconto_valor=desc_val,
                    piso_aplicado=piso if piso != Decimal('-100') else None,
                    teto_aplicado=teto if teto != Decimal('999') else None,
                    parcela_inicial=parcela_inicial,
                    parcela_final=parcela_final,
                    ciclo=ciclo,
                    periodo_referencia_inicio=inicio_ref,
                    periodo_referencia_fim=fim_ref,
                    aplicado_manual=True,
                    usuario=request.user,
                    ip_address=get_client_ip(request),
                    observacoes=observacoes,
                )
                resultado = reajuste.aplicar_reajuste()
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': contrato.numero_contrato,
                'sucesso': True,
                'ciclo': ciclo,
                'percentual': float(perc_final),
                'parcelas_reajustadas': resultado.get('parcelas_reajustadas', 0),
            })
            total_ok += 1
        except Exception as e:
            logger.exception('Erro ao aplicar reajuste informado no contrato %s: %s', contrato_id, e)
            resultados.append({
                'contrato_id': contrato_id,
                'numero_contrato': getattr(contrato, 'numero_contrato', str(contrato_id)),
                'sucesso': False,
                'erro': str(e),
            })
            total_erro += 1

    return JsonResponse({
        'sucesso': True,
        'total_processados': len(resultados),
        'total_ok': total_ok,
        'total_erro': total_erro,
        'resultados': resultados,
    })


@login_required
@require_POST
def excluir_reajuste(request, pk):
    """
    Exclui um reajuste e reverte os valores das parcelas.
    """
    reajuste = get_object_or_404(Reajuste, pk=pk)

    try:
        with transaction.atomic():
            contrato = reajuste.contrato

            # Calcular o fator efetivamente aplicado (percentual já inclui teto/piso)
            perc_aplicado = reajuste.percentual
            fator_reajuste = 1 + (perc_aplicado / 100)

            parcelas = contrato.parcelas.filter(
                numero_parcela__gte=reajuste.parcela_inicial,
                numero_parcela__lte=reajuste.parcela_final,
                pago=False
            )

            for parcela in parcelas:
                # Reverter o valor (dividir pelo fator aplicado)
                if fator_reajuste != 0:
                    parcela.valor_atual = (parcela.valor_atual / fator_reajuste).quantize(Decimal('0.01'))
                parcela.save(update_fields=['valor_atual'])

            # Reverter intermediárias
            intermediarias = contrato.intermediarias.filter(
                paga=False,
                mes_vencimento__gte=reajuste.parcela_inicial,
                mes_vencimento__lte=reajuste.parcela_final,
            )
            for inter in intermediarias:
                if fator_reajuste != 0:
                    inter.valor_atual = (inter.valor_atual / fator_reajuste).quantize(Decimal('0.01'))
                inter.save(update_fields=['valor_atual'])

            # Restaurar ciclo_reajuste_atual do contrato
            if contrato.ciclo_reajuste_atual >= reajuste.ciclo:
                contrato.ciclo_reajuste_atual = reajuste.ciclo - 1
                contrato.save(update_fields=['ciclo_reajuste_atual'])

            # Excluir o registro de reajuste
            reajuste.delete()

        return JsonResponse({
            'sucesso': True,
            'mensagem': 'Reajuste excluido e valores revertidos com sucesso'
        })

    except Exception as e:
        logger.exception(f"Erro ao excluir reajuste: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
def api_reajuste_detail(request, pk):
    """Retorna JSON completo de um reajuste para os modais Visualizar/Alterar."""
    reajuste = get_object_or_404(Reajuste, pk=pk)
    contrato = reajuste.contrato

    parcelas_no_range = contrato.parcelas.filter(
        numero_parcela__gte=reajuste.parcela_inicial,
        numero_parcela__lte=reajuste.parcela_final,
    )
    total_parcelas = parcelas_no_range.count()

    valor_original_total = parcelas_no_range.filter(pago=False).aggregate(
        s=Sum('valor_original')
    )['s'] or Decimal('0')
    valor_atual_total = parcelas_no_range.filter(pago=False).aggregate(
        s=Sum('valor_atual')
    )['s'] or Decimal('0')

    primeira_aberta = parcelas_no_range.filter(pago=False).order_by('numero_parcela').first()
    prestacao_atual = float(primeira_aberta.valor_atual) if primeira_aberta else None
    qtd_parcelas_abertas = parcelas_no_range.filter(pago=False).count()

    # Ciclos afetados — determina pelo range de parcelas e prazo_reajuste_meses
    prazo = contrato.prazo_reajuste_meses or 12
    ciclo_inicial = ((reajuste.parcela_inicial - 1) // prazo) + 1
    ciclo_final = ((reajuste.parcela_final - 1) // prazo) + 1
    afeta_multiplos_ciclos = ciclo_final > ciclo_inicial

    # Se afeta >1 ciclo: pega a 1ª parcela de cada ciclo afetado dentro do range
    parcelas_por_ciclo = []
    if afeta_multiplos_ciclos:
        for c in range(ciclo_inicial, ciclo_final + 1):
            primeira_num = max(reajuste.parcela_inicial, (c - 1) * prazo + 1)
            if primeira_num > reajuste.parcela_final:
                break
            p = contrato.parcelas.filter(numero_parcela=primeira_num).values(
                'numero_parcela', 'valor_original', 'valor_atual', 'pago', 'data_vencimento'
            ).first()
            if p:
                p['ciclo'] = c
                p['data_vencimento'] = p['data_vencimento'].strftime('%d/%m/%Y') if p['data_vencimento'] else None
                p['valor_original'] = float(p['valor_original']) if p['valor_original'] else None
                p['valor_atual'] = float(p['valor_atual']) if p['valor_atual'] else None
                parcelas_por_ciclo.append(p)

    return JsonResponse({
        'pk': reajuste.pk,
        'contrato': contrato.numero_contrato,
        'contrato_pk': contrato.pk,
        'comprador': contrato.comprador.nome,
        'data_reajuste': reajuste.data_reajuste.strftime('%d/%m/%Y'),
        'ciclo': reajuste.ciclo,
        'indice_tipo': reajuste.indice_tipo,
        'percentual_bruto': float(reajuste.percentual_bruto) if reajuste.percentual_bruto is not None else None,
        'percentual': float(reajuste.percentual),
        'spread_aplicado': float(reajuste.spread_aplicado) if reajuste.spread_aplicado else 0,
        'desconto_percentual': float(reajuste.desconto_percentual) if reajuste.desconto_percentual else 0,
        'desconto_valor': float(reajuste.desconto_valor) if reajuste.desconto_valor else 0,
        'piso_aplicado': float(reajuste.piso_aplicado) if reajuste.piso_aplicado is not None else None,
        'teto_aplicado': float(reajuste.teto_aplicado) if reajuste.teto_aplicado is not None else None,
        'parcela_inicial': reajuste.parcela_inicial,
        'parcela_final': reajuste.parcela_final,
        'periodo_referencia_inicio': reajuste.periodo_referencia_inicio.strftime('%m/%Y') if reajuste.periodo_referencia_inicio else None,
        'periodo_referencia_fim': reajuste.periodo_referencia_fim.strftime('%m/%Y') if reajuste.periodo_referencia_fim else None,
        'aplicado': reajuste.aplicado,
        'aplicado_manual': reajuste.aplicado_manual,
        'data_aplicacao': reajuste.data_aplicacao.strftime('%d/%m/%Y %H:%M') if reajuste.data_aplicacao else None,
        'usuario': (reajuste.usuario.get_full_name() or reajuste.usuario.username) if reajuste.usuario else None,
        'observacoes': reajuste.observacoes or '',
        'total_parcelas': total_parcelas,
        'qtd_parcelas_abertas': qtd_parcelas_abertas,
        'prestacao_atual': prestacao_atual,
        'valor_original_total': float(valor_original_total),
        'valor_atual_total': float(valor_atual_total),
        'diferenca_total': float(valor_atual_total - valor_original_total),
        # Resumo por ciclo
        'afeta_multiplos_ciclos': afeta_multiplos_ciclos,
        'ciclo_inicial': ciclo_inicial,
        'ciclo_final': ciclo_final,
        'num_ciclos_afetados': ciclo_final - ciclo_inicial + 1,
        'parcelas_por_ciclo': parcelas_por_ciclo,
    })


@login_required
@require_POST
def alterar_indice_reajuste(request, pk):
    """
    Altera o percentual de um reajuste (entrada manual). O tipo de índice é preservado.
    Se o reajuste já foi aplicado, reverte o fator antigo nas parcelas não pagas
    e reaplica automaticamente com o novo percentual.
    """
    import json

    try:
        body = json.loads(request.body)
        novo_perc_raw = body.get('percentual')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)

    reajuste = get_object_or_404(Reajuste, pk=pk)

    if novo_perc_raw is None or novo_perc_raw == '':
        return JsonResponse({'sucesso': False, 'erro': 'Informe o novo percentual.'}, status=400)

    try:
        novo_percentual = Decimal(str(novo_perc_raw).replace(',', '.'))
    except Exception:
        return JsonResponse({'sucesso': False, 'erro': 'Percentual inválido.'}, status=400)

    if novo_percentual <= Decimal('-100'):
        return JsonResponse({'sucesso': False, 'erro': 'Percentual deve ser maior que -100%.'}, status=400)

    if novo_percentual == reajuste.percentual:
        return JsonResponse({'sucesso': False, 'erro': 'O percentual informado é igual ao atual.'}, status=400)

    try:
        with transaction.atomic():
            contrato = reajuste.contrato

            if reajuste.aplicado:
                # Reverter o fator antigo nas parcelas não pagas do intervalo
                fator_antigo = 1 + (reajuste.percentual / 100)
                parcelas = contrato.parcelas.filter(
                    numero_parcela__gte=reajuste.parcela_inicial,
                    numero_parcela__lte=reajuste.parcela_final,
                    pago=False,
                )
                for p in parcelas:
                    if fator_antigo != 0:
                        p.valor_atual = (p.valor_atual / fator_antigo).quantize(Decimal('0.01'))
                    p.save(update_fields=['valor_atual'])

                intermediarias = contrato.intermediarias.filter(
                    paga=False,
                    mes_vencimento__gte=reajuste.parcela_inicial,
                    mes_vencimento__lte=reajuste.parcela_final,
                )
                for inter in intermediarias:
                    if fator_antigo != 0:
                        inter.valor_atual = (inter.valor_atual / fator_antigo).quantize(Decimal('0.01'))
                    inter.save(update_fields=['valor_atual'])

                # Atualiza só o percentual (índice preservado) e reaplica
                reajuste.percentual = novo_percentual
                reajuste.percentual_bruto = novo_percentual  # entrada manual: bruto = final
                reajuste.aplicado_manual = True
                reajuste.usuario = request.user
                reajuste.aplicado = False
                reajuste.save()

                resultado = reajuste.aplicar_reajuste()
                if not resultado.get('sucesso'):
                    raise Exception(resultado.get('erro', 'Erro ao reaplicar reajuste.'))
            else:
                reajuste.percentual = novo_percentual
                reajuste.percentual_bruto = novo_percentual
                reajuste.aplicado_manual = True
                reajuste.usuario = request.user
                reajuste.save()

    except Exception as e:
        logger.exception(f"Erro ao alterar reajuste {pk}: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)

    return JsonResponse({
        'sucesso': True,
        'mensagem': f'Percentual atualizado para {float(novo_percentual):+.4f}% (índice {reajuste.indice_tipo} preservado).',
        'novo_indice_tipo': reajuste.indice_tipo,
        'novo_percentual_bruto': float(reajuste.percentual_bruto),
        'novo_percentual': float(reajuste.percentual),
    })


@login_required
def obter_indice_reajuste(request):
    """
    Retorna o percentual do indice de reajuste para um tipo e periodo.
    Parametros GET: tipo_indice, ano, mes
    """
    from contratos.models import IndiceReajuste

    tipo_indice = request.GET.get('tipo_indice', '')
    ano = request.GET.get('ano', timezone.now().year)
    mes = request.GET.get('mes', timezone.now().month)

    try:
        ano = int(ano)
        mes = int(mes)

        indice = IndiceReajuste.objects.filter(
            tipo_indice=tipo_indice,
            ano=ano,
            mes=mes
        ).first()

        if indice:
            return JsonResponse({
                'sucesso': True,
                'percentual': float(indice.valor),  # Campo correto e 'valor', nao 'percentual'
                'valor': float(indice.valor),
                'tipo_indice': indice.tipo_indice,
                'ano': indice.ano,
                'mes': indice.mes,
                'valor_acumulado_12m': float(indice.valor_acumulado_12m) if indice.valor_acumulado_12m else None,
                'valor_acumulado_ano': float(indice.valor_acumulado_ano) if indice.valor_acumulado_ano else None,
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'erro': f'Indice {tipo_indice} nao encontrado para {mes}/{ano}'
            }, status=404)

    except Exception as e:
        logger.exception("Erro ao obter indice reajuste: %s", e)
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
def calcular_reajuste_proporcional(request, contrato_id):
    """
    Calcula o percentual de reajuste acumulado proporcional para um contrato.

    O calculo considera:
    - Primeiro mes: proporcional aos dias restantes do mes
    - Meses intermediarios: indice completo
    - Ultimo mes: proporcional aos dias do mes ate a data final

    Parametros GET: tipo_indice (opcional, usa o indice do contrato se nao informado)
    """
    from contratos.models import IndiceReajuste
    from calendar import monthrange

    contrato = get_object_or_404(Contrato, pk=contrato_id)

    try:
        tipo_indice = request.GET.get('tipo_indice', contrato.get_tipo_correcao_display())

        # Determinar periodo do reajuste
        if contrato.data_ultimo_reajuste:
            data_inicio = contrato.data_ultimo_reajuste
        else:
            data_inicio = contrato.data_contrato

        data_fim = timezone.now().date()

        # Se nao passou prazo minimo, retornar 0
        meses_desde = (data_fim.year - data_inicio.year) * 12 + (data_fim.month - data_inicio.month)
        if meses_desde < contrato.prazo_reajuste_meses:
            return JsonResponse({
                'sucesso': False,
                'erro': f'Reajuste ainda nao e necessario. Proximo reajuste em {contrato.prazo_reajuste_meses - meses_desde} mese(s).'
            })

        # Calcular reajuste proporcional
        fator_acumulado = Decimal('1.0')
        detalhes_meses = []

        # Iterar pelos meses do periodo
        mes_atual = data_inicio.month
        ano_atual = data_inicio.year

        while (ano_atual < data_fim.year) or (ano_atual == data_fim.year and mes_atual <= data_fim.month):
            # Buscar indice do mes
            indice = IndiceReajuste.objects.filter(
                tipo_indice=tipo_indice,
                ano=ano_atual,
                mes=mes_atual
            ).first()

            if indice:
                percentual_mes = indice.valor  # Campo correto e 'valor', nao 'percentual'
            else:
                percentual_mes = Decimal('0')

            # Calcular dias do mes
            dias_no_mes = monthrange(ano_atual, mes_atual)[1]

            # Primeiro mes - proporcional
            if ano_atual == data_inicio.year and mes_atual == data_inicio.month:
                dias_considerados = dias_no_mes - data_inicio.day + 1
                proporcao = Decimal(dias_considerados) / Decimal(dias_no_mes)
            # Ultimo mes - proporcional
            elif ano_atual == data_fim.year and mes_atual == data_fim.month:
                dias_considerados = data_fim.day
                proporcao = Decimal(dias_considerados) / Decimal(dias_no_mes)
            # Meses intermediarios - completo
            else:
                dias_considerados = dias_no_mes
                proporcao = Decimal('1.0')

            # Aplicar proporcao ao indice
            percentual_proporcional = percentual_mes * proporcao
            fator_mes = 1 + (percentual_proporcional / 100)
            fator_acumulado *= fator_mes

            detalhes_meses.append({
                'mes': mes_atual,
                'ano': ano_atual,
                'indice_original': float(percentual_mes),
                'dias_considerados': dias_considerados,
                'dias_mes': dias_no_mes,
                'proporcao': float(proporcao),
                'indice_proporcional': float(percentual_proporcional)
            })

            # Avancar para proximo mes
            mes_atual += 1
            if mes_atual > 12:
                mes_atual = 1
                ano_atual += 1

        # Calcular percentual total
        percentual_total = (fator_acumulado - 1) * 100

        return JsonResponse({
            'sucesso': True,
            'percentual_acumulado': float(percentual_total),
            'fator_acumulado': float(fator_acumulado),
            'data_inicio': data_inicio.strftime('%d/%m/%Y'),
            'data_fim': data_fim.strftime('%d/%m/%Y'),
            'tipo_indice': tipo_indice,
            'meses_detalhes': detalhes_meses,
            'total_meses': len(detalhes_meses)
        })

    except Exception as e:
        logger.exception(f"Erro ao calcular reajuste proporcional: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


@login_required
def api_calcular_indice_acumulado(request):
    """
    API para calcular o índice acumulado com detalhes completos.

    Metodologia Cálculo Exato (calculoexato.com.br):
    - Fator = produto de (1 + índice_mensal/100) para cada mês
    - Percentual acumulado = (Fator - 1) * 100
    - Valor atualizado = Valor original * Fator

    Parâmetros GET:
        tipo_indice: Tipo do índice (IPCA, IGPM, etc.)
        ano_inicio: Ano inicial
        mes_inicio: Mês inicial
        ano_fim: Ano final
        mes_fim: Mês final
        proporcional: Se 'true', calcula pro-rata para primeiro/último mês
        valor_original: Valor para simular atualização (opcional)

    Retorno:
        JSON com detalhes completos do cálculo incluindo:
        - fator_acumulado: Fator multiplicador
        - percentual_acumulado: Percentual de correção
        - meses: Lista com detalhes de cada mês
        - calculo_detalhado: Fórmula textual do cálculo
        - valor_atualizado: Se valor_original informado
    """
    from financeiro.services.reajuste_service import ReajusteService
    from datetime import date

    try:
        tipo_indice = request.GET.get('tipo_indice', 'IPCA')
        ano_inicio = int(request.GET.get('ano_inicio', timezone.now().year - 1))
        mes_inicio = int(request.GET.get('mes_inicio', 1))
        ano_fim = int(request.GET.get('ano_fim', timezone.now().year))
        mes_fim = int(request.GET.get('mes_fim', timezone.now().month))
        proporcional = request.GET.get('proporcional', 'false').lower() == 'true'
        valor_original_str = request.GET.get('valor_original', '')

        # Criar datas
        data_inicio = date(ano_inicio, mes_inicio, 1)
        data_fim = date(ano_fim, mes_fim, 1)

        # Validar datas
        if data_inicio > data_fim:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Data inicial deve ser anterior ou igual à data final'
            }, status=400)

        # Calcular usando o serviço
        service = ReajusteService()
        resultado = service.calcular_indice_acumulado_detalhado(
            tipo_indice=tipo_indice,
            data_inicio=data_inicio,
            data_fim=data_fim,
            proporcional=proporcional
        )

        # Se valor original informado, calcular valor atualizado
        if valor_original_str:
            try:
                valor_original = Decimal(valor_original_str.replace(',', '.'))
                fator = Decimal(str(resultado['fator_acumulado']))
                valor_atualizado = valor_original * fator
                resultado['valor_original'] = float(valor_original)
                resultado['valor_atualizado'] = float(valor_atualizado)
                resultado['diferenca'] = float(valor_atualizado - valor_original)
            except (ValueError, TypeError):
                pass

        return JsonResponse(resultado)

    except ValueError as e:
        return JsonResponse({
            'sucesso': False,
            'erro': f'Valor inválido: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.exception(f"Erro ao calcular índice acumulado: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


# =============================================================================
# DASHBOARD CONSOLIDADO DA CONTABILIDADE
# =============================================================================

class DashboardContabilidadeView(LoginRequiredMixin, TemplateView):
    """
    Dashboard consolidado para a Contabilidade.

    Apresenta visão geral de todas as imobiliárias sob gestão,
    com estatísticas consolidadas e alertas importantes.
    """
    template_name = 'financeiro/dashboard_contabilidade.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from core.models import Contabilidade
        from contratos.models import PrestacaoIntermediaria

        hoje = timezone.now().date()

        # Filtro por contabilidade (se usuário tiver acesso restrito)
        contabilidade_id = self.request.GET.get('contabilidade')
        contabilidade_selecionada = None

        if contabilidade_id:
            try:
                contabilidade_selecionada = Contabilidade.objects.get(pk=contabilidade_id)
            except Contabilidade.DoesNotExist:
                pass

        # Lista de contabilidades
        contabilidades = Contabilidade.objects.filter(ativo=True)
        context['contabilidades'] = contabilidades
        context['contabilidade_selecionada'] = contabilidade_selecionada

        # Base querysets
        if contabilidade_selecionada:
            imobiliarias = Imobiliaria.objects.filter(
                contabilidade=contabilidade_selecionada, ativo=True
            )
        else:
            imobiliarias = Imobiliaria.objects.filter(ativo=True)

        imobiliaria_ids = imobiliarias.values_list('id', flat=True)

        contratos_qs = Contrato.objects.filter(imobiliaria__in=imobiliaria_ids)
        parcelas_qs = Parcela.objects.filter(contrato__imobiliaria__in=imobiliaria_ids)

        # =========================================================================
        # ESTATÍSTICAS GERAIS
        # =========================================================================
        context['total_imobiliarias'] = imobiliarias.count()

        # Estatísticas de contratos
        stats_contratos = contratos_qs.aggregate(
            total=Count('id'),
            ativos=Count('id', filter=Q(status=StatusContrato.ATIVO)),
            quitados=Count('id', filter=Q(status=StatusContrato.QUITADO)),
            cancelados=Count('id', filter=Q(status=StatusContrato.CANCELADO)),
            valor_total=Sum('valor_total'),
        )
        context['stats_contratos'] = stats_contratos

        # Estatísticas de parcelas
        stats_parcelas = parcelas_qs.aggregate(
            total=Count('id'),
            pagas=Count('id', filter=Q(pago=True)),
            pendentes=Count('id', filter=Q(pago=False)),
            vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
            valor_total=Sum('valor_atual'),
            valor_recebido=Sum('valor_pago', filter=Q(pago=True)),
            valor_pendente=Sum('valor_atual', filter=Q(pago=False)),
            valor_vencido=Sum('valor_atual', filter=Q(pago=False, data_vencimento__lt=hoje)),
        )
        context['stats_parcelas'] = stats_parcelas

        # =========================================================================
        # ESTATÍSTICAS POR IMOBILIÁRIA
        # =========================================================================
        stats_por_imobiliaria = []
        for imob in imobiliarias:
            parcelas_imob = parcelas_qs.filter(contrato__imobiliaria=imob)
            contratos_imob = contratos_qs.filter(imobiliaria=imob)

            # Contar contratos com bloqueio por reajuste
            # verificar_bloqueio_reajuste() retorna bool (True = bloqueado)
            contratos_bloqueados = sum(
                1 for contrato in contratos_imob.filter(status=StatusContrato.ATIVO)
                if hasattr(contrato, 'verificar_bloqueio_reajuste') and contrato.verificar_bloqueio_reajuste()
            )

            stats = {
                'imobiliaria': imob,
                'total_contratos': contratos_imob.count(),
                'contratos_ativos': contratos_imob.filter(status=StatusContrato.ATIVO).count(),
                'contratos_bloqueados': contratos_bloqueados,
                'parcelas_pendentes': parcelas_imob.filter(pago=False).count(),
                'parcelas_vencidas': parcelas_imob.filter(pago=False, data_vencimento__lt=hoje).count(),
                'valor_pendente': parcelas_imob.filter(pago=False).aggregate(
                    total=Sum('valor_atual')
                )['total'] or Decimal('0.00'),
                'valor_vencido': parcelas_imob.filter(pago=False, data_vencimento__lt=hoje).aggregate(
                    total=Sum('valor_atual')
                )['total'] or Decimal('0.00'),
            }
            stats_por_imobiliaria.append(stats)

        # Ordenar por valor vencido (mais críticos primeiro)
        stats_por_imobiliaria.sort(key=lambda x: x['valor_vencido'], reverse=True)
        context['stats_por_imobiliaria'] = stats_por_imobiliaria

        # =========================================================================
        # CONTRATOS COM REAJUSTE PENDENTE (TODAS IMOBILIÁRIAS)
        # =========================================================================
        contratos_reajuste_pendente = []
        for contrato in contratos_qs.filter(status=StatusContrato.ATIVO):
            if hasattr(contrato, 'data_ultimo_reajuste') and hasattr(contrato, 'prazo_reajuste_meses'):
                data_base = contrato.data_ultimo_reajuste or contrato.data_contrato
                proxima_data_reajuste = data_base + relativedelta(months=contrato.prazo_reajuste_meses)

                if proxima_data_reajuste <= hoje:
                    meses_atraso = (hoje.year - proxima_data_reajuste.year) * 12 + (hoje.month - proxima_data_reajuste.month)
                    contratos_reajuste_pendente.append({
                        'contrato': contrato,
                        'imobiliaria': contrato.imobiliaria,
                        'data_ultimo_reajuste': contrato.data_ultimo_reajuste,
                        'proxima_data_reajuste': proxima_data_reajuste,
                        'meses_atraso': meses_atraso,
                        'tipo_correcao': contrato.get_tipo_correcao_display() if hasattr(contrato, 'get_tipo_correcao_display') else 'N/A',
                    })

        contratos_reajuste_pendente.sort(key=lambda x: x['meses_atraso'], reverse=True)
        context['contratos_reajuste_pendente'] = contratos_reajuste_pendente[:20]
        context['total_reajustes_pendentes'] = len(contratos_reajuste_pendente)

        # =========================================================================
        # ALERTAS E INDICADORES
        # =========================================================================
        # Inadimplência (parcelas vencidas há mais de 30 dias)
        parcelas_inadimplentes = parcelas_qs.filter(
            pago=False,
            data_vencimento__lt=hoje - timedelta(days=30)
        ).count()
        context['parcelas_inadimplentes'] = parcelas_inadimplentes

        # Taxa de inadimplência
        total_parcelas = stats_parcelas['total'] or 1
        taxa_inadimplencia = (parcelas_inadimplentes / total_parcelas) * 100
        context['taxa_inadimplencia'] = round(taxa_inadimplencia, 2)

        # Receita prevista para os próximos 30 dias
        receita_prevista = parcelas_qs.filter(
            pago=False,
            data_vencimento__gte=hoje,
            data_vencimento__lte=hoje + timedelta(days=30)
        ).aggregate(total=Sum('valor_atual'))['total'] or Decimal('0.00')
        context['receita_prevista_30d'] = receita_prevista

        # =========================================================================
        # PRESTAÇÕES INTERMEDIÁRIAS
        # =========================================================================
        intermediarias_qs = PrestacaoIntermediaria.objects.filter(
            contrato__imobiliaria__in=imobiliaria_ids
        )
        stats_intermediarias = intermediarias_qs.aggregate(
            total=Count('id'),
            pagas=Count('id', filter=Q(paga=True)),
            pendentes=Count('id', filter=Q(paga=False)),
            valor_total=Sum('valor'),
            valor_pago=Sum('valor_pago', filter=Q(paga=True)),
            valor_pendente=Sum('valor', filter=Q(paga=False)),
        )
        context['stats_intermediarias'] = stats_intermediarias

        # =========================================================================
        # PERÍODO DO MÊS ATUAL
        # =========================================================================
        primeiro_dia_mes = hoje.replace(day=1)
        if hoje.month == 12:
            ultimo_dia_mes = hoje.replace(day=31)
        else:
            ultimo_dia_mes = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)

        parcelas_mes = parcelas_qs.filter(
            data_vencimento__gte=primeiro_dia_mes,
            data_vencimento__lte=ultimo_dia_mes
        )
        stats_mes = parcelas_mes.aggregate(
            total=Count('id'),
            pagas=Count('id', filter=Q(pago=True)),
            pendentes=Count('id', filter=Q(pago=False)),
            valor_total=Sum('valor_atual'),
            valor_recebido=Sum('valor_pago', filter=Q(pago=True)),
            valor_pendente=Sum('valor_atual', filter=Q(pago=False)),
        )
        context['stats_mes'] = stats_mes
        context['mes_atual'] = hoje.strftime('%B/%Y')

        return context


@login_required
def api_dashboard_contabilidade(request):
    """
    API para retornar dados do dashboard de contabilidade em JSON.
    Usado para gráficos e atualizações via AJAX.
    """

    hoje = timezone.now().date()

    # Filtro por contabilidade
    contabilidade_id = request.GET.get('contabilidade')
    if contabilidade_id:
        imobiliarias = Imobiliaria.objects.filter(
            contabilidade_id=contabilidade_id, ativo=True
        )
    else:
        imobiliarias = Imobiliaria.objects.filter(ativo=True)

    imobiliaria_ids = imobiliarias.values_list('id', flat=True)
    parcelas_qs = Parcela.objects.filter(contrato__imobiliaria__in=imobiliaria_ids)
    contratos_qs = Contrato.objects.filter(imobiliaria__in=imobiliaria_ids)

    # Dados para gráficos
    meses = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

    # Recebimentos por mês (últimos 12 meses)
    recebimentos_mensais = {'labels': [], 'recebido': [], 'esperado': []}
    for i in range(11, -1, -1):
        data = hoje - relativedelta(months=i)
        primeiro_dia = data.replace(day=1)
        ultimo_dia = (primeiro_dia + relativedelta(months=1)) - timedelta(days=1)

        parcelas_mes = parcelas_qs.filter(
            data_vencimento__gte=primeiro_dia,
            data_vencimento__lte=ultimo_dia
        )

        recebido = parcelas_mes.filter(pago=True).aggregate(
            total=Sum('valor_pago')
        )['total'] or 0
        esperado = parcelas_mes.aggregate(
            total=Sum('valor_atual')
        )['total'] or 0

        recebimentos_mensais['labels'].append(f"{meses[data.month-1]}/{data.year % 100}")
        recebimentos_mensais['recebido'].append(float(recebido))
        recebimentos_mensais['esperado'].append(float(esperado))

    # Distribuição por imobiliária
    distribuicao_imobiliarias = {'labels': [], 'valores': [], 'cores': []}
    cores = ['#007bff', '#28a745', '#dc3545', '#ffc107', '#17a2b8', '#6c757d', '#fd7e14', '#6610f2']

    for i, imob in enumerate(imobiliarias[:8]):  # Máximo 8 imobiliárias
        valor = parcelas_qs.filter(
            contrato__imobiliaria=imob, pago=False
        ).aggregate(total=Sum('valor_atual'))['total'] or 0

        distribuicao_imobiliarias['labels'].append(imob.nome_fantasia or imob.razao_social[:20])
        distribuicao_imobiliarias['valores'].append(float(valor))
        distribuicao_imobiliarias['cores'].append(cores[i % len(cores)])

    # Status dos contratos
    status_contratos = {
        'labels': ['Ativos', 'Quitados', 'Cancelados', 'Suspensos'],
        'data': [
            contratos_qs.filter(status=StatusContrato.ATIVO).count(),
            contratos_qs.filter(status=StatusContrato.QUITADO).count(),
            contratos_qs.filter(status=StatusContrato.CANCELADO).count(),
            contratos_qs.filter(status=StatusContrato.SUSPENSO).count(),
        ],
        'cores': ['#28a745', '#007bff', '#dc3545', '#6c757d']
    }

    return JsonResponse({
        'recebimentos_mensais': recebimentos_mensais,
        'distribuicao_imobiliarias': distribuicao_imobiliarias,
        'status_contratos': status_contratos,
    })


# =============================================================================
# VIEWS DE RELATÓRIOS AVANÇADOS
# =============================================================================

class RelatorioPrestacoesAPagarView(LoginRequiredMixin, TemplateView):
    """
    Relatório de prestações a pagar.
    Permite filtrar por contrato, período, imobiliária e status.
    """
    template_name = 'financeiro/relatorios/prestacoes_a_pagar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .services import RelatorioService, FiltroRelatorio

        # Obter parâmetros de filtro
        contrato_id = self.request.GET.get('contrato')
        imobiliaria_id = self.request.GET.get('imobiliaria')
        data_inicio = self.request.GET.get('data_inicio')
        data_fim = self.request.GET.get('data_fim')
        apenas_vencidas = self.request.GET.get('apenas_vencidas') == 'true'

        # Criar filtro
        filtro = FiltroRelatorio()
        if contrato_id:
            filtro.contrato_id = int(contrato_id)
        if imobiliaria_id:
            filtro.imobiliaria_id = int(imobiliaria_id)
        if data_inicio:
            try:
                from datetime import datetime
                filtro.data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            except ValueError:
                pass
        if data_fim:
            try:
                from datetime import datetime
                filtro.data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Gerar relatório
        service = RelatorioService()
        relatorio = service.gerar_relatorio_prestacoes_a_pagar(filtro)

        context['relatorio'] = relatorio
        context['filtros'] = {
            'contrato_id': contrato_id,
            'imobiliaria_id': imobiliaria_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'apenas_vencidas': apenas_vencidas,
        }

        # Listas para os filtros
        context['imobiliarias'] = Imobiliaria.objects.filter(ativo=True)
        context['contratos'] = Contrato.objects.filter(status=StatusContrato.ATIVO)

        return context


class RelatorioPrestacoesPageasView(LoginRequiredMixin, TemplateView):
    """
    Relatório de prestações pagas.
    Permite filtrar por período de pagamento, contrato e imobiliária.
    """
    template_name = 'financeiro/relatorios/prestacoes_pagas.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .services import RelatorioService, FiltroRelatorio

        # Obter parâmetros de filtro
        contrato_id = self.request.GET.get('contrato')
        imobiliaria_id = self.request.GET.get('imobiliaria')
        data_inicio = self.request.GET.get('data_inicio')
        data_fim = self.request.GET.get('data_fim')

        # Criar filtro
        filtro = FiltroRelatorio()
        if contrato_id:
            filtro.contrato_id = int(contrato_id)
        if imobiliaria_id:
            filtro.imobiliaria_id = int(imobiliaria_id)
        if data_inicio:
            try:
                from datetime import datetime
                filtro.data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            except ValueError:
                pass
        if data_fim:
            try:
                from datetime import datetime
                filtro.data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Gerar relatório
        service = RelatorioService()
        relatorio = service.gerar_relatorio_prestacoes_pagas(filtro)

        context['relatorio'] = relatorio
        context['filtros'] = {
            'contrato_id': contrato_id,
            'imobiliaria_id': imobiliaria_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
        }

        # Listas para os filtros
        context['imobiliarias'] = Imobiliaria.objects.filter(ativo=True)
        context['contratos'] = Contrato.objects.all()

        return context


class RelatorioPosicaoContratosView(LoginRequiredMixin, TemplateView):
    """
    Relatório de posição de contratos.
    Mostra resumo de cada contrato com saldo devedor, progresso, etc.
    """
    template_name = 'financeiro/relatorios/posicao_contratos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .services import RelatorioService, FiltroRelatorio

        # Obter parâmetros de filtro
        imobiliaria_id = self.request.GET.get('imobiliaria')
        status = self.request.GET.get('status')

        # Criar filtro
        filtro = FiltroRelatorio()
        if imobiliaria_id:
            filtro.imobiliaria_id = int(imobiliaria_id)

        # Gerar relatório
        service = RelatorioService()
        relatorio = service.gerar_relatorio_posicao_contratos(filtro)

        # Filtrar por status se necessário
        if status:
            relatorio['itens'] = [
                item for item in relatorio['itens']
                if item.get('status') == status
            ]

        context['relatorio'] = relatorio
        context['filtros'] = {
            'imobiliaria_id': imobiliaria_id,
            'status': status,
        }

        # Listas para os filtros
        context['imobiliarias'] = Imobiliaria.objects.filter(ativo=True)
        context['status_choices'] = StatusContrato.choices

        return context


class RelatorioPrevisaoReajustesView(LoginRequiredMixin, TemplateView):
    """
    Relatório de previsão de reajustes.
    Mostra contratos que precisarão de reajuste nos próximos dias.
    """
    template_name = 'financeiro/relatorios/previsao_reajustes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .services import RelatorioService

        # Obter parâmetros
        dias_antecedencia = int(self.request.GET.get('dias', 60))
        imobiliaria_id = self.request.GET.get('imobiliaria')

        # Gerar relatório
        service = RelatorioService()
        relatorio = service.gerar_relatorio_previsao_reajustes(dias_antecedencia)

        # Filtrar por imobiliária se necessário
        if imobiliaria_id:
            relatorio['itens'] = [
                item for item in relatorio['itens']
                if str(item.get('imobiliaria_id')) == imobiliaria_id
            ]

        context['relatorio'] = relatorio
        context['filtros'] = {
            'dias': dias_antecedencia,
            'imobiliaria_id': imobiliaria_id,
        }

        # Listas para os filtros
        context['imobiliarias'] = Imobiliaria.objects.filter(ativo=True)

        return context


@login_required
def exportar_relatorio(request, tipo):
    """
    Exporta um relatório para CSV ou JSON.

    Tipos disponíveis:
    - prestacoes_a_pagar
    - prestacoes_pagas
    - posicao_contratos
    - previsao_reajustes

    Formatos: csv, json
    """
    from .services import RelatorioService, FiltroRelatorio

    formato = request.GET.get('formato', 'csv')

    # Criar filtro com base nos parâmetros
    filtro = FiltroRelatorio()
    contrato_id = request.GET.get('contrato')
    imobiliaria_id = request.GET.get('imobiliaria')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if contrato_id:
        filtro.contrato_id = int(contrato_id)
    if imobiliaria_id:
        filtro.imobiliaria_id = int(imobiliaria_id)
    if data_inicio:
        try:
            from datetime import datetime
            filtro.data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass
    if data_fim:
        try:
            from datetime import datetime
            filtro.data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Gerar relatório
    service = RelatorioService()

    if tipo == 'prestacoes_a_pagar':
        relatorio = service.gerar_relatorio_prestacoes_a_pagar(filtro)
    elif tipo == 'prestacoes_pagas':
        relatorio = service.gerar_relatorio_prestacoes_pagas(filtro)
    elif tipo == 'posicao_contratos':
        relatorio = service.gerar_relatorio_posicao_contratos(filtro)
    elif tipo == 'previsao_reajustes':
        try:
            dias = max(1, int(request.GET.get('dias', 60)))
        except (ValueError, TypeError):
            dias = 60
        relatorio = service.gerar_relatorio_previsao_reajustes(dias)
    else:
        return HttpResponse('Tipo de relatório inválido', status=400)

    # Exportar
    from django.utils import timezone
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')

    if formato == 'json':
        conteudo = service.exportar_para_json(relatorio)
        content_type = 'application/json'
        extensao = 'json'
    elif formato == 'excel' or formato == 'xlsx':
        try:
            conteudo = service.exportar_para_excel(relatorio)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            extensao = 'xlsx'
        except ImportError as e:
            logger.exception("Biblioteca openpyxl não disponível para exportação Excel: %s", e)
            return HttpResponse(str(e), status=500)
    elif formato == 'pdf':
        try:
            conteudo = service.exportar_para_pdf(relatorio)
            content_type = 'application/pdf'
            extensao = 'pdf'
        except ImportError as e:
            logger.exception("Biblioteca reportlab não disponível para exportação PDF: %s", e)
            return HttpResponse(str(e), status=500)
    else:
        conteudo = service.exportar_para_csv(relatorio)
        content_type = 'text/csv'
        extensao = 'csv'

    # Criar response
    filename = f'relatorio_{tipo}_{timestamp}.{extensao}'

    response = HttpResponse(conteudo, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def exportar_relatorio_consolidado(request):
    """
    3.21 — Exporta relatório consolidado (multi-abas) em Excel ou PDF.

    Combina em um único arquivo:
    - Aba 1: Prestações a pagar (próximos 90 dias)
    - Aba 2: Prestações pagas (últimos 90 dias)
    - Aba 3: Posição dos contratos

    GET params:
      formato (xlsx|pdf, default xlsx)
      imobiliaria (opcional)
      data_inicio / data_fim (opcional, YYYY-MM-DD)
    """
    from .services import RelatorioService, FiltroRelatorio
    from django.utils import timezone as tz

    formato = request.GET.get('formato', 'xlsx')
    imobiliaria_id = request.GET.get('imobiliaria')
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')

    hoje = tz.now().date()

    filtro = FiltroRelatorio()
    if imobiliaria_id:
        try:
            filtro.imobiliaria_id = int(imobiliaria_id)
        except (ValueError, TypeError):
            pass

    if data_inicio_str:
        try:
            from datetime import datetime
            filtro.data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        except ValueError:
            filtro.data_inicio = hoje - timedelta(days=90)
    else:
        filtro.data_inicio = hoje - timedelta(days=90)

    if data_fim_str:
        try:
            from datetime import datetime
            filtro.data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            filtro.data_fim = hoje + timedelta(days=90)
    else:
        filtro.data_fim = hoje + timedelta(days=90)

    service = RelatorioService()

    filtro_pagar = FiltroRelatorio()
    if imobiliaria_id:
        try:
            filtro_pagar.imobiliaria_id = int(imobiliaria_id)
        except (ValueError, TypeError):
            pass
    filtro_pagar.data_inicio = hoje
    filtro_pagar.data_fim = hoje + timedelta(days=90)

    rel_a_pagar = service.gerar_relatorio_prestacoes_a_pagar(filtro_pagar)
    rel_pagas = service.gerar_relatorio_prestacoes_pagas(filtro)
    rel_posicao = service.gerar_relatorio_posicao_contratos(filtro)

    timestamp = tz.now().strftime('%Y%m%d_%H%M%S')

    if formato == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from io import BytesIO

            buf = BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=40, bottomMargin=40)
            styles = getSampleStyleSheet()
            elements = []

            def _relatorio_para_elementos(relatorio, titulo):
                elements.append(Paragraph(titulo, styles['Heading2']))
                elements.append(Spacer(1, 6))
                itens = relatorio.get('itens', [])
                if not itens:
                    elements.append(Paragraph('Nenhum item encontrado.', styles['Normal']))
                    elements.append(PageBreak())
                    return
                # Usar exportar_para_csv para extrair dados
                csv_str = service.exportar_para_csv(relatorio)
                rows = [line.split(',') for line in csv_str.strip().split('\n')]
                if rows:
                    data = [rows[0]] + rows[1:51]  # cabeçalho + até 50 linhas
                    tbl = Table(data, repeatRows=1)
                    tbl.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                    ]))
                    elements.append(tbl)
                elements.append(PageBreak())

            _relatorio_para_elementos(rel_a_pagar, 'Prestações a Pagar (próximos 90 dias)')
            _relatorio_para_elementos(rel_pagas, 'Prestações Pagas (últimos 90 dias)')
            _relatorio_para_elementos(rel_posicao, 'Posição dos Contratos')

            doc.build(elements)
            conteudo = buf.getvalue()
            content_type = 'application/pdf'
            extensao = 'pdf'
        except Exception as e:
            logger.exception("Erro ao gerar relatório consolidado PDF: %s", e)
            return HttpResponse(f'Erro ao gerar PDF: {e}', status=500)
    else:
        # Excel com múltiplas abas
        try:
            from openpyxl import Workbook
            from io import BytesIO

            wb = Workbook()
            wb.remove(wb.active)  # remove aba padrão

            for relatorio, sheet_title in [
                (rel_a_pagar, 'A Pagar (90d)'),
                (rel_pagas, 'Pagas (90d)'),
                (rel_posicao, 'Posição Contratos'),
            ]:
                xlsx_bytes = service.exportar_para_excel(relatorio)
                import openpyxl
                wb_tmp = openpyxl.load_workbook(BytesIO(xlsx_bytes))
                ws_tmp = wb_tmp.active
                ws_new = wb.create_sheet(title=sheet_title)
                for row in ws_tmp.iter_rows(values_only=True):
                    ws_new.append(list(row) if row else [])

            buf = BytesIO()
            wb.save(buf)
            conteudo = buf.getvalue()
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            extensao = 'xlsx'
        except Exception as e:
            logger.exception("Erro ao gerar relatório consolidado Excel: %s", e)
            return HttpResponse(f'Erro ao gerar Excel: {e}', status=500)

    filename = f'relatorio_consolidado_{timestamp}.{extensao}'
    response = HttpResponse(conteudo, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def api_relatorio_resumo(request):
    """
    API para retornar resumo de relatórios em JSON.
    Útil para widgets e dashboards.
    """
    from .services import RelatorioService, FiltroRelatorio

    hoje = timezone.now().date()

    # Filtro por imobiliária
    imobiliaria_id = request.GET.get('imobiliaria')
    filtro = FiltroRelatorio()
    if imobiliaria_id:
        filtro.imobiliaria_id = int(imobiliaria_id)

    service = RelatorioService()

    # Prestações a pagar (próximos 30 dias)
    filtro.data_inicio = hoje
    filtro.data_fim = hoje + timedelta(days=30)
    relatorio_a_pagar = service.gerar_relatorio_prestacoes_a_pagar(filtro)

    # Prestações pagas (últimos 30 dias)
    filtro.data_inicio = hoje - timedelta(days=30)
    filtro.data_fim = hoje
    relatorio_pagas = service.gerar_relatorio_prestacoes_pagas(filtro)

    # Previsão de reajustes
    relatorio_reajustes = service.gerar_relatorio_previsao_reajustes(60)

    return JsonResponse({
        'a_pagar': {
            'total_parcelas': relatorio_a_pagar['totalizador'].total_parcelas if hasattr(relatorio_a_pagar['totalizador'], 'total_parcelas') else 0,
            'valor_total': float(relatorio_a_pagar['totalizador'].valor_total) if hasattr(relatorio_a_pagar['totalizador'], 'valor_total') else 0,
        },
        'pagas': {
            'total_parcelas': relatorio_pagas['totalizador'].total_parcelas if hasattr(relatorio_pagas['totalizador'], 'total_parcelas') else 0,
            'valor_total': float(relatorio_pagas['totalizador'].valor_total) if hasattr(relatorio_pagas['totalizador'], 'valor_total') else 0,
        },
        'reajustes_pendentes': len(relatorio_reajustes.get('itens', [])),
    })


# =============================================================================
# APIs REST - IMOBILIÁRIAS
# =============================================================================

@login_required
@require_GET
def api_imobiliarias_lista(request):
    """
    API para listar imobiliárias da contabilidade.

    GET /api/contabilidade/imobiliarias/

    Parâmetros:
        - contabilidade: ID da contabilidade (opcional)
        - ativo: true/false (opcional)
    """
    contabilidade_id = request.GET.get('contabilidade')
    ativo = request.GET.get('ativo', 'true').lower() == 'true'

    queryset = Imobiliaria.objects.all()

    if contabilidade_id:
        queryset = queryset.filter(contabilidade_id=contabilidade_id)

    if ativo is not None:
        queryset = queryset.filter(ativo=ativo)

    imobiliarias = []
    for imob in queryset.select_related('contabilidade'):
        # Estatísticas básicas
        contratos_ativos = Contrato.objects.filter(
            imobiliaria=imob, status=StatusContrato.ATIVO
        ).count()
        valor_a_receber = Parcela.objects.filter(
            contrato__imobiliaria=imob, pago=False
        ).aggregate(total=Sum('valor_atual'))['total'] or Decimal('0.00')

        imobiliarias.append({
            'id': imob.id,
            'razao_social': imob.razao_social,
            'nome_fantasia': imob.nome_fantasia,
            'cnpj': imob.cnpj,
            'email': imob.email,
            'telefone': imob.telefone,
            'ativo': imob.ativo,
            'contabilidade': {
                'id': imob.contabilidade.id,
                'nome': imob.contabilidade.razao_social,
            } if imob.contabilidade else None,
            'contratos_ativos': contratos_ativos,
            'valor_a_receber': float(valor_a_receber),
        })

    return JsonResponse({
        'sucesso': True,
        'imobiliarias': imobiliarias,
        'total': len(imobiliarias),
    })


@login_required
@require_GET
def api_imobiliaria_dashboard(request, imobiliaria_id):
    """
    API para dados do dashboard de uma imobiliária específica.

    GET /api/imobiliaria/{id}/dashboard/
    """
    imobiliaria = get_object_or_404(Imobiliaria, pk=imobiliaria_id)
    hoje = timezone.now().date()

    # Contratos
    contratos_qs = Contrato.objects.filter(imobiliaria=imobiliaria)
    parcelas_qs = Parcela.objects.filter(contrato__imobiliaria=imobiliaria)

    stats_contratos = contratos_qs.aggregate(
        total=Count('id'),
        ativos=Count('id', filter=Q(status=StatusContrato.ATIVO)),
        quitados=Count('id', filter=Q(status=StatusContrato.QUITADO)),
        cancelados=Count('id', filter=Q(status=StatusContrato.CANCELADO)),
        valor_total=Sum('valor_total'),
    )

    stats_parcelas = parcelas_qs.aggregate(
        total=Count('id'),
        pagas=Count('id', filter=Q(pago=True)),
        pendentes=Count('id', filter=Q(pago=False)),
        vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
        valor_recebido=Sum('valor_pago', filter=Q(pago=True)),
        valor_pendente=Sum('valor_atual', filter=Q(pago=False)),
        valor_vencido=Sum('valor_atual', filter=Q(pago=False, data_vencimento__lt=hoje)),
    )

    # Próximas parcelas a vencer
    proximas_parcelas = parcelas_qs.filter(
        pago=False,
        data_vencimento__gte=hoje
    ).order_by('data_vencimento')[:10]

    proximas = [{
        'id': p.id,
        'contrato': p.contrato.numero_contrato,
        'comprador': p.contrato.comprador.nome,
        'numero_parcela': p.numero_parcela,
        'data_vencimento': p.data_vencimento.strftime('%Y-%m-%d'),
        'valor': float(p.valor_atual),
        'status_boleto': p.status_boleto,
    } for p in proximas_parcelas]

    # Parcelas vencidas
    parcelas_vencidas = parcelas_qs.filter(
        pago=False,
        data_vencimento__lt=hoje
    ).order_by('-data_vencimento')[:10]

    vencidas = [{
        'id': p.id,
        'contrato': p.contrato.numero_contrato,
        'comprador': p.contrato.comprador.nome,
        'numero_parcela': p.numero_parcela,
        'data_vencimento': p.data_vencimento.strftime('%Y-%m-%d'),
        'dias_atraso': (hoje - p.data_vencimento).days,
        'valor': float(p.valor_atual),
    } for p in parcelas_vencidas]

    # Contratos com reajuste pendente
    contratos_reajuste = []
    for contrato in contratos_qs.filter(status=StatusContrato.ATIVO):
        if hasattr(contrato, 'verificar_reajuste_necessario'):
            if contrato.verificar_reajuste_necessario():
                contratos_reajuste.append({
                    'id': contrato.id,
                    'numero': contrato.numero_contrato,
                    'comprador': contrato.comprador.nome,
                    'tipo_correcao': contrato.tipo_correcao,
                    'data_proximo_reajuste': contrato.data_proximo_reajuste.strftime('%Y-%m-%d') if contrato.data_proximo_reajuste else None,
                })

    return JsonResponse({
        'sucesso': True,
        'imobiliaria': {
            'id': imobiliaria.id,
            'nome': imobiliaria.nome_fantasia or imobiliaria.razao_social,
        },
        'contratos': stats_contratos,
        'parcelas': stats_parcelas,
        'proximas_parcelas': proximas,
        'parcelas_vencidas': vencidas,
        'reajustes_pendentes': contratos_reajuste,
    })


# =============================================================================
# APIs REST - CONTRATOS
# =============================================================================

@login_required
@require_GET
def api_contratos_lista(request):
    """
    API para listar contratos com filtros.

    GET /api/contratos/

    Parâmetros:
        - imobiliaria: ID da imobiliária (opcional)
        - status: Status do contrato (opcional)
        - comprador: ID do comprador (opcional)
        - page: Página (default 1)
        - per_page: Itens por página (default 20)
    """
    imobiliaria_id = request.GET.get('imobiliaria')
    status = request.GET.get('status')
    comprador_id = request.GET.get('comprador')
    try:
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(max(1, int(request.GET.get('per_page', 20))), 100)
    except (ValueError, TypeError):
        page, per_page = 1, 20

    queryset = Contrato.objects.all().select_related(
        'imobiliaria', 'comprador', 'imovel'
    )

    if imobiliaria_id:
        queryset = queryset.filter(imobiliaria_id=imobiliaria_id)
    if status:
        queryset = queryset.filter(status=status)
    if comprador_id:
        queryset = queryset.filter(comprador_id=comprador_id)

    total = queryset.count()
    offset = (page - 1) * per_page
    contratos_page = queryset.order_by('-data_contrato')[offset:offset + per_page]

    contratos = []
    for c in contratos_page:
        saldo_devedor = c.calcular_saldo_devedor()
        progresso = c.calcular_progresso()

        contratos.append({
            'id': c.id,
            'numero_contrato': c.numero_contrato,
            'data_contrato': c.data_contrato.strftime('%Y-%m-%d'),
            'status': c.status,
            'imobiliaria': {
                'id': c.imobiliaria.id,
                'nome': c.imobiliaria.nome_fantasia or c.imobiliaria.razao_social,
            },
            'comprador': {
                'id': c.comprador.id,
                'nome': c.comprador.nome,
                'documento': c.comprador.cpf or c.comprador.cnpj,
            },
            'imovel': {
                'id': c.imovel.id,
                'identificacao': c.imovel.identificacao,
            },
            'valor_total': float(c.valor_total),
            'valor_entrada': float(c.valor_entrada),
            'numero_parcelas': c.numero_parcelas,
            'saldo_devedor': float(saldo_devedor),
            'progresso': round(progresso, 2),
            'tipo_correcao': c.tipo_correcao,
        })

    return JsonResponse({
        'sucesso': True,
        'contratos': contratos,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page,
    })


@login_required
@require_GET
def api_contrato_detalhe(request, contrato_id):
    """
    API para detalhes de um contrato específico.

    GET /api/contratos/{id}/
    """
    contrato = get_object_or_404(
        Contrato.objects.select_related('imobiliaria', 'comprador', 'imovel'),
        pk=contrato_id
    )

    resumo = contrato.get_resumo_financeiro()

    return JsonResponse({
        'sucesso': True,
        'contrato': {
            'id': contrato.id,
            'numero_contrato': contrato.numero_contrato,
            'data_contrato': contrato.data_contrato.strftime('%Y-%m-%d'),
            'data_primeiro_vencimento': contrato.data_primeiro_vencimento.strftime('%Y-%m-%d'),
            'status': contrato.status,
            'imobiliaria': {
                'id': contrato.imobiliaria.id,
                'nome': contrato.imobiliaria.nome_fantasia or contrato.imobiliaria.razao_social,
                'cnpj': contrato.imobiliaria.cnpj,
            },
            'comprador': {
                'id': contrato.comprador.id,
                'nome': contrato.comprador.nome,
                'documento': contrato.comprador.cpf or contrato.comprador.cnpj,
                'email': contrato.comprador.email,
                'telefone': contrato.comprador.telefone,
            },
            'imovel': {
                'id': contrato.imovel.id,
                'identificacao': contrato.imovel.identificacao,
                'endereco': contrato.imovel.endereco,
            },
            'valores': {
                'total': float(contrato.valor_total),
                'entrada': float(contrato.valor_entrada),
                'financiado': float(contrato.valor_financiado),
                'parcela_original': float(contrato.valor_parcela_original),
            },
            'parcelas': {
                'numero_total': contrato.numero_parcelas,
                'dia_vencimento': contrato.dia_vencimento,
            },
            'reajuste': {
                'tipo_correcao': contrato.tipo_correcao,
                'prazo_meses': contrato.prazo_reajuste_meses,
                'data_ultimo': contrato.data_ultimo_reajuste.strftime('%Y-%m-%d') if contrato.data_ultimo_reajuste else None,
                'data_proximo': contrato.data_proximo_reajuste.strftime('%Y-%m-%d') if contrato.data_proximo_reajuste else None,
                'ciclo_atual': contrato.ciclo_reajuste_atual,
                'bloqueio_ativo': contrato.bloqueio_boleto_reajuste,
            },
            'encargos': {
                'juros_mora': float(contrato.percentual_juros_mora),
                'multa': float(contrato.percentual_multa),
            },
            'resumo_financeiro': {
                'valor_pago': float(resumo.get('total_pago', 0)),
                'valor_a_pagar': float(resumo.get('total_a_pagar', 0)),
                'valor_vencido': float(resumo.get('total_vencido', 0)),
                'parcelas_pagas': resumo.get('parcelas_pagas', 0),
                'parcelas_a_pagar': resumo.get('parcelas_a_pagar', 0),
                'parcelas_vencidas': resumo.get('parcelas_vencidas', 0),
                'saldo_devedor': float(resumo.get('saldo_devedor', 0)),
                'progresso': round(resumo.get('progresso_percentual', 0), 2),
            },
        },
    })


@login_required
@require_GET
def api_contrato_parcelas(request, contrato_id):
    """
    API para listar parcelas de um contrato.

    GET /api/contratos/{id}/parcelas/

    Parâmetros:
        - status: pago, pendente, vencido (opcional)
        - page: Página (default 1)
        - per_page: Itens por página (default 50)
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    status_filter = request.GET.get('status')
    try:
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(max(1, int(request.GET.get('per_page', 50))), 100)
    except (ValueError, TypeError):
        page, per_page = 1, 50
    hoje = timezone.now().date()

    queryset = contrato.parcelas.all()

    if status_filter == 'pago':
        queryset = queryset.filter(pago=True)
    elif status_filter == 'pendente':
        queryset = queryset.filter(pago=False, data_vencimento__gte=hoje)
    elif status_filter == 'vencido':
        queryset = queryset.filter(pago=False, data_vencimento__lt=hoje)

    total = queryset.count()
    offset = (page - 1) * per_page
    parcelas_page = queryset.order_by('numero_parcela')[offset:offset + per_page]

    parcelas = []
    for p in parcelas_page:
        parcelas.append({
            'id': p.id,
            'numero_parcela': p.numero_parcela,
            'tipo': p.tipo_parcela,
            'ciclo_reajuste': p.ciclo_reajuste,
            'data_vencimento': p.data_vencimento.strftime('%Y-%m-%d'),
            'valor_original': float(p.valor_original),
            'valor_atual': float(p.valor_atual),
            'valor_juros': float(p.valor_juros),
            'valor_multa': float(p.valor_multa),
            'valor_desconto': float(p.valor_desconto),
            'valor_total': float(p.valor_total),
            'pago': p.pago,
            'data_pagamento': p.data_pagamento.strftime('%Y-%m-%d') if p.data_pagamento else None,
            'valor_pago': float(p.valor_pago) if p.valor_pago else None,
            'dias_atraso': p.dias_atraso,
            'vencida': p.esta_vencida,
            'boleto': {
                'status': p.status_boleto,
                'nosso_numero': p.nosso_numero,
                'tem_boleto': p.tem_boleto,
            },
        })

    return JsonResponse({
        'sucesso': True,
        'contrato_id': contrato.id,
        'parcelas': parcelas,
        'total': total,
        'page': page,
        'per_page': per_page,
    })


@login_required
@require_GET
def api_contrato_reajustes(request, contrato_id):
    """
    API para listar histórico de reajustes de um contrato.

    GET /api/contratos/{id}/reajustes/
    """
    contrato = get_object_or_404(Contrato, pk=contrato_id)

    reajustes = Reajuste.objects.filter(contrato=contrato).order_by('-data_reajuste')

    lista_reajustes = [{
        'id': r.id,
        'data_reajuste': r.data_reajuste.strftime('%Y-%m-%d'),
        'indice_tipo': r.indice_tipo,
        'percentual': float(r.percentual),
        'ciclo': r.ciclo,
        'parcela_inicial': r.parcela_inicial,
        'parcela_final': r.parcela_final,
        'aplicado': r.aplicado,
        'data_aplicacao': r.data_aplicacao.strftime('%Y-%m-%d %H:%M') if r.data_aplicacao else None,
        'aplicado_manual': r.aplicado_manual,
        'observacoes': r.observacoes,
    } for r in reajustes]

    # Informações de reajuste pendente
    reajuste_pendente = None
    if contrato.verificar_reajuste_necessario():
        reajuste_pendente = {
            'tipo_correcao': contrato.tipo_correcao,
            'ciclo_atual': contrato.ciclo_reajuste_atual,
            'data_proximo': contrato.data_proximo_reajuste.strftime('%Y-%m-%d') if contrato.data_proximo_reajuste else None,
            'bloqueio_ativo': contrato.bloqueio_boleto_reajuste,
        }

    return JsonResponse({
        'sucesso': True,
        'contrato_id': contrato.id,
        'reajustes': lista_reajustes,
        'total': len(lista_reajustes),
        'reajuste_pendente': reajuste_pendente,
    })


# =============================================================================
# APIs REST - PARCELAS
# =============================================================================

@login_required
@require_GET
def api_parcelas_lista(request):
    """
    API para listar parcelas com filtros avançados.

    GET /api/parcelas/

    Parâmetros:
        - imobiliaria: ID da imobiliária (opcional)
        - contrato: ID do contrato (opcional)
        - status: pago, pendente, vencido (opcional)
        - data_inicio: Data inicial de vencimento (YYYY-MM-DD)
        - data_fim: Data final de vencimento (YYYY-MM-DD)
        - page: Página (default 1)
        - per_page: Itens por página (default 50)
    """
    imobiliaria_id = request.GET.get('imobiliaria')
    contrato_id = request.GET.get('contrato')
    status_filter = request.GET.get('status')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    try:
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(max(1, int(request.GET.get('per_page', 50))), 100)
    except (ValueError, TypeError):
        page, per_page = 1, 50

    hoje = timezone.now().date()
    queryset = Parcela.objects.all().select_related(
        'contrato', 'contrato__comprador', 'contrato__imobiliaria'
    )

    if imobiliaria_id:
        queryset = queryset.filter(contrato__imobiliaria_id=imobiliaria_id)
    if contrato_id:
        queryset = queryset.filter(contrato_id=contrato_id)

    if status_filter == 'pago':
        queryset = queryset.filter(pago=True)
    elif status_filter == 'pendente':
        queryset = queryset.filter(pago=False, data_vencimento__gte=hoje)
    elif status_filter == 'vencido':
        queryset = queryset.filter(pago=False, data_vencimento__lt=hoje)

    if data_inicio:
        try:
            from datetime import datetime
            queryset = queryset.filter(
                data_vencimento__gte=datetime.strptime(data_inicio, '%Y-%m-%d').date()
            )
        except ValueError:
            pass
    if data_fim:
        try:
            from datetime import datetime
            queryset = queryset.filter(
                data_vencimento__lte=datetime.strptime(data_fim, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    total = queryset.count()
    offset = (page - 1) * per_page
    parcelas_page = queryset.order_by('data_vencimento')[offset:offset + per_page]

    # Totalizadores
    totais = queryset.aggregate(
        valor_total=Sum('valor_atual'),
        valor_pago=Sum('valor_pago', filter=Q(pago=True)),
    )

    parcelas = [{
        'id': p.id,
        'contrato': {
            'id': p.contrato.id,
            'numero': p.contrato.numero_contrato,
        },
        'comprador': p.contrato.comprador.nome,
        'imobiliaria': p.contrato.imobiliaria.nome_fantasia or p.contrato.imobiliaria.razao_social,
        'numero_parcela': p.numero_parcela,
        'data_vencimento': p.data_vencimento.strftime('%Y-%m-%d'),
        'valor_atual': float(p.valor_atual),
        'valor_total': float(p.valor_total),
        'pago': p.pago,
        'data_pagamento': p.data_pagamento.strftime('%Y-%m-%d') if p.data_pagamento else None,
        'dias_atraso': p.dias_atraso,
        'status_boleto': p.status_boleto,
    } for p in parcelas_page]

    return JsonResponse({
        'sucesso': True,
        'parcelas': parcelas,
        'total': total,
        'page': page,
        'per_page': per_page,
        'totais': {
            'valor_total': float(totais['valor_total'] or 0),
            'valor_pago': float(totais['valor_pago'] or 0),
        },
    })


@login_required
@require_POST
def api_parcela_registrar_pagamento(request, parcela_id):
    """
    API para registrar pagamento de uma parcela.

    POST /api/parcelas/{id}/pagamento/

    Body JSON:
        - valor_pago: Valor pago (obrigatório)
        - data_pagamento: Data do pagamento (opcional, default hoje)
        - forma_pagamento: Forma de pagamento (opcional)
        - observacoes: Observações (opcional)
    """
    import json
    parcela = get_object_or_404(Parcela, pk=parcela_id)

    if parcela.pago:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Esta parcela já está paga.'
        }, status=400)

    try:
        data = json.loads(request.body)

        valor_pago = Decimal(str(data.get('valor_pago', 0)))
        if valor_pago <= 0:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Informe um valor pago válido.'
            }, status=400)

        data_pagamento = data.get('data_pagamento')
        if data_pagamento:
            from datetime import datetime
            data_pagamento = datetime.strptime(data_pagamento, '%Y-%m-%d').date()
        else:
            data_pagamento = timezone.now().date()

        observacoes = data.get('observacoes', '')
        forma_pagamento = data.get('forma_pagamento', 'DINHEIRO')

        # Registrar pagamento
        parcela.registrar_pagamento(
            valor_pago=valor_pago,
            data_pagamento=data_pagamento,
            observacoes=observacoes
        )

        # Criar histórico
        from financeiro.models import HistoricoPagamento
        HistoricoPagamento.objects.create(
            parcela=parcela,
            data_pagamento=data_pagamento,
            valor_pago=valor_pago,
            valor_parcela=parcela.valor_atual,
            valor_juros=parcela.valor_juros,
            valor_multa=parcela.valor_multa,
            valor_desconto=parcela.valor_desconto,
            forma_pagamento=forma_pagamento,
            observacoes=observacoes,
            origem_pagamento='MANUAL',
        )

        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Pagamento de R$ {valor_pago} registrado com sucesso.',
            'parcela': {
                'id': parcela.id,
                'numero_parcela': parcela.numero_parcela,
                'pago': parcela.pago,
                'data_pagamento': parcela.data_pagamento.strftime('%Y-%m-%d'),
                'valor_pago': float(parcela.valor_pago),
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Dados inválidos.'
        }, status=400)
    except Exception as e:
        logger.exception(f"Erro ao registrar pagamento: {e}")
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


# =============================================================================
# APIs REST - BOLETO E CNAB
# =============================================================================

@login_required
def api_boleto_gerar(request, parcela_id):
    """
    API para gerar boleto de uma parcela.

    POST /api/boletos/{parcela_id}/gerar/

    Body JSON (opcional):
        - conta_bancaria_id: ID da conta bancária (usa padrão se não informado)
        - force: Força regeneração mesmo se já existe (default: false)
        - enviar_email: Envia email ao comprador (default: true)

    Returns:
        - sucesso: boolean
        - boleto: dados do boleto gerado
        - erro: mensagem de erro (se houver)
    """
    import json

    parcela = get_object_or_404(Parcela, pk=parcela_id)

    if request.method != 'POST':
        return JsonResponse({
            'sucesso': False,
            'erro': 'Método não permitido. Use POST.'
        }, status=405)

    try:
        data = json.loads(request.body) if request.body else {}

        conta_bancaria_id = data.get('conta_bancaria_id')
        force = data.get('force', False)
        enviar_email = data.get('enviar_email', True)

        if parcela.pago:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Parcela já está paga.'
            }, status=400)

        pode_gerar, motivo = parcela.pode_gerar_boleto()
        if not pode_gerar and not force:
            return JsonResponse({
                'sucesso': False,
                'erro': motivo,
                'bloqueado_reajuste': True
            }, status=400)

        if parcela.tem_boleto and not force:
            return JsonResponse({
                'sucesso': True,
                'mensagem': 'Boleto já existe.',
                'boleto': {
                    'parcela_id': parcela.id,
                    'nosso_numero': parcela.nosso_numero,
                    'linha_digitavel': parcela.linha_digitavel,
                    'codigo_barras': parcela.codigo_barras,
                    'valor': float(parcela.valor_boleto or parcela.valor_atual),
                    'data_vencimento': parcela.data_vencimento.strftime('%Y-%m-%d'),
                    'status': parcela.status_boleto,
                    'tem_pdf': bool(parcela.boleto_pdf),
                }
            })

        conta_bancaria = None
        if conta_bancaria_id:
            conta_bancaria = get_object_or_404(ContaBancaria, pk=conta_bancaria_id, ativo=True)

        resultado = parcela.gerar_boleto(
            conta_bancaria=conta_bancaria,
            force=force,
            enviar_email=enviar_email
        )

        if resultado and resultado.get('sucesso'):
            parcela.refresh_from_db()
            return JsonResponse({
                'sucesso': True,
                'mensagem': 'Boleto gerado com sucesso.',
                'boleto': {
                    'parcela_id': parcela.id,
                    'nosso_numero': parcela.nosso_numero,
                    'linha_digitavel': parcela.linha_digitavel,
                    'codigo_barras': parcela.codigo_barras,
                    'valor': float(parcela.valor_boleto or parcela.valor_atual),
                    'data_vencimento': parcela.data_vencimento.strftime('%Y-%m-%d'),
                    'status': parcela.status_boleto,
                    'tem_pdf': bool(parcela.boleto_pdf),
                    'pix_copia_cola': parcela.pix_copia_cola or None,
                }
            })
        else:
            return JsonResponse({
                'sucesso': False,
                'erro': resultado.get('erro') if resultado else 'Erro ao gerar boleto.'
            }, status=400)

    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)
    except Exception as e:
        logger.exception(f"Erro ao gerar boleto via API: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
def api_boleto_detalhe(request, parcela_id):
    """
    API para obter detalhes do boleto de uma parcela.
    GET /api/boletos/{parcela_id}/
    """
    parcela = get_object_or_404(Parcela, pk=parcela_id)

    if not parcela.tem_boleto:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Esta parcela não possui boleto gerado.'
        }, status=404)

    return JsonResponse({
        'sucesso': True,
        'boleto': {
            'parcela_id': parcela.id,
            'contrato_id': parcela.contrato_id,
            'numero_parcela': parcela.numero_parcela,
            'nosso_numero': parcela.nosso_numero,
            'numero_documento': parcela.numero_documento,
            'linha_digitavel': parcela.linha_digitavel,
            'codigo_barras': parcela.codigo_barras,
            'valor': float(parcela.valor_boleto or parcela.valor_atual),
            'valor_original': float(parcela.valor_original),
            'data_vencimento': parcela.data_vencimento.strftime('%Y-%m-%d'),
            'status': parcela.status_boleto,
            'status_display': parcela.get_status_boleto_display(),
            'pago': parcela.pago,
            'data_pagamento': parcela.data_pagamento.strftime('%Y-%m-%d') if parcela.data_pagamento else None,
            'valor_pago': float(parcela.valor_pago) if parcela.valor_pago else None,
            'data_geracao': parcela.data_geracao_boleto.isoformat() if parcela.data_geracao_boleto else None,
            'tem_pdf': bool(parcela.boleto_pdf),
            'boleto_url': parcela.boleto_url or None,
            'pix_copia_cola': parcela.pix_copia_cola or None,
            'comprador': {
                'nome': parcela.contrato.comprador.nome,
                'documento': parcela.contrato.comprador.cpf or parcela.contrato.comprador.cnpj,
            }
        }
    })


@login_required
@require_POST
def api_boleto_cancelar(request, parcela_id):
    """
    API para cancelar boleto de uma parcela.
    POST /api/boletos/{parcela_id}/cancelar/
    Body: {"motivo": "texto"}
    """
    import json
    parcela = get_object_or_404(Parcela, pk=parcela_id)

    if not parcela.tem_boleto:
        return JsonResponse({'sucesso': False, 'erro': 'Sem boleto para cancelar.'}, status=400)
    if parcela.pago:
        return JsonResponse({'sucesso': False, 'erro': 'Parcela já paga.'}, status=400)

    try:
        data = json.loads(request.body) if request.body else {}
        motivo = data.get('motivo', '')
        if not motivo:
            return JsonResponse({'sucesso': False, 'erro': 'Informe o motivo.'}, status=400)

        parcela.cancelar_boleto(motivo=motivo)
        return JsonResponse({'sucesso': True, 'mensagem': 'Boleto cancelado.'})

    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)
    except Exception as e:
        logger.exception(f"Erro ao cancelar boleto: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
def api_boletos_lote(request):
    """
    API para gerar boletos em lote.
    POST /api/boletos/lote/
    Body: {"parcela_ids": [1,2,3], "conta_bancaria_id": 1, "force": false}
    """
    import json

    if request.method != 'POST':
        return JsonResponse({'sucesso': False, 'erro': 'Use POST.'}, status=405)

    try:
        data = json.loads(request.body)
        parcela_ids = data.get('parcela_ids', [])
        conta_bancaria_id = data.get('conta_bancaria_id')
        force = data.get('force', False)
        enviar_email = data.get('enviar_email', False)

        if not parcela_ids:
            return JsonResponse({'sucesso': False, 'erro': 'Informe parcelas.'}, status=400)

        parcelas = Parcela.objects.filter(pk__in=parcela_ids, pago=False)
        if not parcelas.exists():
            return JsonResponse({'sucesso': False, 'erro': 'Nenhuma parcela válida.'}, status=400)

        conta_bancaria = None
        if conta_bancaria_id:
            conta_bancaria = get_object_or_404(ContaBancaria, pk=conta_bancaria_id, ativo=True)

        resultados = []
        gerados = erros = bloqueados = 0

        for parcela in parcelas:
            pode, motivo = parcela.pode_gerar_boleto()
            if not pode and not force:
                resultados.append({'parcela_id': parcela.id, 'sucesso': False, 'bloqueado': True, 'erro': motivo})
                bloqueados += 1
                continue

            if parcela.tem_boleto and not force:
                resultados.append({'parcela_id': parcela.id, 'sucesso': True, 'ja_existe': True})
                continue

            try:
                res = parcela.gerar_boleto(conta_bancaria=conta_bancaria, force=force, enviar_email=enviar_email)
                if res and res.get('sucesso'):
                    gerados += 1
                    resultados.append({'parcela_id': parcela.id, 'sucesso': True, 'nosso_numero': res.get('nosso_numero')})
                else:
                    erros += 1
                    resultados.append({'parcela_id': parcela.id, 'sucesso': False, 'erro': res.get('erro') if res else 'Erro'})
            except Exception as e:
                logger.exception('Erro ao gerar boleto parcela pk=%s: %s', parcela.id, e)
                erros += 1
                resultados.append({'parcela_id': parcela.id, 'sucesso': False, 'erro': str(e)})

        return JsonResponse({'sucesso': True, 'gerados': gerados, 'erros': erros, 'bloqueados': bloqueados, 'resultados': resultados})

    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)
    except Exception as e:
        logger.exception(f"Erro em lote: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


# =============================================================================
# APIs REST - REVALIDAR BOLETOS
# =============================================================================

@login_required
@require_POST
def api_boletos_revalidar(request):
    """
    Revalida boletos com status GERADO, verificando se todos os dados
    necessários ainda estão presentes e consistentes.

    POST /api/boletos/revalidar/
    Body: {"parcela_ids": [1,2,3]}  -- se vazio, revalida todos com status GERADO
    """
    import json

    CAMPOS_OBRIG_BANCO = {
        '001': ['convenio'],
        '033': ['convenio'],
        '104': ['convenio'],
        '748': ['posto', 'byte_idt'],
        '756': [],
    }

    NOMES_BANCO = {
        '001': 'Banco do Brasil',
        '033': 'Santander',
        '104': 'Caixa Econômica',
        '237': 'Bradesco',
        '341': 'Itaú',
        '748': 'Sicredi',
        '756': 'Sicoob',
    }

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)

    parcela_ids = data.get('parcela_ids', [])

    qs = Parcela.objects.filter(
        status_boleto=StatusBoleto.GERADO
    ).select_related(
        'conta_bancaria',
        'contrato',
        'contrato__comprador',
        'contrato__imovel',
        'contrato__imovel__imobiliaria',
    )

    if parcela_ids:
        qs = qs.filter(pk__in=parcela_ids)

    resultados = []
    validos = invalidos = 0

    for parcela in qs:
        erros_parcela = []

        # 1. Nosso número
        if not parcela.nosso_numero:
            erros_parcela.append('Sem nosso_numero')

        # 2. Conta bancária
        conta = parcela.conta_bancaria
        if not conta:
            erros_parcela.append('Sem conta bancária associada')
        else:
            banco = conta.banco
            nome_banco = NOMES_BANCO.get(banco, f'Banco {banco}')

            # Campos obrigatórios por banco
            campos_obrig = CAMPOS_OBRIG_BANCO.get(banco, [])
            for campo in campos_obrig:
                valor = getattr(conta, campo, '') or ''
                if not valor.strip():
                    erros_parcela.append(f'{nome_banco}: campo "{campo}" obrigatório ausente na conta bancária')

            # Agência e conta corrente básicos
            if not (conta.agencia or '').strip():
                erros_parcela.append('Agência bancária ausente')
            if not (conta.conta or '').strip():
                erros_parcela.append('Conta corrente ausente')

            # CNPJ da imobiliária
            imobiliaria = getattr(conta, 'imobiliaria', None)
            if imobiliaria and not (imobiliaria.cnpj or '').strip():
                erros_parcela.append('CNPJ da imobiliária ausente')

        # 3. Dados do comprador
        comprador = parcela.contrato.comprador if parcela.contrato_id else None
        if comprador:
            doc = (comprador.cpf or '') or (getattr(comprador, 'cnpj', '') or '')
            if not doc.strip():
                erros_parcela.append('Comprador sem CPF/CNPJ')
        else:
            erros_parcela.append('Sem comprador no contrato')

        if erros_parcela:
            invalidos += 1
            resultados.append({
                'parcela_id': parcela.id,
                'valido': False,
                'erros': erros_parcela,
                'contrato': parcela.contrato.numero_contrato if parcela.contrato_id else '-',
                'nosso_numero': parcela.nosso_numero or '-',
            })
        else:
            validos += 1
            resultados.append({
                'parcela_id': parcela.id,
                'valido': True,
                'erros': [],
                'contrato': parcela.contrato.numero_contrato if parcela.contrato_id else '-',
                'nosso_numero': parcela.nosso_numero,
            })

    total = validos + invalidos
    return JsonResponse({
        'sucesso': True,
        'total': total,
        'validos': validos,
        'invalidos': invalidos,
        'resultados': resultados,
    })


# =============================================================================
# APIs REST - CNAB REMESSA
# =============================================================================

@login_required
def api_cnab_remessa_listar(request):
    """API para listar remessas CNAB. GET /api/cnab/remessas/"""
    import json
    from .models import ArquivoRemessa

    CAMPOS_OBRIG_BANCO = {
        '001': ['convenio'],
        '033': ['convenio'],
        '104': ['convenio'],
        '748': ['posto', 'byte_idt'],
        '756': [],
    }

    qs = ArquivoRemessa.objects.select_related('conta_bancaria', 'conta_bancaria__imobiliaria').order_by('-data_geracao')

    if request.GET.get('conta_bancaria_id'):
        qs = qs.filter(conta_bancaria_id=request.GET['conta_bancaria_id'])
    if request.GET.get('status'):
        qs = qs.filter(status=request.GET['status'])

    NOMES_BANCO = {
        '001': 'Banco do Brasil',
        '033': 'Santander',
        '104': 'Caixa Econômica',
        '237': 'Bradesco',
        '341': 'Itaú',
        '748': 'Sicredi',
        '756': 'Sicoob',
    }

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)

    parcela_ids = data.get('parcela_ids', [])

    qs = Parcela.objects.filter(
        status_boleto=StatusBoleto.GERADO
    ).select_related(
        'conta_bancaria',
        'contrato',
        'contrato__comprador',
        'contrato__imovel',
        'contrato__imovel__imobiliaria',
    )

    if parcela_ids:
        qs = qs.filter(pk__in=parcela_ids)

    resultados = []
    validos = invalidos = 0

    for parcela in qs:
        erros_parcela = []

        # 1. Nosso número
        if not parcela.nosso_numero:
            erros_parcela.append('Sem nosso_numero')

        # 2. Conta bancária
        conta = parcela.conta_bancaria
        if not conta:
            erros_parcela.append('Sem conta bancária associada')
        else:
            banco = conta.banco
            nome_banco = NOMES_BANCO.get(banco, f'Banco {banco}')

            # Campos obrigatórios por banco
            campos_obrig = CAMPOS_OBRIG_BANCO.get(banco, [])
            for campo in campos_obrig:
                valor = getattr(conta, campo, '') or ''
                if not valor.strip():
                    erros_parcela.append(f'{nome_banco}: campo "{campo}" obrigatório ausente na conta bancária')

            # Agência e conta corrente básicos
            if not (conta.agencia or '').strip():
                erros_parcela.append('Agência bancária ausente')
            if not (conta.conta or '').strip():
                erros_parcela.append('Conta corrente ausente')

            # CNPJ da imobiliária
            imobiliaria = getattr(conta, 'imobiliaria', None)
            if imobiliaria and not (imobiliaria.cnpj or '').strip():
                erros_parcela.append('CNPJ da imobiliária ausente')

        # 3. Dados do comprador
        comprador = parcela.contrato.comprador if parcela.contrato_id else None
        if comprador:
            doc = (comprador.cpf or '') or (getattr(comprador, 'cnpj', '') or '')
            if not doc.strip():
                erros_parcela.append('Comprador sem CPF/CNPJ')
        else:
            erros_parcela.append('Sem comprador no contrato')

        if erros_parcela:
            invalidos += 1
            resultados.append({
                'parcela_id': parcela.id,
                'valido': False,
                'erros': erros_parcela,
                'contrato': parcela.contrato.numero_contrato if parcela.contrato_id else '-',
                'nosso_numero': parcela.nosso_numero or '-',
            })
        else:
            validos += 1
            resultados.append({
                'parcela_id': parcela.id,
                'valido': True,
                'erros': [],
                'contrato': parcela.contrato.numero_contrato if parcela.contrato_id else '-',
                'nosso_numero': parcela.nosso_numero,
            })

    total = validos + invalidos
    return JsonResponse({
        'sucesso': True,
        'total': total,
        'validos': validos,
        'invalidos': invalidos,
        'resultados': resultados,
    })


# =============================================================================
# APIs REST - CNAB REMESSA
# =============================================================================

@login_required
def api_cnab_remessa_listar(request):  # noqa: F811
    """API para listar remessas CNAB. GET /api/cnab/remessas/"""
    from .models import ArquivoRemessa

    qs = ArquivoRemessa.objects.select_related('conta_bancaria', 'conta_bancaria__imobiliaria').order_by('-data_geracao')

    if request.GET.get('conta_bancaria_id'):
        qs = qs.filter(conta_bancaria_id=request.GET['conta_bancaria_id'])
    if request.GET.get('status'):
        qs = qs.filter(status=request.GET['status'])

    try:
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(max(1, int(request.GET.get('per_page', 20))), 100)
    except (ValueError, TypeError):
        page, per_page = 1, 20
    total = qs.count()
    arquivos = qs[(page-1)*per_page:page*per_page]

    remessas = [{
        'id': a.id, 'numero_remessa': a.numero_remessa, 'layout': a.layout,
        'nome_arquivo': a.nome_arquivo, 'status': a.status,
        'data_geracao': a.data_geracao.isoformat() if a.data_geracao else None,
        'quantidade_boletos': a.quantidade_boletos, 'valor_total': float(a.valor_total),
        'conta_bancaria': {'id': a.conta_bancaria.id, 'banco': a.conta_bancaria.banco},
    } for a in arquivos]

    return JsonResponse({'sucesso': True, 'remessas': remessas, 'total': total, 'page': page})


@login_required
def api_cnab_remessa_detalhe(request, remessa_id):
    """API para detalhes de remessa. GET /api/cnab/remessas/{id}/"""
    from .models import ArquivoRemessa

    arq = get_object_or_404(ArquivoRemessa.objects.select_related('conta_bancaria').prefetch_related('itens', 'itens__parcela'), pk=remessa_id)

    itens = [{'id': i.id, 'parcela_id': i.parcela_id, 'nosso_numero': i.nosso_numero,
              'valor': float(i.valor), 'data_vencimento': i.data_vencimento.strftime('%Y-%m-%d')} for i in arq.itens.all()]

    return JsonResponse({
        'sucesso': True,
        'remessa': {
            'id': arq.id, 'numero_remessa': arq.numero_remessa, 'layout': arq.layout,
            'nome_arquivo': arq.nome_arquivo, 'status': arq.status,
            'quantidade_boletos': arq.quantidade_boletos, 'valor_total': float(arq.valor_total),
            'pode_reenviar': arq.pode_reenviar,
        },
        'itens': itens
    })


@login_required
@require_POST
def api_cnab_remessa_gerar(request):
    """
    API para gerar remessa CNAB.
    POST /api/cnab/remessas/gerar/
    Body: {"conta_bancaria_id": 1, "parcela_ids": [1,2,3], "layout": "CNAB_240"}
    """
    import json
    from .models import StatusBoleto
    from .services.cnab_service import CNABService

    try:
        data = json.loads(request.body)
        conta_id = data.get('conta_bancaria_id')
        parcela_ids = data.get('parcela_ids', [])
        layout = data.get('layout', 'CNAB_240')

        if not conta_id:
            return JsonResponse({'sucesso': False, 'erro': 'Informe conta bancária.'}, status=400)
        if not parcela_ids:
            return JsonResponse({'sucesso': False, 'erro': 'Informe parcelas.'}, status=400)

        conta = get_object_or_404(ContaBancaria, pk=conta_id, ativo=True)

        # Contar pagas para informar ao chamador
        pagas_count = Parcela.objects.filter(pk__in=parcela_ids, pago=True).count()

        parcelas = Parcela.objects.filter(pk__in=parcela_ids, status_boleto=StatusBoleto.GERADO, pago=False)

        if not parcelas.exists():
            msg = 'Nenhuma parcela válida.'
            if pagas_count:
                msg = f'Todas as {pagas_count} parcela(s) selecionada(s) estão PAGAS. Parcelas pagas não entram em remessa.'
            return JsonResponse({'sucesso': False, 'erro': msg}, status=400)

        service = CNABService()
        resultado = service.gerar_remessa(list(parcelas), conta, layout)

        if resultado.get('sucesso'):
            arq = resultado['arquivo_remessa']
            return JsonResponse({
                'sucesso': True,
                'remessa': {
                    'id': arq.id, 'numero_remessa': arq.numero_remessa, 'nome_arquivo': arq.nome_arquivo,
                    'quantidade_boletos': resultado.get('quantidade_boletos'),
                    'valor_total': float(resultado.get('valor_total', 0))}
            })
        return JsonResponse({'sucesso': False, 'erro': resultado.get('erro')}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({'sucesso': False, 'erro': 'JSON inválido.'}, status=400)
    except Exception as e:
        logger.exception(f"Erro ao gerar remessa: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
@require_GET
def api_cnab_boletos_pendentes_count(request):
    """Contagem rápida de boletos disponíveis para nova remessa (badge navbar)."""
    try:
        from .services.cnab_service import CNABService
        count = len(CNABService().obter_boletos_sem_remessa())
        return JsonResponse({'count': count})
    except Exception as e:
        logger.exception("Erro em api_cnab_boletos_pendentes_count: %s", e)
        return JsonResponse({'count': 0})


@login_required
def api_cnab_boletos_disponiveis(request):
    """
    API para listar boletos disponíveis para remessa.
    GET /api/cnab/boletos-disponiveis/
    Params: conta_bancaria_id, imobiliaria_id, contrato_id (todos opcionais)
    """
    from .services.cnab_service import CNABService

    conta_id = request.GET.get('conta_bancaria_id')
    imobiliaria_id = request.GET.get('imobiliaria_id') or None
    contrato_id = request.GET.get('contrato_id') or None

    conta = None
    if conta_id:
        conta = get_object_or_404(ContaBancaria, pk=conta_id, ativo=True)

    service = CNABService()
    boletos = service.obter_boletos_sem_remessa(
        conta_bancaria=conta,
        imobiliaria_id=imobiliaria_id,
        contrato_id=contrato_id,
    )

    data = [{
        'parcela_id': p.id,
        'numero_contrato': p.contrato.numero_contrato,
        'numero_parcela': p.numero_parcela,
        'nosso_numero': p.nosso_numero,
        'valor': float(p.valor_boleto or p.valor_atual),
        'data_vencimento': p.data_vencimento.strftime('%Y-%m-%d'),
        'comprador': p.contrato.comprador.nome,
        'conta_bancaria_id': p.conta_bancaria_id,
        'conta_bancaria': str(p.conta_bancaria) if p.conta_bancaria else None,
    } for p in boletos]

    return JsonResponse({'sucesso': True, 'boletos': data, 'total': len(data)})


# =============================================================================
# APIs REST - CNAB RETORNO
# =============================================================================

@login_required
def api_cnab_retorno_listar(request):
    """API para listar retornos CNAB. GET /api/cnab/retornos/"""
    from .models import ArquivoRetorno

    qs = ArquivoRetorno.objects.select_related('conta_bancaria').order_by('-data_upload')

    if request.GET.get('conta_bancaria_id'):
        qs = qs.filter(conta_bancaria_id=request.GET['conta_bancaria_id'])
    if request.GET.get('status'):
        qs = qs.filter(status=request.GET['status'])

    try:
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(max(1, int(request.GET.get('per_page', 20))), 100)
    except (ValueError, TypeError):
        page, per_page = 1, 20
    total = qs.count()
    arquivos = qs[(page-1)*per_page:page*per_page]

    retornos = [{
        'id': a.id, 'nome_arquivo': a.nome_arquivo, 'layout': a.layout, 'status': a.status,
        'data_upload': a.data_upload.isoformat() if a.data_upload else None,
        'total_registros': a.total_registros, 'registros_processados': a.registros_processados,
        'valor_total_pago': float(a.valor_total_pago),
    } for a in arquivos]

    return JsonResponse({'sucesso': True, 'retornos': retornos, 'total': total, 'page': page})


@login_required
def api_cnab_retorno_detalhe(request, retorno_id):
    """API para detalhes de retorno. GET /api/cnab/retornos/{id}/"""
    from .models import ArquivoRetorno

    arq = get_object_or_404(ArquivoRetorno.objects.select_related('conta_bancaria').prefetch_related('itens'), pk=retorno_id)

    itens = [{'id': i.id, 'nosso_numero': i.nosso_numero, 'codigo_ocorrencia': i.codigo_ocorrencia,
              'tipo_ocorrencia': i.tipo_ocorrencia, 'valor_pago': float(i.valor_pago) if i.valor_pago else None,
              'data_credito': i.data_credito.strftime('%Y-%m-%d') if i.data_credito else None, 'processado': i.processado} for i in arq.itens.all()]

    return JsonResponse({
        'sucesso': True,
        'retorno': {
            'id': arq.id, 'nome_arquivo': arq.nome_arquivo, 'layout': arq.layout, 'status': arq.status,
            'total_registros': arq.total_registros, 'registros_processados': arq.registros_processados,
            'registros_erro': arq.registros_erro, 'valor_total_pago': float(arq.valor_total_pago),
        },
        'itens': itens
    })


@login_required
@require_POST
def api_cnab_retorno_processar(request, retorno_id):
    """API para processar retorno CNAB. POST /api/cnab/retornos/{id}/processar/"""
    from .models import ArquivoRetorno, StatusArquivoRetorno
    from .services.cnab_service import CNABService

    arquivo = get_object_or_404(ArquivoRetorno, pk=retorno_id)

    if arquivo.status == StatusArquivoRetorno.PROCESSADO:
        return JsonResponse({'sucesso': False, 'erro': 'Já processado.'}, status=400)

    try:
        service = CNABService()
        resultado = service.processar_retorno(arquivo)

        if resultado.get('sucesso'):
            return JsonResponse({
                'sucesso': True, 'total_registros': resultado.get('total_registros', 0),
                'processados': resultado.get('processados', 0), 'liquidacoes': resultado.get('liquidacoes', 0),
            })
        return JsonResponse({'sucesso': False, 'erro': resultado.get('erro')}, status=400)

    except Exception as e:
        logger.exception(f"Erro ao processar retorno: {e}")
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
def api_contas_bancarias(request):
    """API para listar contas bancárias. GET /api/contas-bancarias/"""
    qs = ContaBancaria.objects.filter(ativo=True).select_related('imobiliaria')

    if request.GET.get('imobiliaria_id'):
        qs = qs.filter(imobiliaria_id=request.GET['imobiliaria_id'])

    contas = [{
        'id': c.id, 'banco': c.banco, 'descricao': c.descricao, 'agencia': c.agencia, 'conta': c.conta,
        'convenio': c.convenio, 'carteira': c.carteira, 'layout_cnab': c.layout_cnab, 'principal': c.principal,
        'imobiliaria': {'id': c.imobiliaria.id, 'nome': c.imobiliaria.nome}
    } for c in qs]

    return JsonResponse({'sucesso': True, 'contas': contas})


# ==========================================================================
# API - NOTIFICAÇÕES
# ==========================================================================

@login_required
@require_GET
def api_reajustes_pendentes_count(request):
    """Retorna contagem de contratos ativos com reajuste pendente (badge navbar).

    Usa prefetch_related para evitar N×M queries: 2 queries no total
    em vez de uma por ciclo por contrato.
    """
    try:
        from contratos.models import TipoCorrecao
        from django.db.models import Prefetch

        hoje = timezone.localdate()

        contratos_ativos = Contrato.objects.filter(
            status=StatusContrato.ATIVO
        ).exclude(tipo_correcao=TipoCorrecao.FIXO).only(
            'pk', 'data_contrato', 'numero_parcelas', 'prazo_reajuste_meses', 'tipo_correcao'
        ).prefetch_related(
            Prefetch(
                'reajustes',
                queryset=Reajuste.objects.filter(aplicado=True).only('contrato_id', 'ciclo'),
                to_attr='reajustes_aplicados_cache',
            )
        )

        count = 0
        for contrato in contratos_ativos:
            # Pular contratos com dados incompletos
            if not contrato.data_contrato or not contrato.numero_parcelas:
                continue

            ciclos_aplicados = {r.ciclo for r in contrato.reajustes_aplicados_cache}
            prazo = contrato.prazo_reajuste_meses or 12
            total_ciclos = (contrato.numero_parcelas - 1) // prazo + 1

            for ciclo in range(2, total_ciclos + 1):
                data_reajuste = contrato.data_contrato + relativedelta(months=(ciclo - 1) * prazo)
                if hoje < data_reajuste:
                    break
                if ciclo not in ciclos_aplicados:
                    count += 1
                    break

        return JsonResponse({'count': count})
    except Exception as e:
        logger.exception("Erro em api_reajustes_pendentes_count: %s", e)
        return JsonResponse({'count': 0})


@login_required
@require_GET
def api_sidebar_pendencias(request):
    """Retorna todos os contadores de pendências para a sidebar em uma única chamada."""
    from contratos.models import TipoCorrecao
    from django.db.models import Prefetch
    from .services.cnab_service import CNABService

    hoje = timezone.localdate()

    # 1. Parcelas vencidas não pagas
    try:
        parcelas_vencidas = Parcela.objects.filter(
            pago=False, data_vencimento__lt=hoje
        ).count()
    except Exception:
        parcelas_vencidas = 0

    # 2. Boletos não gerados (parcelas em aberto, vencimento nos próximos 30 dias)
    try:
        boletos_nao_gerados = Parcela.objects.filter(
            pago=False,
            boleto_gerado=False,
            data_vencimento__lte=hoje + relativedelta(days=30),
        ).count()
    except Exception:
        boletos_nao_gerados = 0

    # 3. Boletos aguardando remessa CNAB
    try:
        boletos_sem_remessa = len(CNABService().obter_boletos_sem_remessa())
    except Exception:
        boletos_sem_remessa = 0

    # 4. Reajustes pendentes
    try:
        contratos_ativos = Contrato.objects.filter(
            status=StatusContrato.ATIVO
        ).exclude(tipo_correcao=TipoCorrecao.FIXO).only(
            'pk', 'data_contrato', 'numero_parcelas', 'prazo_reajuste_meses', 'tipo_correcao'
        ).prefetch_related(
            Prefetch(
                'reajustes',
                queryset=Reajuste.objects.filter(aplicado=True).only('contrato_id', 'ciclo'),
                to_attr='_reaj_cache',
            )
        )
        reajustes_pendentes = 0
        for contrato in contratos_ativos:
            if not contrato.data_contrato or not contrato.numero_parcelas:
                continue
            ciclos_aplicados = {r.ciclo for r in contrato._reaj_cache}
            prazo = contrato.prazo_reajuste_meses or 12
            total_ciclos = (contrato.numero_parcelas - 1) // prazo + 1
            for ciclo in range(2, total_ciclos + 1):
                data_reajuste = contrato.data_contrato + relativedelta(months=(ciclo - 1) * prazo)
                if hoje < data_reajuste:
                    break
                if ciclo not in ciclos_aplicados:
                    reajustes_pendentes += 1
                    break
    except Exception:
        reajustes_pendentes = 0

    return JsonResponse({
        'parcelas_vencidas': parcelas_vencidas,
        'boletos_nao_gerados': boletos_nao_gerados,
        'boletos_sem_remessa': boletos_sem_remessa,
        'reajustes_pendentes': reajustes_pendentes,
    })


# ==========================================================================
# FASE 9 — APIs P2
# ==========================================================================

def _serializar_parcela(p, hoje):
    """Serializa uma parcela para JSON (helper compartilhado)."""
    return {
        'id': p.id,
        'contrato': {
            'id': p.contrato.id,
            'numero': p.contrato.numero_contrato,
        },
        'comprador': p.contrato.comprador.nome,
        'imobiliaria': {
            'id': p.contrato.imobiliaria_id,
            'nome': p.contrato.imobiliaria.nome_fantasia or p.contrato.imobiliaria.razao_social,
        },
        'numero_parcela': p.numero_parcela,
        'total_parcelas': p.contrato.numero_parcelas,
        'data_vencimento': p.data_vencimento.strftime('%Y-%m-%d'),
        'valor_atual': float(p.valor_atual),
        'pago': p.pago,
        'data_pagamento': p.data_pagamento.strftime('%Y-%m-%d') if p.data_pagamento else None,
        'dias_atraso': max(0, (hoje - p.data_vencimento).days) if not p.pago and p.data_vencimento < hoje else 0,
        'status_boleto': p.status_boleto,
        'tem_boleto': p.tem_boleto,
        'linha_digitavel': p.linha_digitavel or '',
    }


def _filtrar_parcelas_periodo(qs, request, hoje):
    """Aplica filtros comuns de período e status."""
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')

    if status == 'pago':
        qs = qs.filter(pago=True)
    elif status == 'pendente':
        qs = qs.filter(pago=False, data_vencimento__gte=hoje)
    elif status == 'vencido':
        qs = qs.filter(pago=False, data_vencimento__lt=hoje)
    elif status == 'a_vencer':
        qs = qs.filter(pago=False, data_vencimento__gte=hoje)

    if data_inicio:
        try:
            from datetime import datetime as dt
            qs = qs.filter(data_vencimento__gte=dt.strptime(data_inicio, '%Y-%m-%d').date())
        except ValueError:
            pass  # Invalid date format — ignore filter
    if data_fim:
        try:
            from datetime import datetime as dt
            qs = qs.filter(data_vencimento__lte=dt.strptime(data_fim, '%Y-%m-%d').date())
        except ValueError:
            pass  # Invalid date format — ignore filter
    return qs


def _paginar(qs, request):
    """Retorna (items, total, page, per_page)."""
    try:
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(max(1, int(request.GET.get('per_page', 50))), 200)
    except (ValueError, TypeError):
        page, per_page = 1, 50
    total = qs.count()
    offset = (page - 1) * per_page
    return qs[offset:offset + per_page], total, page, per_page


# --------------------------------------------------------------------------
# 4-P2-1 : GET /financeiro/api/contabilidade/vencimentos/
# --------------------------------------------------------------------------

@login_required
@require_GET
def api_contabilidade_vencimentos(request):
    """
    Lista vencimentos consolidados para a Contabilidade.

    Filtros: imobiliaria, status, data_inicio, data_fim, comprador, page, per_page
    """
    hoje = timezone.now().date()

    qs = Parcela.objects.select_related(
        'contrato', 'contrato__comprador', 'contrato__imobiliaria'
    )

    imobiliaria_id = request.GET.get('imobiliaria')
    if imobiliaria_id:
        qs = qs.filter(contrato__imobiliaria_id=imobiliaria_id)

    comprador_id = request.GET.get('comprador')
    if comprador_id:
        qs = qs.filter(contrato__comprador_id=comprador_id)

    qs = _filtrar_parcelas_periodo(qs, request, hoje)
    qs = qs.order_by('data_vencimento', 'contrato_id', 'numero_parcela')

    totais = qs.aggregate(
        total_valor=Sum('valor_atual'),
        total_pago=Sum('valor_pago', filter=Q(pago=True)),
        quantidade=Count('id'),
        vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
    )

    page_qs, total, page, per_page = _paginar(qs, request)
    parcelas = [_serializar_parcela(p, hoje) for p in page_qs]

    return JsonResponse({
        'sucesso': True,
        'parcelas': parcelas,
        'total': total,
        'page': page,
        'per_page': per_page,
        'totais': {
            'quantidade': totais['quantidade'] or 0,
            'valor_total': float(totais['total_valor'] or 0),
            'valor_pago': float(totais['total_pago'] or 0),
            'quantidade_vencidas': totais['vencidas'] or 0,
        },
    })


# --------------------------------------------------------------------------
# 4-P2-2 : POST /financeiro/api/contabilidade/boletos/gerar/massa/
# --------------------------------------------------------------------------
# Re-usa api_gerar_boletos_lote (registrado como alias na urls.py)

api_contabilidade_boletos_massa = api_gerar_boletos_lote


# --------------------------------------------------------------------------
# 4-P2-3 : GET /financeiro/api/imobiliaria/<id>/vencimentos/
# --------------------------------------------------------------------------

@login_required
@require_GET
def api_imobiliaria_vencimentos(request, imobiliaria_id):
    """
    Lista vencimentos de uma imobiliária específica.

    Filtros: status, data_inicio, data_fim, comprador, page, per_page
    """
    imobiliaria = get_object_or_404(Imobiliaria, pk=imobiliaria_id)
    hoje = timezone.now().date()

    qs = Parcela.objects.select_related(
        'contrato', 'contrato__comprador', 'contrato__imobiliaria'
    ).filter(contrato__imobiliaria=imobiliaria)

    comprador_id = request.GET.get('comprador')
    if comprador_id:
        qs = qs.filter(contrato__comprador_id=comprador_id)

    qs = _filtrar_parcelas_periodo(qs, request, hoje)
    qs = qs.order_by('data_vencimento', 'numero_parcela')

    totais = qs.aggregate(
        total_valor=Sum('valor_atual'),
        total_pago=Sum('valor_pago', filter=Q(pago=True)),
        quantidade=Count('id'),
        vencidas=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
    )

    page_qs, total, page, per_page = _paginar(qs, request)
    parcelas = [_serializar_parcela(p, hoje) for p in page_qs]

    return JsonResponse({
        'sucesso': True,
        'imobiliaria': {
            'id': imobiliaria.id,
            'nome': imobiliaria.nome_fantasia or imobiliaria.razao_social,
        },
        'parcelas': parcelas,
        'total': total,
        'page': page,
        'per_page': per_page,
        'totais': {
            'quantidade': totais['quantidade'] or 0,
            'valor_total': float(totais['total_valor'] or 0),
            'valor_pago': float(totais['total_pago'] or 0),
            'quantidade_vencidas': totais['vencidas'] or 0,
        },
    })


# --------------------------------------------------------------------------
# 4-P2-4 : GET /financeiro/api/imobiliaria/<id>/fluxo-caixa/
# --------------------------------------------------------------------------

@login_required
@require_GET
def api_imobiliaria_fluxo_caixa(request, imobiliaria_id):
    """
    Fluxo de caixa mensal previsto vs realizado para uma imobiliária.

    Retorna os últimos 6 meses + próximos 6 meses.
    """
    imobiliaria = get_object_or_404(Imobiliaria, pk=imobiliaria_id)
    hoje = timezone.now().date()
    meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

    base_qs = Parcela.objects.filter(contrato__imobiliaria=imobiliaria)

    meses = []
    for i in range(-5, 7):  # -5 (5 meses atrás) até +6 (6 meses à frente)
        ref = hoje + relativedelta(months=i)
        primeiro = ref.replace(day=1)
        ultimo = (primeiro + relativedelta(months=1)) - timedelta(days=1)

        qs_mes = base_qs.filter(data_vencimento__gte=primeiro, data_vencimento__lte=ultimo)
        agg = qs_mes.aggregate(
            esperado=Sum('valor_atual'),
            recebido=Sum('valor_pago', filter=Q(pago=True)),
            qtd_total=Count('id'),
            qtd_pago=Count('id', filter=Q(pago=True)),
            qtd_vencido=Count('id', filter=Q(pago=False, data_vencimento__lt=hoje)),
        )

        meses.append({
            'mes': f"{meses_nomes[ref.month - 1]}/{ref.year}",
            'ano_mes': ref.strftime('%Y-%m'),
            'passado': i < 0,
            'corrente': i == 0,
            'futuro': i > 0,
            'esperado': float(agg['esperado'] or 0),
            'recebido': float(agg['recebido'] or 0),
            'pendente': float((agg['esperado'] or 0) - (agg['recebido'] or 0)),
            'qtd_total': agg['qtd_total'] or 0,
            'qtd_pago': agg['qtd_pago'] or 0,
            'qtd_vencido': agg['qtd_vencido'] or 0,
        })

    return JsonResponse({
        'sucesso': True,
        'imobiliaria': {
            'id': imobiliaria.id,
            'nome': imobiliaria.nome_fantasia or imobiliaria.razao_social,
        },
        'meses': meses,
    })


# =============================================================================
# OFX — Importação de Extrato Bancário para Quitação de Parcelas
# =============================================================================

@login_required
def upload_ofx(request):
    """
    GET  → página de upload do extrato OFX
    POST → processa o arquivo OFX e retorna JSON com resultado da reconciliação
    """
    from financeiro.services.ofx_service import processar_ofx_upload

    if request.method == 'GET':
        # Listar imobiliárias disponíveis para o filtro
        from core.models import Imobiliaria
        imobiliarias = Imobiliaria.objects.filter(ativo=True).order_by('razao_social')
        return render(request, 'financeiro/ofx_upload.html', {
            'imobiliarias': imobiliarias,
        })

    # POST — processa arquivo
    arquivo = request.FILES.get('arquivo_ofx')
    if not arquivo:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhum arquivo enviado'}, status=400)

    if not arquivo.name.lower().endswith('.ofx'):
        return JsonResponse({'sucesso': False, 'erro': 'Arquivo deve ter extensão .ofx'}, status=400)

    # Limite de tamanho: 5 MB
    if arquivo.size > 5 * 1024 * 1024:
        return JsonResponse({'sucesso': False, 'erro': 'Arquivo muito grande (máximo 5 MB)'}, status=400)

    imobiliaria_id = request.POST.get('imobiliaria_id')
    imobiliaria = None
    if imobiliaria_id:
        from core.models import Imobiliaria
        imobiliaria = get_object_or_404(Imobiliaria, pk=imobiliaria_id)

    dry_run = request.POST.get('dry_run', '').lower() in ('1', 'true', 'yes')

    try:
        conteudo = arquivo.read()
        resultado = processar_ofx_upload(
            conteudo,
            imobiliaria=imobiliaria,
            dry_run=dry_run,
        )
    except Exception as e:
        logger.exception('OFX upload: erro ao processar arquivo')
        return JsonResponse({'sucesso': False, 'erro': f'Erro ao processar arquivo: {e}'}, status=500)

    if dry_run:
        return JsonResponse({'sucesso': True, **resultado})

    # Serializar resultados para JSON
    resultados_json = []
    for rec in resultado.get('resultados', []):
        tx = rec.transacao
        item = {
            'fitid': tx.fitid,
            'data': str(tx.data) if tx.data else None,
            'valor': str(tx.valor),
            'memo': tx.memo,
            'reconciliada': rec.reconciliada,
            'confianca': rec.confianca,
            'motivo': rec.motivo,
        }
        if rec.parcela:
            item['parcela'] = {
                'id': rec.parcela.pk,
                'numero_parcela': rec.parcela.numero_parcela,
                'nosso_numero': rec.parcela.nosso_numero,
                'contrato_id': rec.parcela.contrato_id,
                'numero_contrato': rec.parcela.contrato.numero_contrato,
            }
        resultados_json.append(item)

    return JsonResponse({
        'sucesso': True,
        'total_transacoes': resultado['total_transacoes'],
        'reconciliadas': resultado['reconciliadas'],
        'nao_reconciliadas': resultado['nao_reconciliadas'],
        'parcelas_quitadas': [p.pk for p in resultado.get('parcelas_quitadas', [])],
        'resultados': resultados_json,
        'parser': resultado.get('parser', 'python'),
    })


# =============================================================================
# WhatsApp / SMS — Envio de Boleto via Twilio
# =============================================================================

@login_required
@require_POST
def enviar_boleto_whatsapp(request, pk):
    """
    POST /financeiro/parcelas/<pk>/boleto/whatsapp/

    Envia dados do boleto (linha digitável + vencimento) por WhatsApp via Twilio.
    Campos opcionais no body (JSON ou form):
      - telefone: destinatário (ex: +5511999999999). Se omitido, usa comprador.telefone.
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    import json as _json

    if not parcela.tem_boleto:
        return JsonResponse({'sucesso': False, 'erro': 'Parcela não possui boleto gerado'}, status=400)

    # Determinar destinatário
    telefone = None
    try:
        body = _json.loads(request.body) if request.body else {}
        telefone = body.get('telefone')
    except (_json.JSONDecodeError, ValueError):
        telefone = request.POST.get('telefone')

    if not telefone:
        try:
            telefone = parcela.contrato.comprador.telefone
        except AttributeError:
            pass

    if not telefone:
        return JsonResponse({'sucesso': False, 'erro': 'Telefone não informado e comprador sem telefone cadastrado'}, status=400)

    # Montar mensagem
    vencimento = parcela.data_vencimento.strftime('%d/%m/%Y') if parcela.data_vencimento else '—'
    valor = f"R$ {parcela.valor_atual:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    linha = parcela.linha_digitavel or '(linha digitável não disponível)'
    nome = getattr(parcela.contrato.comprador, 'nome', 'Cliente')
    mensagem = (
        f"Ola, {nome}!\n"
        f"Seu boleto esta disponivel:\n\n"
        f"Parcela: {parcela.numero_parcela}\n"
        f"Valor: {valor}\n"
        f"Vencimento: {vencimento}\n\n"
        f"Linha Digitavel:\n{linha}\n\n"
        f"Em caso de duvidas, entre em contato conosco."
    )

    try:
        from notificacoes.services import ServicoWhatsApp
        ServicoWhatsApp.enviar(telefone, mensagem)
        logger.info('WhatsApp boleto enviado: parcela pk=%s → %s', pk, telefone)
        return JsonResponse({'sucesso': True, 'destinatario': telefone})
    except Exception as e:
        logger.exception('WhatsApp boleto erro: parcela pk=%s → %s', pk, e)
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


@login_required
@require_POST
def enviar_boleto_sms(request, pk):
    """
    POST /financeiro/parcelas/<pk>/boleto/sms/

    Envia dados do boleto (linha digitável + vencimento) por SMS via Twilio.
    Campos opcionais no body (JSON ou form):
      - telefone: destinatário (ex: +5511999999999). Se omitido, usa comprador.telefone.
    """
    parcela = get_object_or_404(Parcela, pk=pk)

    import json as _json

    if not parcela.tem_boleto:
        return JsonResponse({'sucesso': False, 'erro': 'Parcela não possui boleto gerado'}, status=400)

    # Determinar destinatário
    telefone = None
    try:
        body = _json.loads(request.body) if request.body else {}
        telefone = body.get('telefone')
    except (_json.JSONDecodeError, ValueError):
        telefone = request.POST.get('telefone')

    if not telefone:
        try:
            telefone = parcela.contrato.comprador.telefone
        except AttributeError:
            pass

    if not telefone:
        return JsonResponse({'sucesso': False, 'erro': 'Telefone não informado e comprador sem telefone cadastrado'}, status=400)

    # Montar mensagem SMS (160 chars idealmente)
    vencimento = parcela.data_vencimento.strftime('%d/%m/%Y') if parcela.data_vencimento else '—'
    valor = f"R${parcela.valor_atual:.2f}"
    linha = parcela.linha_digitavel or 'indisponível'
    mensagem = (
        f"Boleto parcela {parcela.numero_parcela} - {valor} - venc {vencimento}. "
        f"Linha: {linha}"
    )[:160]

    try:
        from notificacoes.services import ServicoSMS
        ServicoSMS.enviar(telefone, mensagem)
        logger.info('SMS boleto enviado: parcela pk=%s → %s', pk, telefone)
        return JsonResponse({'sucesso': True, 'destinatario': telefone})
    except Exception as e:
        logger.exception('SMS boleto erro: parcela pk=%s → %s', pk, e)
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=500)


# =============================================================================
# SEÇÃO 18 — SIMULADOR DE ANTECIPAÇÃO / RENEGOCIAÇÃO
# =============================================================================

@login_required
def simulador_antecipacao(request, contrato_id):
    """
    R-01: Tela simulador de antecipação de parcelas com desconto.
    R-02: Preview do valor original vs. antecipado (economia).
    R-03: Aplicar antecipação — cria HistoricoPagamento com antecipado=True.

    GET  → formulário: lista parcelas NORMAL não pagas + campo % desconto
    POST action=preview → tabela preview sem persistir
    POST action=aplicar → quita as parcelas selecionadas com desconto
    """
    from django.db import transaction
    from contratos.models import Contrato
    from .models import HistoricoPagamento

    contrato = get_object_or_404(Contrato, pk=contrato_id)

    parcelas_disponiveis = (
        Parcela.objects.filter(
            contrato=contrato,
            pago=False,
            tipo_parcela='NORMAL',
        )
        .order_by('numero_parcela')
    )

    preview = None
    parcelas_selecionadas_ids = []
    desconto_perc = Decimal('0')

    if request.method == 'POST':
        action = request.POST.get('action', 'preview')
        parcelas_selecionadas_ids = request.POST.getlist('parcelas')
        desconto_str = request.POST.get('desconto', '0').replace(',', '.')
        try:
            desconto_perc = Decimal(desconto_str)
        except Exception:
            desconto_perc = Decimal('0')
        desconto_perc = max(Decimal('0'), min(Decimal('100'), desconto_perc))

        parcelas_sel = list(
            parcelas_disponiveis.filter(id__in=parcelas_selecionadas_ids).order_by('numero_parcela')
        )

        preview_itens = []
        total_original = Decimal('0')
        total_antecipado = Decimal('0')

        for p in parcelas_sel:
            valor_original = p.valor_atual
            desconto_valor = (valor_original * desconto_perc / 100).quantize(Decimal('0.01'))
            valor_antecipado = valor_original - desconto_valor
            preview_itens.append({
                'parcela': p,
                'valor_original': valor_original,
                'desconto_valor': desconto_valor,
                'valor_antecipado': valor_antecipado,
            })
            total_original += valor_original
            total_antecipado += valor_antecipado

        preview = {
            'itens': preview_itens,
            'total_original': total_original,
            'total_antecipado': total_antecipado,
            'economia': total_original - total_antecipado,
            'desconto_perc': desconto_perc,
            'qtd': len(preview_itens),
        }

        if action == 'aplicar' and preview_itens:
            data_pagamento = timezone.now().date()
            obs = f'Antecipação com {desconto_perc}% de desconto'
            with transaction.atomic():
                for item in preview_itens:
                    p = item['parcela']
                    p.pago = True
                    p.data_pagamento = data_pagamento
                    p.valor_pago = item['valor_antecipado']
                    p.valor_desconto = item['desconto_valor']
                    if p.tem_boleto:
                        p.status_boleto = StatusBoleto.PAGO
                        p.data_pagamento_boleto = timezone.now()
                        p.valor_pago_boleto = item['valor_antecipado']
                    p.save()

                    HistoricoPagamento.objects.create(
                        parcela=p,
                        data_pagamento=data_pagamento,
                        valor_pago=item['valor_antecipado'],
                        valor_parcela=item['valor_original'],
                        valor_desconto=item['desconto_valor'],
                        forma_pagamento='DINHEIRO',
                        antecipado=True,
                        observacoes=obs,
                        origem_pagamento='ANTECIPACAO',
                    )

            messages.success(
                request,
                f'{len(preview_itens)} parcela(s) antecipada(s) com sucesso. '
                f'Economia total: R$ {(total_original - total_antecipado):,.2f}.'
            )
            return redirect('contratos:detalhe', pk=contrato_id)

    return render(request, 'financeiro/simulador_antecipacao.html', {
        'contrato': contrato,
        'parcelas_disponiveis': parcelas_disponiveis,
        'preview': preview,
        'parcelas_selecionadas_ids': [str(x) for x in parcelas_selecionadas_ids],
        'desconto_perc': desconto_perc,
    })


@login_required
def download_recibo_antecipacao(request, contrato_id):
    """
    R-05: Gera e baixa o recibo PDF de quitação antecipada.

    GET ?historico_ids=1,2,3  → PDF com os IDs informados
    GET (sem params)           → PDF com as antecipações mais recentes do contrato
    """
    from contratos.models import Contrato
    from .models import HistoricoPagamento
    from .services.recibo_service import gerar_recibo_antecipacao_pdf

    contrato = get_object_or_404(Contrato, pk=contrato_id)

    ids_param = request.GET.get('historico_ids', '')
    if ids_param:
        try:
            ids = [int(x) for x in ids_param.split(',') if x.strip()]
        except ValueError:
            ids = []
    else:
        ids = []

    if ids:
        historicos = HistoricoPagamento.objects.filter(
            pk__in=ids,
            parcela__contrato=contrato,
            antecipado=True,
        ).select_related('parcela').order_by('parcela__numero_parcela')
    else:
        historicos = HistoricoPagamento.objects.filter(
            parcela__contrato=contrato,
            antecipado=True,
        ).select_related('parcela').order_by('-data_pagamento', 'parcela__numero_parcela')[:20]

    if not historicos.exists():
        messages.error(request, 'Nenhuma antecipação encontrada para este contrato.')
        return redirect('contratos:detalhe', pk=contrato_id)

    try:
        pdf_bytes = gerar_recibo_antecipacao_pdf(contrato, historicos)
    except Exception as e:
        messages.error(request, f'Erro ao gerar recibo: {e}')
        return redirect('contratos:detalhe', pk=contrato_id)

    filename = f'recibo_antecipacao_{contrato.numero_contrato}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# R-04: Renegociação de Parcelas em Atraso
# =============================================================================

@login_required
def renegociar_parcelas(request, contrato_id):
    """
    R-04: Tela de renegociação — permite alterar data de vencimento e/ou valor
    de parcelas em atraso (não pagas e vencidas). Zera juros/multa acumulados
    ao confirmar a renegociação.

    GET  → exibe lista de parcelas elegíveis com campos editáveis
    POST → aplica as alterações às parcelas selecionadas
    """
    from contratos.models import Contrato
    from datetime import date

    contrato = get_object_or_404(Contrato, pk=contrato_id)
    hoje = date.today()

    # Parcelas elegíveis: não pagas e vencidas
    parcelas_em_atraso = Parcela.objects.filter(
        contrato=contrato,
        pago=False,
        data_vencimento__lt=hoje,
    ).order_by('numero_parcela')

    if request.method == 'POST':
        parcelas_ids = request.POST.getlist('parcela_ids')
        if not parcelas_ids:
            messages.warning(request, 'Selecione ao menos uma parcela para renegociar.')
            return redirect('financeiro:renegociar_parcelas', contrato_id=contrato_id)

        alteradas = 0
        erros = []

        with transaction.atomic():
            for pk in parcelas_ids:
                try:
                    parcela = Parcela.objects.select_for_update().get(
                        pk=pk, contrato=contrato, pago=False
                    )
                except Parcela.DoesNotExist:
                    continue

                nova_data_str = request.POST.get(f'nova_data_{pk}', '').strip()
                novo_valor_str = request.POST.get(f'novo_valor_{pk}', '').strip()

                mudou = False

                if nova_data_str:
                    try:
                        from datetime import datetime
                        nova_data = datetime.strptime(nova_data_str, '%Y-%m-%d').date()
                        if nova_data <= hoje:
                            erros.append(
                                f'Parcela {parcela.numero_parcela}: nova data deve ser futura.'
                            )
                            continue
                        parcela.data_vencimento = nova_data
                        mudou = True
                    except ValueError:
                        erros.append(f'Parcela {parcela.numero_parcela}: data inválida.')
                        continue

                if novo_valor_str:
                    try:
                        novo_valor = Decimal(novo_valor_str.replace(',', '.'))
                        if novo_valor <= 0:
                            erros.append(
                                f'Parcela {parcela.numero_parcela}: valor deve ser positivo.'
                            )
                            continue
                        parcela.valor_atual = novo_valor
                        mudou = True
                    except Exception:
                        erros.append(f'Parcela {parcela.numero_parcela}: valor inválido.')
                        continue

                if mudou:
                    # Zera juros/multa acumulados — novo prazo começa limpo
                    parcela.valor_juros = Decimal('0.00')
                    parcela.valor_multa = Decimal('0.00')
                    parcela.save(update_fields=[
                        'data_vencimento', 'valor_atual', 'valor_juros', 'valor_multa'
                    ])
                    alteradas += 1

        for erro in erros:
            messages.warning(request, erro)

        if alteradas:
            messages.success(
                request,
                f'{alteradas} parcela(s) renegociada(s) com sucesso. '
                f'Juros e multa foram zerados.'
            )
        return redirect('contratos:detalhe', pk=contrato_id)

    # GET — renderiza formulário
    return render(request, 'financeiro/renegociar_parcelas.html', {
        'contrato': contrato,
        'parcelas': parcelas_em_atraso,
        'hoje': hoje,
    })


# =============================================================================
# SECTION 4 P3 — APIs Pendentes
# =============================================================================

@login_required
@require_GET
def api_contabilidade_relatorios_vencimentos(request):
    """
    4-P3-1 : GET /financeiro/api/contabilidade/relatorios/vencimentos/

    Retorna relatório de vencimentos agrupado por período
    (semanal, mensal ou trimestral).

    Parâmetros GET:
      periodo  : semanal | mensal | trimestral (default: mensal)
      meses    : quantidade de meses a projetar (default: 3)
      imobiliaria: ID da imobiliária (opcional)
    """
    from dateutil.relativedelta import relativedelta

    hoje = timezone.now().date()
    periodo = request.GET.get('periodo', 'mensal')
    imobiliaria_id = request.GET.get('imobiliaria')

    try:
        meses = max(1, min(int(request.GET.get('meses', 3)), 12))
    except (ValueError, TypeError):
        meses = 3

    qs = Parcela.objects.filter(pago=False).select_related(
        'contrato', 'contrato__imobiliaria'
    )
    if imobiliaria_id:
        qs = qs.filter(contrato__imobiliaria_id=imobiliaria_id)

    grupos = []

    if periodo == 'semanal':
        # Próximas 4 semanas
        for i in range(4):
            inicio = hoje + timedelta(weeks=i)
            fim = inicio + timedelta(days=6)
            parcelas_periodo = qs.filter(
                data_vencimento__gte=inicio,
                data_vencimento__lte=fim,
            )
            total = parcelas_periodo.aggregate(
                valor=Sum('valor_atual'), qtd=Count('id')
            )
            grupos.append({
                'periodo': f'Semana {i + 1} ({inicio.strftime("%d/%m")}–{fim.strftime("%d/%m")})',
                'data_inicio': inicio.isoformat(),
                'data_fim': fim.isoformat(),
                'quantidade': total['qtd'] or 0,
                'valor': float(total['valor'] or 0),
            })
    elif periodo == 'trimestral':
        # Próximos trimestres
        for i in range(2):
            inicio = hoje + relativedelta(months=i * 3)
            fim = inicio + relativedelta(months=3) - timedelta(days=1)
            parcelas_periodo = qs.filter(
                data_vencimento__gte=inicio,
                data_vencimento__lte=fim,
            )
            total = parcelas_periodo.aggregate(
                valor=Sum('valor_atual'), qtd=Count('id')
            )
            grupos.append({
                'periodo': f'T{i + 1} ({inicio.strftime("%m/%Y")}–{fim.strftime("%m/%Y")})',
                'data_inicio': inicio.isoformat(),
                'data_fim': fim.isoformat(),
                'quantidade': total['qtd'] or 0,
                'valor': float(total['valor'] or 0),
            })
    else:
        # Mensal (default)
        for i in range(meses):
            inicio = hoje.replace(day=1) + relativedelta(months=i)
            fim = inicio + relativedelta(months=1) - timedelta(days=1)
            parcelas_periodo = qs.filter(
                data_vencimento__gte=inicio,
                data_vencimento__lte=fim,
            )
            total = parcelas_periodo.aggregate(
                valor=Sum('valor_atual'), qtd=Count('id')
            )
            grupos.append({
                'periodo': inicio.strftime('%B/%Y'),
                'data_inicio': inicio.isoformat(),
                'data_fim': fim.isoformat(),
                'quantidade': total['qtd'] or 0,
                'valor': float(total['valor'] or 0),
            })

    vencidas = qs.filter(data_vencimento__lt=hoje).aggregate(
        valor=Sum('valor_atual'), qtd=Count('id')
    )

    return JsonResponse({
        'sucesso': True,
        'periodo': periodo,
        'grupos': grupos,
        'vencidas': {
            'quantidade': vencidas['qtd'] or 0,
            'valor': float(vencidas['valor'] or 0),
        },
    })


@login_required
@require_GET
def api_imobiliaria_pendencias(request, imobiliaria_id):
    """
    4-P3-3 : GET /financeiro/api/imobiliaria/<id>/pendencias/

    Retorna parcelas vencidas com encargos calculados para uma imobiliária.
    """
    imobiliaria = get_object_or_404(Imobiliaria, pk=imobiliaria_id)
    hoje = timezone.now().date()

    qs = Parcela.objects.filter(
        contrato__imobiliaria=imobiliaria,
        pago=False,
        data_vencimento__lt=hoje,
    ).select_related(
        'contrato', 'contrato__comprador'
    ).order_by('data_vencimento')

    pendencias = []
    for p in qs[:200]:  # limita a 200 registros
        dias_atraso = (hoje - p.data_vencimento).days

        # Calcula encargos se o contrato tiver configuração
        try:
            valor_juros = p.calcular_juros_mora()
            valor_multa = p.calcular_multa()
        except Exception:
            valor_juros = p.valor_juros or Decimal('0.00')
            valor_multa = p.valor_multa or Decimal('0.00')

        total = p.valor_atual + valor_juros + valor_multa

        pendencias.append({
            'parcela_id': p.pk,
            'contrato': p.contrato.numero_contrato,
            'comprador': p.contrato.comprador.nome,
            'numero_parcela': p.numero_parcela,
            'data_vencimento': p.data_vencimento.isoformat(),
            'dias_atraso': dias_atraso,
            'valor_original': float(p.valor_atual),
            'valor_juros': float(valor_juros),
            'valor_multa': float(valor_multa),
            'valor_total': float(total),
            'nosso_numero': p.nosso_numero or '',
        })

    totais = {
        'quantidade': len(pendencias),
        'valor_total': sum(p['valor_total'] for p in pendencias),
        'valor_original': sum(p['valor_original'] for p in pendencias),
    }

    return JsonResponse({
        'sucesso': True,
        'imobiliaria': {'id': imobiliaria.pk, 'nome': imobiliaria.nome},
        'pendencias': pendencias,
        'totais': totais,
    })
